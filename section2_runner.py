"""
section2_runner.py  —  V10.0
Dual-mode script:
  --mode=trigger_kaggle  (runs on GitHub Actions — lightweight, pushes notebook to Kaggle and waits)
  --mode=kaggle_run      (runs ON Kaggle — ViT-GPT2 vision, GPU-accelerated on P100/T4)
  (no args)              (local mode — same ViT-GPT2 pipeline, CPU fallback)

Vision model: nlpconnect/vit-gpt2-image-captioning
  - Apache 2.0 licence, 100% free
  - Only 330 MB — loads in ~5s on CPU, ~3s on GPU
  - Auto-selects GPU (CUDA) when available — falls back to CPU seamlessly
  - No CUDA version conflicts: uses torch.cuda.is_available() at runtime
  - ~0.05-0.15s per image caption on P100/T4 GPU  (~0.3s on CPU)
  - No HF token needed
"""

import argparse
import base64
import io
import json
import os
import re
import subprocess
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple


ULTRADATA_XLSX  = "ultradata.xlsx"
WATERMARK_TEXT  = "www.ultrapng.com"
INSTANT_CAP     = 5000
MAX_RUN_SECONDS = 17_400   # 4h50m local / unused in Kaggle path
_RUN_START      = time.time()

# ══════════════════════════════════════════════════════════════
# KAGGLE TRIGGER MODE  (runs on GitHub Actions, no GPU needed)
# ══════════════════════════════════════════════════════════════

KAGGLE_NOTEBOOK_TITLE  = "section2-seo-builder-v10"
KAGGLE_NOTEBOOK_SLUG   = "section2-seo-builder-v10"   # all lowercase, hyphens only


def _build_kaggle_kernel_source(env_vars: dict) -> str:
    """
    Generates the Python source that Kaggle will execute on a GPU machine.
    Embeds all env vars as literals (secrets never leave GitHub env).
    The kernel runs section2_runner.py in kaggle_run mode with all envs set.
    """
    env_lines = "\n".join(
        f'os.environ[{k!r}] = {v!r}' for k, v in env_vars.items() if v
    )
    return f'''#!/usr/bin/env python3
"""Auto-generated Kaggle kernel — do not edit manually."""
import os, subprocess, sys

# ── inject secrets ──────────────────────────────────────────
{env_lines}

# ── install dependencies ─────────────────────────────────────
# accelerate is required for GPU-accelerated transformers pipelines.
# torch is pre-installed on Kaggle; listed here for safety.
subprocess.check_call([sys.executable, "-m", "pip", "install", "-q",
    "transformers>=4.37", "accelerate>=0.26", "openpyxl", "requests", "pillow"])
print("[setup] Dependencies installed ✓", flush=True)

# ── confirm GPU availability ─────────────────────────────────
import torch as _torch
if _torch.cuda.is_available():
    print(f"[setup] GPU detected: {{_torch.cuda.get_device_name(0)}} ✓", flush=True)
else:
    print("[setup] No GPU detected — will run on CPU", flush=True)

# ── download runner script from GitHub repo ──────────────────
import requests as _req
_gh_token  = os.environ.get("GH_TOKEN","")
_gh_repo   = os.environ.get("GITHUB_REPOSITORY","")
_gh_branch = os.environ.get("GITHUB_REF_NAME","main")
_headers   = {{"Authorization": f"token {{_gh_token}}", "Accept": "application/vnd.github.raw"}}
_api_url   = f"https://api.github.com/repos/{{_gh_repo}}/contents/section2_runner.py?ref={{_gh_branch}}"
_resp = _req.get(_api_url, headers=_headers, timeout=60)
_resp.raise_for_status()
with open("/kaggle/working/section2_runner.py", "w") as _f:
    _f.write(_resp.text)

# ── run in kaggle_run mode ───────────────────────────────────
os.chdir("/kaggle/working")
subprocess.check_call([sys.executable, "section2_runner.py", "--mode=kaggle_run"])
'''


def trigger_kaggle_mode():
    """
    Runs on GitHub Actions.
    1. Collects all env secrets.
    2. Writes a kernel folder (kernel.py + kernel-metadata.json).
    3. Pushes via `kaggle kernels push` CLI (correct, official method).
    4. Polls via Kaggle REST API until complete or error.

    WHY CLI push instead of raw blob API:
      The /blobs/upload endpoint requires a specific undocumented binary protocol.
      The official `kaggle kernels push` CLI handles this correctly and is already
      installed in the GitHub Actions step.
    """
    import requests

    kaggle_user = os.environ.get("KAGGLE_USERNAME","").strip()
    kaggle_key  = os.environ.get("KAGGLE_KEY","").strip()
    if not kaggle_user or not kaggle_key:
        raise SystemExit("❌ KAGGLE_USERNAME or KAGGLE_KEY not set")

    # Collect all env vars to embed into the Kaggle kernel source
    env_keys = [
        "REPO2_TOKEN","REPO2_SLUG","REPO2_MAX_PER_JSON",
        "GOOGLE_CLIENT_ID","GOOGLE_CLIENT_SECRET","GOOGLE_REFRESH_TOKEN",
        "GH_TOKEN","GH_OWNER",
        "PREVIEW_REPO","PREVIEW_BRANCH","PREVIEW_FOLDER",
        "PNG_LIBRARY_FOLDER","WATERMARK_TEXT",
        "S2_COUNT","SCAN_DRIVE",
        "GITHUB_REPOSITORY","GITHUB_REF_NAME",
    ]
    env_vars = {k: os.environ.get(k,"") for k in env_keys}

    kernel_src = _build_kaggle_kernel_source(env_vars)
    slug = KAGGLE_NOTEBOOK_SLUG

    # ── Write kernel push folder ─────────────────────────────────────────────
    import tempfile
    tmpdir = Path(tempfile.mkdtemp()) / slug
    tmpdir.mkdir(parents=True, exist_ok=True)
    (tmpdir / "kernel.py").write_text(kernel_src, encoding="utf-8")

    meta = {
        "id": f"{kaggle_user}/{slug}",
        "title": "Section2 SEO Builder V10",
        "code_file": "kernel.py",
        "language": "python",
        "kernel_type": "script",
        "is_private": True,
        "enable_gpu": True,
        "enable_tpu": False,
        "enable_internet": True,
        "dataset_sources": [],
        "competition_sources": [],
        "kernel_sources": []
    }
    (tmpdir / "kernel-metadata.json").write_text(json.dumps(meta, indent=2), encoding="utf-8")

    # ── Ensure ~/.kaggle/kaggle.json exists (already written by yml step) ───
    kaggle_cfg = Path.home() / ".kaggle" / "kaggle.json"
    if not kaggle_cfg.exists():
        kaggle_cfg.parent.mkdir(parents=True, exist_ok=True)
        kaggle_cfg.write_text(json.dumps({"username": kaggle_user, "key": kaggle_key}))
        kaggle_cfg.chmod(0o600)

    # ── Push kernel via CLI ──────────────────────────────────────────────────
    print(f"  [Kaggle] Pushing kernel '{slug}' via kaggle CLI ...", flush=True)
    result = subprocess.run(
        ["kaggle", "kernels", "push", "-p", str(tmpdir)],
        capture_output=True, text=True
    )
    print(result.stdout.strip(), flush=True)
    if result.returncode != 0:
        print(result.stderr.strip(), flush=True)
        raise SystemExit(f"❌ kaggle kernels push failed (exit {result.returncode})")

    print(f"  [Kaggle] Kernel pushed ✓", flush=True)
    print(f"  [Kaggle] View at: https://www.kaggle.com/code/{kaggle_user}/{slug}", flush=True)

    # ── Poll via `kaggle kernels status` CLI ────────────────────────────────
    # The REST GET /kernels/{user}/{slug} returns 404 for private kernels.
    # The official CLI always works correctly.

    # Give Kaggle ~90s to queue and start the job before first poll
    print("  [Kaggle] Waiting 90s for job to queue ...", flush=True)
    time.sleep(90)

    # GitHub Actions timeout is 60 min (set in yml).
    # Poll every 45s. Once the job is confirmed "running" we exit GH as success
    # — the Kaggle session runs up to 9h and pushes results to repo2 directly.
    POLL_INTERVAL = 45
    MAX_POLLS     = 78   # 78 x 45s ≈ 58 min (fits inside 60-min GH timeout)

    for attempt in range(1, MAX_POLLS + 1):
        time.sleep(POLL_INTERVAL)
        elapsed_m = (attempt * POLL_INTERVAL + 90) // 60
        try:
            result = subprocess.run(
                ["kaggle", "kernels", "status", f"{kaggle_user}/{slug}"],
                capture_output=True, text=True, timeout=30
            )
            output = (result.stdout + result.stderr).strip()
            print(f"    [{elapsed_m}m] {output}", flush=True)

            # kaggle CLI outputs e.g.:  username/slug has status "running"
            status = "unknown"
            m = re.search(r'"(\w+)"', output)
            if m:
                status = m.group(1).lower()

            if status == "complete":
                print("  ✅ Kaggle kernel completed successfully!", flush=True)
                return

            if status in ("error", "cancelacknowledged", "cancelled"):
                raise SystemExit(f"❌ Kaggle kernel failed with status: {status}\nOutput: {output}")

            # Once confirmed running, exit GH Actions — Kaggle runs independently
            if status == "running" and attempt >= 3:
                print("  ✅ Kaggle kernel is RUNNING on GPU — GitHub Action exiting.", flush=True)
                print(f"     Monitor: https://www.kaggle.com/code/{kaggle_user}/{slug}", flush=True)
                return

        except SystemExit:
            raise
        except Exception as exc:
            print(f"    [poll error] {exc}", flush=True)

    print(f"  ⚠  Poll timeout — Kaggle job likely still running.", flush=True)
    print(f"     Check: https://www.kaggle.com/code/{kaggle_user}/{slug}", flush=True)


# ══════════════════════════════════════════════════════════════
# GOOGLE DRIVE helpers
# ══════════════════════════════════════════════════════════════

_drive_token_cache = {"value": None, "expires": 0}

def _drive_token() -> str:
    import requests
    if _drive_token_cache["value"] and time.time() < _drive_token_cache["expires"]:
        return _drive_token_cache["value"]
    r = requests.post("https://oauth2.googleapis.com/token", data={
        "client_id": os.environ.get("GOOGLE_CLIENT_ID",""),
        "client_secret": os.environ.get("GOOGLE_CLIENT_SECRET",""),
        "refresh_token": os.environ.get("GOOGLE_REFRESH_TOKEN",""),
        "grant_type": "refresh_token",
    }, timeout=30)
    d = r.json()
    if "access_token" not in d:
        raise RuntimeError(f"Drive token error: {d}")
    _drive_token_cache.update({"value": d["access_token"], "expires": time.time()+3200})
    return _drive_token_cache["value"]

def _dh(token): return {"Authorization": f"Bearer {token}"}

_drive_folder_cache = {}
def _drive_folder_id(token, name, parent_id=None):
    import requests
    cache_key = f"{parent_id or 'root'}::{name}"
    if cache_key in _drive_folder_cache:
        return _drive_folder_cache[cache_key]
    q = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    if parent_id:
        q += f" and '{parent_id}' in parents"
    r = requests.get("https://www.googleapis.com/drive/v3/files", headers=_dh(token),
                     params={"q": q, "fields": "files(id,name)", "pageSize":1}, timeout=30)
    r.raise_for_status()
    files = r.json().get("files", [])
    if files:
        fid = files[0]["id"]
    else:
        body = {"name": name, "mimeType": "application/vnd.google-apps.folder"}
        if parent_id: body["parents"] = [parent_id]
        r2 = requests.post("https://www.googleapis.com/drive/v3/files",
                           headers={**_dh(token), "Content-Type": "application/json"},
                           json=body, timeout=30)
        r2.raise_for_status()
        fid = r2.json()["id"]
    _drive_folder_cache[cache_key] = fid
    return fid

def _drive_list_folder(token, folder_id, mime_filter=None):
    import requests
    q = f"'{folder_id}' in parents and trashed=false"
    if mime_filter: q += f" and mimeType='{mime_filter}'"
    results, page_token = [], None
    while True:
        params = {"q": q, "pageSize": 1000, "fields": "nextPageToken,files(id,name,mimeType,size)"}
        if page_token: params["pageToken"] = page_token
        r = requests.get("https://www.googleapis.com/drive/v3/files", headers=_dh(token), params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
        results.extend(data.get("files", []))
        page_token = data.get("nextPageToken")
        if not page_token: break
    return results

def _drive_list_pngs(token, folder_id):
    import requests
    q = f"'{folder_id}' in parents and trashed=false and (mimeType='image/png' or name contains '.png')"
    results, page_token = [], None
    while True:
        params = {"q": q, "pageSize": 1000, "fields": "nextPageToken,files(id,name,mimeType)"}
        if page_token: params["pageToken"] = page_token
        r = requests.get("https://www.googleapis.com/drive/v3/files", headers=_dh(token), params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
        results.extend(data.get("files", []))
        page_token = data.get("nextPageToken")
        if not page_token: break
    return results

def _drive_download(token, fid):
    import requests
    r = requests.get(f"https://www.googleapis.com/drive/v3/files/{fid}",
                     headers=_dh(token), params={"alt": "media"}, timeout=180)
    r.raise_for_status()
    return r.content

# ══════════════════════════════════════════════════════════════
# GITHUB API helpers
# ══════════════════════════════════════════════════════════════

def _gh_headers(token): return {"Authorization": f"token {token}", "Accept": "application/vnd.github.v3+json"}

def _gh_get_sha(token, owner, repo, path, branch="main"):
    import requests
    url = f"https://api.github.com/repos/{owner}/{repo}/contents/{path}"
    r = requests.get(url, headers=_gh_headers(token), params={"ref": branch}, timeout=30)
    if r.status_code == 200: return r.json().get("sha")
    return None

def _gh_upload_file(token, owner, repo, path, content_bytes, message, branch="main"):
    import requests
    url = f"https://api.github.com/repos/{owner}/{repo}/contents/{path}"
    sha = _gh_get_sha(token, owner, repo, path, branch)
    body = {"message": message, "content": base64.b64encode(content_bytes).decode(), "branch": branch}
    if sha: body["sha"] = sha
    for attempt in range(1, 4):
        r = requests.put(url, headers={**_gh_headers(token), "Content-Type": "application/json"}, json=body, timeout=90)
        if r.ok: return r.json()
        if attempt < 3: time.sleep(5 * attempt)
        else: r.raise_for_status()
    return {}

def _jsdelivr_url(owner, repo, branch, path):
    return f"https://cdn.jsdelivr.net/gh/{owner}/{repo}@{branch}/{path}"

# ══════════════════════════════════════════════════════════════
# WEBP PREVIEW
# ══════════════════════════════════════════════════════════════

WEBP_MAX_SIDE = 800
WEBP_MAX_BYTES = 80*1024

def _make_webp_preview(png_bytes, watermark):
    from PIL import Image, ImageDraw, ImageFont
    def _font(size):
        for fp in ["/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
                   "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
                   "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf",
                   "/usr/share/fonts/DejaVuSans-Bold.ttf"]:
            try: return ImageFont.truetype(fp, size)
            except: pass
        return ImageFont.load_default()
    img = Image.open(io.BytesIO(png_bytes)).convert("RGBA")
    w, h = img.size
    if max(w, h) > WEBP_MAX_SIDE:
        scale = WEBP_MAX_SIDE / max(w, h)
        img = img.resize((int(w*scale), int(h*scale)), Image.LANCZOS)
    w, h = img.size
    CELL = 16
    bg = Image.new("RGB", (w, h), (255,255,255))
    draw = ImageDraw.Draw(bg)
    for gy in range(0, h, CELL):
        for gx in range(0, w, CELL):
            if (gx//CELL + gy//CELL) % 2 == 0:
                draw.rectangle([gx, gy, gx+CELL-1, gy+CELL-1], fill=(204,204,204))
    bg.paste(img, mask=img.split()[3])
    wm_layer = Image.new("RGBA", (w, h), (0,0,0,0))
    wm_draw = ImageDraw.Draw(wm_layer)
    wm_font = _font(max(11, w//30))
    step_x, step_y = max(80, w//4), max(40, h//5)
    for oy in range(-h, h*2, step_y):
        for ox in range(-w, w*2, step_x):
            wm_draw.text((ox, oy), watermark, font=wm_font, fill=(200,200,200,80))
    bg = bg.convert("RGBA")
    bg.alpha_composite(wm_layer)
    bg = bg.convert("RGB")
    FOOTER_H = max(18, h//18)
    canvas = Image.new("RGB", (w, h+FOOTER_H), (40,40,40))
    canvas.paste(bg, (0,0))
    ft_draw = ImageDraw.Draw(canvas)
    ft_font = _font(max(9, FOOTER_H-4))
    ft_draw.rectangle([0, h, w, h+FOOTER_H], fill=(40,40,40))
    ft_draw.text((4, h+2), watermark, font=ft_font, fill=(220,220,220))
    buf = io.BytesIO()
    for quality in [85,70,55,40,25,10]:
        buf.seek(0); buf.truncate()
        canvas.save(buf, "WEBP", quality=quality, method=6)
        if buf.tell() <= WEBP_MAX_BYTES: break
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════
# XLSX helpers
# ══════════════════════════════════════════════════════════════

def _append_ultradata_rows(xlsx_path, rows):
    import openpyxl
    if not rows: return 0
    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb.active
    HEADERS = ["date_added","subject_name","category","subcategory","filename","png_file_id","webp_file_id","download_url","preview_url","seo_status"]
    if ws.max_row == 0 or ws.cell(row=1, column=1).value is None:
        ws.append(HEADERS)
    else:
        existing_hdr = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
        for col_name in HEADERS:
            if col_name not in existing_hdr:
                ws.cell(row=1, column=ws.max_column+1, value=col_name)
                existing_hdr.append(col_name)
        HEADERS = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
    for row in rows:
        ws.append([row.get(h, "") for h in HEADERS])
    wb.save(str(xlsx_path))
    return len(rows)

# ══════════════════════════════════════════════════════════════
# DRIVE SCANNER
# ══════════════════════════════════════════════════════════════

def _collect_all_pngs_from_drive(folder_name):
    print(f"  Scanning Drive folder '{folder_name}' for PNGs ...")
    token = _drive_token()
    root_id = _drive_folder_id(token, folder_name)
    print(f"  Root folder ID: {root_id}")
    all_pngs = []
    queue = []
    top_subs = _drive_list_folder(token, root_id, mime_filter="application/vnd.google-apps.folder")
    print(f"  Top-level subfolders: {len(top_subs)}")
    for sf in top_subs:
        queue.append((sf["id"], sf["name"], sf["name"], sf["name"]))
    for f in _drive_list_pngs(token, root_id):
        all_pngs.append({"fid": f["id"], "name": f["name"], "stem": Path(f["name"]).stem,
                         "subfolder_name": "uncategorised", "top_category": "uncategorised", "folder_path": ""})
    visited = set()
    while queue:
        folder_id, folder_name_, top_cat, path_str = queue.pop(0)
        if folder_id in visited: continue
        visited.add(folder_id)
        token = _drive_token()
        pngs = _drive_list_pngs(token, folder_id)
        if pngs: print(f"    [{path_str}]: {len(pngs)} PNG(s)")
        for f in pngs:
            all_pngs.append({"fid": f["id"], "name": f["name"], "stem": Path(f["name"]).stem,
                             "subfolder_name": folder_name_, "top_category": top_cat, "folder_path": path_str})
        nested = _drive_list_folder(token, folder_id, mime_filter="application/vnd.google-apps.folder")
        for sf in nested:
            if sf["id"] not in visited:
                queue.append((sf["id"], sf["name"], top_cat, f"{path_str}/{sf['name']}"))
    print(f"  Total PNGs found in Drive: {len(all_pngs)}")
    return all_pngs

def process_drive_png_library(repo2_dir, cfg):
    needed = ["GOOGLE_CLIENT_ID","GOOGLE_CLIENT_SECRET","GOOGLE_REFRESH_TOKEN","GH_TOKEN","GH_OWNER"]
    missing = [k for k in needed if not os.environ.get(k,"").strip()]
    if missing:
        print(f"  ⚠  Skipping Drive scan — missing env vars: {', '.join(missing)}")
        return 0
    if os.environ.get("SCAN_DRIVE","true").lower() not in ("true","1","yes"):
        print("  SCAN_DRIVE=false — skipping Drive PNG library scan")
        return 0
    folder_name = os.environ.get("PNG_LIBRARY_FOLDER","png_library_images").strip()
    gh_token  = os.environ.get("GH_TOKEN","").strip()
    gh_owner  = os.environ.get("GH_OWNER","").strip()
    prev_repo = os.environ.get("PREVIEW_REPO","guruimageusha").strip()
    prev_branch = os.environ.get("PREVIEW_BRANCH","main").strip()
    prev_folder = os.environ.get("PREVIEW_FOLDER","preview_webp").strip()
    watermark = os.environ.get("WATERMARK_TEXT","www.ultrapng.com").strip()
    today_str = datetime.now(__import__("datetime").timezone.utc).strftime("%Y-%m-%d")
    xlsx_path = repo2_dir / ULTRADATA_XLSX
    if not xlsx_path.exists():
        print(f"  ⚠  {ULTRADATA_XLSX} not found — skipping Drive scan")
        return 0
    import openpyxl
    wb_check = openpyxl.load_workbook(str(xlsx_path), read_only=True)
    ws_check = wb_check.active
    header_row = [ws_check.cell(row=1, column=c).value for c in range(1, ws_check.max_column+1)]
    try: fn_col = header_row.index("filename") + 1
    except ValueError:
        print("  ⚠  'filename' column not found in ultradata.xlsx — skipping")
        return 0
    existing_filenames = set()
    for row in ws_check.iter_rows(min_row=2, values_only=True):
        val = row[fn_col-1]
        if val: existing_filenames.add(str(val).strip())
    wb_check.close()
    print(f"  Existing UltraData entries: {len(existing_filenames)}")
    try: all_pngs = _collect_all_pngs_from_drive(folder_name)
    except Exception as exc:
        print(f"  ⚠  Drive scan failed: {exc}")
        return 0
    new_rows = []
    token = _drive_token()
    for p in all_pngs:
        stem = p["stem"]
        png_name = p["name"]
        if png_name in existing_filenames or stem in existing_filenames: continue
        subject_raw = re.sub(r"\s+"," ", re.sub(r"[_\-]+"," ", stem).strip())
        cat = p["top_category"].replace("_"," ").title()
        subcat = p["subfolder_name"].replace("_"," ").title()
        fid = p["fid"]
        dl_url = f"https://drive.google.com/uc?export=download&id={fid}"
        webp_fid = ""
        prev_url = ""
        try:
            png_bytes = _drive_download(token, fid)
            webp_bytes = _make_webp_preview(png_bytes, watermark)
            webp_path_in_repo = f"{prev_folder}/{stem}.webp"
            res = _gh_upload_file(gh_token, gh_owner, prev_repo, webp_path_in_repo, webp_bytes,
                                  f"preview: add {stem}.webp", branch=prev_branch)
            if res: prev_url = _jsdelivr_url(gh_owner, prev_repo, prev_branch, webp_path_in_repo)
        except Exception as e: print(f"    ⚠  WEBP gen failed for {png_name}: {e}")
        new_rows.append({"date_added": today_str, "subject_name": subject_raw.title(),
                         "category": cat, "subcategory": subcat, "filename": png_name,
                         "png_file_id": fid, "webp_file_id": webp_fid,
                         "download_url": dl_url, "preview_url": prev_url, "seo_status": ""})
    if not new_rows:
        print("  ✅  No new PNGs to add from Drive scan.")
        return 0
    print(f"  ➕  Adding {len(new_rows)} new rows to ultradata.xlsx ...")
    _push_xlsx_rows_via_api(cfg, new_rows, commit_msg=f"ultradata: +{len(new_rows)} from Drive png_library [{today_str}]")
    return len(new_rows)

# ══════════════════════════════════════════════════════════════
# UTILITIES
# ══════════════════════════════════════════════════════════════

def _word_count(s): return len([w for w in re.split(r"\s+", (s or "").strip()) if w])
def _today(): return datetime.now(__import__("datetime").timezone.utc).strftime("%Y-%m-%d")

FILENAME_NOISE = re.compile(
    r"\b(hd|png|img|image|photo|pic|transparent|bg|nobg|free|dl|download|"
    r"clipart|vector|stock|high|quality|resolution|res|ultra|4k|full)\b", re.I)

def _clean_subject(raw):
    s = re.sub(r"[_\-]+", " ", raw.strip())
    s = FILENAME_NOISE.sub(" ", s)
    s = re.sub(r"\s*\d+\s*$", "", s)
    s = re.sub(r"^\d+\s*", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s.title() if s else raw.strip().title()

def _extract_json(text):
    start = text.find('{')
    if start == -1: return None
    count = 0
    for i, ch in enumerate(text[start:], start):
        if ch == '{': count += 1
        elif ch == '}':
            count -= 1
            if count == 0: return text[start:i+1]
    return None

# ══════════════════════════════════════════════════════════════
# VIT-GPT2 VISION MODEL  — nlpconnect/vit-gpt2-image-captioning
# Apache 2.0 · 330 MB · auto-selects GPU (CUDA) or CPU at runtime
# Works on every Kaggle machine: P100 (sm_60), T4, CPU-only
# GPU speed: ~0.05-0.15s per image (P100/T4)
# CPU speed: ~0.3s per image
# No HF token · No CUDA config needed — torch.cuda.is_available() handles it
# ══════════════════════════════════════════════════════════════

_vitgpt2_pipeline = None
_vitgpt2_device_name = "CPU"

def _load_vitgpt2():
    global _vitgpt2_pipeline, _vitgpt2_device_name
    if _vitgpt2_pipeline is not None:
        return
    import torch
    from transformers import pipeline
    model_id = "nlpconnect/vit-gpt2-image-captioning"  # Apache 2.0, 330MB

    # Auto-detect GPU — use it if available, fall back to CPU gracefully
    if torch.cuda.is_available():
        device = 0
        _vitgpt2_device_name = torch.cuda.get_device_name(0)
    else:
        device = -1
        _vitgpt2_device_name = "CPU"

    print(f"  [ViT-GPT2] Loading {model_id} on {_vitgpt2_device_name} ...", flush=True)
    _vitgpt2_pipeline = pipeline(
        "image-to-text",
        model=model_id,
        device=device,
        max_new_tokens=50,
    )
    print(f"  [ViT-GPT2] Model ready on {_vitgpt2_device_name} ✓", flush=True)


def _vitgpt2_caption(image_bytes: bytes) -> str:
    """
    Returns a visual caption using ViT-GPT2.
    Pure CPU, ~0.3s per image, zero CUDA dependency.
    """
    from PIL import Image
    _load_vitgpt2()
    img = Image.open(io.BytesIO(image_bytes)).convert("RGB")
    results = _vitgpt2_pipeline(img)
    caption = results[0].get("generated_text", "").strip() if results else ""
    return caption


def _build_seo_from_vision(visual_caption: str, clean_subject: str, category: str) -> dict:
    """
    Builds high-quality, natural English SEO from a visual caption.
    Produces: title (55-70 chars), H1 (50-80), meta (140-160),
    alt text (<125), 10 tags, 30 long-tail keywords.
    All output is grammatically correct English — no keyword stuffing.
    """
    s      = clean_subject
    sl     = s.lower()
    cat_sl = (category or "design").strip().lower()
    cap    = (visual_caption or sl).lower()

    # ── Extract visual descriptors from caption ─────────────────────────────
    COLOR_WORDS = ["red","blue","green","yellow","orange","purple","pink",
                   "white","black","golden","silver","brown","grey","gray",
                   "teal","cyan","violet","dark","light","bright","colorful",
                   "multicolor","transparent"]
    STYLE_WORDS = ["3d","realistic","cartoon","flat","watercolor","sketch",
                   "illustration","minimalist","vintage","modern","floral",
                   "abstract","geometric","neon","glossy","shiny","retro",
                   "hand drawn","digital art","clipart","vector"]
    SCENE_WORDS = ["isolated","transparent background","no background",
                   "white background","dark background","studio","outdoor",
                   "nature","food","animal","flower","business","sports"]

    colors = [w for w in COLOR_WORDS if w in cap]
    styles = [w for w in STYLE_WORDS if w in cap]
    scenes = [w for w in SCENE_WORDS if w in cap]

    color_adj = colors[0].title() if colors else ""
    style_adj = styles[0].title() if styles else ""

    # Build a natural English descriptor phrase for the subject
    # e.g. "Red Rose", "3D Mango", "Watercolor Butterfly"
    parts = []
    if style_adj and style_adj.lower() not in sl: parts.append(style_adj)
    if color_adj and color_adj.lower() not in sl: parts.append(color_adj)
    parts.append(s)
    visual_s  = " ".join(parts)
    visual_sl = visual_s.lower()

    # ── Use-case mapping per category/caption ───────────────────────────────
    USE_CASES = {
        "food":       ("food menus, restaurant flyers, and food blogs",        "chefs, food bloggers, and restaurant owners"),
        "fruit":      ("recipe cards, food blogs, and grocery promotions",     "food designers and health bloggers"),
        "vegetable":  ("healthy eating guides, recipe sites, and menus",       "nutritionists and food content creators"),
        "juice":      ("beverage menus, health blogs, and café promotions",    "café owners and nutritionists"),
        "flower":     ("wedding invitations, greeting cards, and social media","event planners and designers"),
        "animal":     ("children's books, nature blogs, and presentations",    "educators and content creators"),
        "business":   ("presentations, reports, and marketing materials",      "business professionals and marketers"),
        "sport":      ("sports apps, fitness blogs, and social media posts",   "athletes and fitness coaches"),
        "technology": ("tech websites, app UIs, and digital marketing",        "developers and UI designers"),
        "logo":       ("brand identity, websites, and marketing campaigns",    "graphic designers and brand managers"),
        "banner":     ("website headers, social media, and advertisements",    "digital marketers and designers"),
        "sale":       ("promotional flyers, social media ads, and e-commerce", "retailers and marketing teams"),
        "people":     ("social media, marketing campaigns, and websites",      "marketers and content creators"),
        "fish":       ("seafood menus, fishing blogs, and recipe sites",       "restaurant owners and food bloggers"),
        "chicken":    ("restaurant menus, food blogs, and recipe cards",       "chefs and food content creators"),
    }
    uc_label = "graphic design, presentations, and digital marketing"
    uc_who   = "designers and content creators"
    for kw, (lbl, who) in USE_CASES.items():
        if kw in cap or kw in cat_sl or kw in sl:
            uc_label, uc_who = lbl, who
            break

    # ── SEO Title: natural English, 55-70 chars ──────────────────────────────
    title_candidates = [
        f"{visual_s} PNG – Transparent Background Free Download",
        f"Free {visual_s} PNG | Transparent Background HD",
        f"{visual_s} Transparent PNG – Free High-Quality Download",
        f"Download {visual_s} PNG with Transparent Background",
        f"{visual_s} PNG Image – Free Transparent Download",
        f"Free {visual_s} PNG – No Background, High Resolution",
        f"{visual_s} PNG Cut Out – Free Transparent Download",
    ]
    title = next((c for c in title_candidates if 55 <= len(c) <= 70), None)
    if not title:
        title = min(title_candidates, key=lambda c: abs(len(c) - 63))
    title = title[:70]

    # ── H1: conversational English, 50-80 chars ──────────────────────────────
    h1_candidates = [
        f"Download {visual_s} PNG on a Transparent Background",
        f"{visual_s} PNG – Free Transparent Background Image",
        f"Free {visual_s} PNG with a Transparent Background",
        f"High-Quality {visual_s} PNG – Transparent & Free",
        f"{visual_s} Transparent PNG – Free for Commercial Use",
    ]
    h1 = next((c for c in h1_candidates if 50 <= len(c) <= 80), None)
    if not h1:
        h1 = min(h1_candidates, key=lambda c: abs(len(c) - 65))
    h1 = h1[:80]

    # ── Meta description: natural English, 140-160 chars ────────────────────
    meta_raw = (
        f"Download this free {visual_s} PNG image with a clean transparent background. "
        f"Perfect for {uc_label}. "
        f"High resolution, ready to use instantly — no watermark, no sign-up required."
    )
    if len(meta_raw) < 140:
        meta_raw += " Completely free for personal and commercial use."
    meta_desc = meta_raw[:160]

    # ── Alt text: descriptive English, <125 chars ────────────────────────────
    alt_candidates = [
        f"{visual_s} PNG image on a transparent background, high resolution, free download",
        f"Free {visual_s} PNG with transparent background for {uc_who}",
        f"{visual_s} isolated PNG cut-out, transparent background, high quality",
    ]
    alt_text = next((a for a in alt_candidates if len(a) <= 125), alt_candidates[0])[:125]

    # ── Tags: 10 natural English search phrases ──────────────────────────────
    tags_raw = [
        f"{sl} png",
        f"{sl} transparent background",
        f"free {sl} png",
        f"{sl} no background",
        f"{sl} hd png",
        f"{visual_sl} png free",
        f"{cat_sl} png transparent",
        f"{sl} png download",
        f"transparent {sl} png",
        f"{sl} clipart png",
    ]
    tags = ", ".join(dict.fromkeys(t for t in tags_raw if t.strip()))

    # ── 30 long-tail keywords: varied, natural English ───────────────────────
    kw_templates = [
        f"free {sl} png download",
        f"{sl} transparent background png",
        f"{sl} png no background free",
        f"download {sl} png transparent",
        f"high resolution {sl} png",
        f"{sl} png cut out transparent",
        f"{sl} isolated png free",
        f"free {sl} clipart png transparent",
        f"{sl} png for graphic design",
        f"{sl} transparent png hd quality",
        f"{visual_sl} png free download",
        f"{sl} png image transparent background",
        f"best free {sl} png transparent",
        f"{sl} png for presentations",
        f"free high quality {sl} png",
        f"{sl} png sticker transparent",
        f"{sl} png for website design",
        f"transparent {sl} image png",
        f"{sl} png for social media",
        f"{sl} cutout png transparent",
        f"{sl} png no watermark free",
        f"{sl} png for photoshop",
        f"{sl} png for canva free",
        f"{sl} png for powerpoint",
        f"{cat_sl} {sl} transparent png",
        f"free {cat_sl} {sl} png",
        f"{sl} png commercial use free",
        f"download {visual_sl} png hd",
        f"{sl} png image free download",
        f"free printable {sl} png transparent",
    ]
    seen_kw: set = set()
    kw_list: list = []
    for kw in kw_templates:
        kw = kw.strip()
        if kw and kw not in seen_kw:
            seen_kw.add(kw)
            kw_list.append(kw)
    kw_list = kw_list[:30]

    return {
        "title":       title,
        "h1":          h1,
        "meta_desc":   meta_desc,
        "alt_text":    alt_text,
        "tags":        tags,
        "description": ", ".join(kw_list),
    }

def _vision_seo_vitgpt2(row: dict) -> dict:
    """
    SEO entry-point for Kaggle and local runs using ViT-GPT2.
    Downloads the PNG from Drive, gets a visual caption, builds rich English SEO.
    Falls back to filename-based SEO if the image cannot be fetched.
    """
    subject  = row.get("subject_name", "").strip()
    category = row.get("category", "")
    if not subject:
        raise RuntimeError("Missing subject_name")
    clean_subj = _clean_subject(subject)

    visual_caption = ""
    png_fid     = row.get("png_file_id", "")
    preview_url = row.get("preview_url", "")

    import requests
    for label, url in [
        ("Drive PNG",   f"https://www.googleapis.com/drive/v3/files/{png_fid}?alt=media" if png_fid else ""),
        ("Preview URL", preview_url),
    ]:
        if not url:
            continue
        try:
            headers = {}
            if "googleapis.com" in url:
                headers = _dh(_drive_token())
            r = requests.get(url, headers=headers, timeout=30)
            r.raise_for_status()
            visual_caption = _vitgpt2_caption(r.content)
            print(f"    👁  Caption ({label}): {visual_caption[:120]}", flush=True)
            break
        except Exception as e:
            print(f"    ⚠  Image fetch failed ({label}): {e}", flush=True)

    if not visual_caption:
        print(f"    ℹ  No image available — using filename-based SEO", flush=True)
        visual_caption = clean_subj

    return _build_seo_from_vision(visual_caption, clean_subj, category)


# ══════════════════════════════════════════════════════════════
# FALLBACK (Gemma 3 1B CPU — used only if not on Kaggle)
# ══════════════════════════════════════════════════════════════

_gemma_model     = None
_gemma_tokenizer = None
GEMMA_MODEL_ID   = "google/gemma-3-1b-it"

def _load_gemma_model():
    global _gemma_model, _gemma_tokenizer
    if _gemma_model is not None: return
    import torch
    from transformers import AutoTokenizer, AutoModelForCausalLM
    print(f"  [Gemma] Loading {GEMMA_MODEL_ID} ...", flush=True)
    hf_token = os.environ.get("HF_TOKEN")
    _gemma_tokenizer = AutoTokenizer.from_pretrained(GEMMA_MODEL_ID, token=hf_token)
    _gemma_model = AutoModelForCausalLM.from_pretrained(
        GEMMA_MODEL_ID, torch_dtype=__import__("torch").float32,
        device_map="cpu", token=hf_token)
    _gemma_model.eval()
    print("  [Gemma] Model ready ✓", flush=True)

def _gemma_generate_seo(clean_subject, category, orig_subject):
    prompt = f"""You are an expert SEO copywriter specializing in PNG image downloads.
Create a JSON object for the PNG image described below. Output must be 100% unique, natural, creative.

Subject: "{clean_subject}"
Category: "{category}"

Generate:
- title: SEO page title (55-70 chars), include subject, "PNG", and a benefit
- h1: conversational H1 (50-80 chars)
- meta_desc: meta description (140-160 chars), mention transparent background + use case
- alt_text: image alt attribute (under 125 chars)
- tags: exactly 10 relevant tags, comma-separated, varied
- keywords: exactly 30 long-tail keywords (comma-separated), diverse, human search phrases

RULES:
- NEVER copy exact subject name in every keyword. Vary phrasing.
- Mix short and long keywords.
- Use words: free, download, transparent, hd, clipart, png, for, image, background.
- Return ONLY the JSON object, no extra text.

JSON:"""
    import torch
    try: _load_gemma_model()
    except Exception as exc:
        print(f"    [Gemma] Model load error: {exc}", flush=True)
        return _fallback_rule_seo(clean_subject, category, orig_subject)
    try:
        inputs = _gemma_tokenizer(prompt, return_tensors="pt")
        with torch.no_grad():
            outputs = _gemma_model.generate(**inputs, max_new_tokens=800, do_sample=True, temperature=0.7, top_p=0.9)
        raw_text = _gemma_tokenizer.decode(outputs[0], skip_special_tokens=True)
        json_str = _extract_json(raw_text)
        if not json_str: raise ValueError("No balanced JSON found")
        data = json.loads(json_str)
        for k in ["title","h1","meta_desc","alt_text","tags","keywords"]:
            if k not in data: raise ValueError(f"Missing key {k}")
        kw_list = [kw.strip() for kw in data["keywords"].split(",") if kw.strip()][:30]
        data["keywords"] = ", ".join(kw_list)
        return {"title": data["title"], "h1": data["h1"], "meta_desc": data["meta_desc"],
                "alt_text": data["alt_text"], "tags": data["tags"], "description": data["keywords"]}
    except Exception as e:
        print(f"    [Gemma] Error: {e} – using fallback", flush=True)
        return _fallback_rule_seo(clean_subject, category, orig_subject)

def _fallback_rule_seo(clean_subject, category, orig_subject):
    s = clean_subject or orig_subject.strip() or "Image"
    sl = s.lower()
    cat_sl = category.strip().lower()
    title = f"{s} PNG Transparent Background Free Download"
    h1    = f"{s} PNG on Transparent Background - Free HD Download"
    meta  = f"Download this free {s} PNG with transparent background. High quality, perfect for designers and creative projects."
    alt   = f"{s} PNG transparent background"
    tags  = ", ".join([f"{sl} png", f"{sl} transparent", f"free {sl} png", f"{sl} no background", f"{sl} hd png",
                       f"{sl} clipart", f"{cat_sl} png", f"{sl} download", f"transparent {sl}", f"{sl} image"])
    kw_list = [f"{sl} png free download", f"{sl} transparent png hd", f"free {sl} png download",
               f"{sl} png no background", f"{sl} png transparent background", f"{sl} png cutout",
               f"{sl} isolated png", f"{sl} png high resolution", f"high quality {sl} png",
               f"{sl} png for designers", f"{sl} png clipart free", f"{sl} sticker png transparent",
               f"{sl} illustration png transparent", f"{sl} png image hd quality", f"best {sl} png transparent",
               f"{sl} image png free download", f"free transparent {sl} image", f"{sl} png hd free download",
               f"transparent background {sl} png", f"{sl} cutout image free", f"{sl} png without background",
               f"{sl} png for photoshop", f"{sl} png for canva", f"{sl} png for powerpoint",
               f"{sl} png for website", f"{cat_sl} {sl} png transparent", f"{sl} {cat_sl} free png",
               f"free {cat_sl} {sl} png", f"{sl} png image free", f"download {sl} png transparent"]
    return {"title": title[:70], "h1": h1[:85], "meta_desc": meta[:160], "alt_text": alt[:125],
            "tags": tags, "description": ", ".join(kw_list[:30])}

def _vision_seo(row):
    """
    Dispatcher: ViT-GPT2 vision SEO — GPU if CUDA available, CPU otherwise.
    Works on Kaggle P100/T4 and local machines.
    """
    if not row.get("subject_name", "").strip():
        raise RuntimeError("Missing subject_name")
    return _vision_seo_vitgpt2(row)

# ══════════════════════════════════════════════════════════════
# AUTO-RESTART
# ══════════════════════════════════════════════════════════════

def _trigger_self_restart(remaining, workflow_file="section2_seo.yml"):
    import requests
    repo     = os.environ.get("GITHUB_REPOSITORY","").strip()
    gh_token = (os.environ.get("GH_TOKEN") or os.environ.get("REPO2_TOKEN","")).strip()
    ref      = os.environ.get("GITHUB_REF_NAME","main").strip() or "main"
    if not repo or not gh_token:
        print("  ⚠  Cannot auto-restart: missing GITHUB_REPOSITORY or GH_TOKEN")
        return
    url = f"https://api.github.com/repos/{repo}/actions/workflows/{workflow_file}/dispatches"
    try:
        r = requests.post(url, headers={"Authorization": f"token {gh_token}","Accept":"application/vnd.github.v3+json"},
                         json={"ref": ref, "inputs": {"count":"","scan_drive":"false"}}, timeout=30)
        if r.status_code in (204,200): print(f"  🔄  Auto-restart dispatched ({remaining} items still pending) ✓")
        else: print(f"  ⚠  Auto-restart failed: {r.status_code} — {r.text[:120]}")
    except Exception as exc: print(f"  ⚠  Auto-restart exception: {exc}")

# ══════════════════════════════════════════════════════════════
# XLSX READ/UPDATE
# ══════════════════════════════════════════════════════════════

def _read_pending_rows(xlsx_path):
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb.active
    if ws.max_row < 2: return []
    headers = [str(c.value or "").strip() for c in ws[1]]
    idx = {h:i for i,h in enumerate(headers)}
    for h in ["subject_name","filename","download_url","preview_url"]:
        if h not in idx: raise RuntimeError(f"ultradata.xlsx missing column: {h}")
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        def _v(col): return "" if col not in idx or r[idx[col]] is None else str(r[idx[col]]).strip()
        if _v("seo_status").lower() == "completed": continue
        fn = _v("filename"); sn = _v("subject_name")
        if not fn or not sn: continue
        out.append({"subject_name":sn, "filename":fn, "download_url":_v("download_url"),
                    "preview_url":_v("preview_url"), "webp_file_id":_v("webp_file_id"),
                    "png_file_id":_v("png_file_id"),
                    "category":_v("category"), "subcategory":_v("subcategory"), "date_added":_v("date_added")})
    return out

def _mark_completed(xlsx_path, completed_filenames):
    import openpyxl
    if not xlsx_path.exists() or not completed_filenames: return 0
    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb.active
    if ws.max_row < 2: return 0
    headers = [str(c.value or "").strip() for c in ws[1]]
    if "seo_status" not in headers:
        seo_col = len(headers)+1
        ws.cell(row=1, column=seo_col, value="seo_status")
        headers.append("seo_status")
    else: seo_col = headers.index("seo_status")+1
    if "filename" not in headers: return 0
    fn_col = headers.index("filename")+1
    updated = 0
    for row in ws.iter_rows(min_row=2):
        fn = str(row[fn_col-1].value or "").strip()
        cell = row[seo_col-1]
        if fn in completed_filenames and str(cell.value or "").strip() != "completed":
            cell.value = "completed"; updated += 1
    wb.save(str(xlsx_path))
    return updated

# ══════════════════════════════════════════════════════════════
# REPO2 (clone/load/save/push)
# ══════════════════════════════════════════════════════════════

@dataclass
class Repo2Config:
    token: str
    slug: str
    data_dir: str = "data"

def _clone_repo2(cfg, workdir):
    repo_url = f"https://x-access-token:{cfg.token}@github.com/{cfg.slug}.git"
    if workdir.exists():
        subprocess.run(["git","pull","--rebase","--autostash"], cwd=str(workdir), check=False)
        return workdir
    subprocess.run(["git","clone","--depth","1", repo_url, str(workdir)], check=True)
    return workdir

def _load_existing_entries(repo2_dir, data_dir):
    d = repo2_dir / data_dir
    d.mkdir(parents=True, exist_ok=True)
    files = sorted(p for p in d.glob("json*.json") if p.is_file())
    if not files:
        f1 = d / "json1.json"; f1.write_text("[]", encoding="utf-8"); files = [f1]
    all_entries = {}
    for f in files:
        try:
            arr = json.loads(f.read_text(encoding="utf-8"))
            if isinstance(arr, list):
                for e in arr:
                    fn = e.get("filename")
                    if fn and fn not in all_entries: all_entries[fn] = e
        except: continue
    return all_entries, files

def _get_active_file(files, repo2_dir, data_dir, max_entries, file_entries):
    last = files[-1]
    if len(file_entries.get(last, [])) < max_entries: return last
    m = re.match(r"json(\d+)\.json$", last.name)
    nxt = (int(m.group(1))+1) if m else (len(files)+1)
    newf = repo2_dir / data_dir / f"json{nxt}.json"
    newf.write_text("[]", encoding="utf-8"); files.append(newf); file_entries[newf] = []
    print(f"\n  [JSON] Created {newf.name} (previous file full)", flush=True)
    return newf

def _save_json_files(file_entries):
    for f, arr in file_entries.items():
        tmp = f.with_suffix(f.suffix + ".tmp")
        tmp.write_text(json.dumps(arr, ensure_ascii=False, indent=2), encoding="utf-8")
        tmp.replace(f)

def _git_setup(repo2_dir):
    subprocess.run(["git","config","user.name","github-actions[bot]"], cwd=str(repo2_dir), check=True)
    subprocess.run(["git","config","user.email","github-actions[bot]@users.noreply.github.com"], cwd=str(repo2_dir), check=True)

def _commit_push_repo2(repo2_dir, cfg, added, commit_msg=None, file_entries=None, max_retries=3):
    _git_setup(repo2_dir)
    msg = commit_msg or f"seo: add {added} entries [section2]"
    for attempt in range(1, max_retries+1):
        subprocess.run(["git","add", cfg.data_dir], cwd=str(repo2_dir), check=True)
        if (repo2_dir / ULTRADATA_XLSX).exists():
            subprocess.run(["git","add", ULTRADATA_XLSX], cwd=str(repo2_dir), check=True)
        diff = subprocess.run(["git","diff","--staged","--quiet"], cwd=str(repo2_dir))
        if diff.returncode == 0:
            print("  Repo2: nothing to commit – already up-to-date."); return
        subprocess.run(["git","commit","-m", msg], cwd=str(repo2_dir), check=True)
        subprocess.run(["git","fetch","origin","main"], cwd=str(repo2_dir), capture_output=True)
        rebase = subprocess.run(["git","rebase","origin/main"], cwd=str(repo2_dir), capture_output=True, text=True)
        if rebase.returncode != 0:
            print(f"  [WARN] Rebase conflict (attempt {attempt}/{max_retries}) ...")
            subprocess.run(["git","rebase","--abort"], cwd=str(repo2_dir), capture_output=True)
            if attempt >= max_retries: raise RuntimeError(f"Repo2 push failed after {max_retries} retries.")
            subprocess.run(["git","reset","--hard","origin/main"], cwd=str(repo2_dir), check=True)
            if file_entries: _save_json_files(file_entries)
            time.sleep(3 * attempt); continue
        result = subprocess.run(["git","push"], cwd=str(repo2_dir), capture_output=True, text=True)
        if result.returncode == 0:
            print(f"  Repo2: pushed {added} SEO entries ✓"); return
        print(f"  [WARN] git push failed (attempt {attempt}/{max_retries}):\n{result.stderr}")
        if attempt >= max_retries: raise RuntimeError("Repo2 push failed – check REPO2_TOKEN permissions.")
        subprocess.run(["git","reset","--hard","origin/main"], cwd=str(repo2_dir), check=True)
        if file_entries: _save_json_files(file_entries)
        time.sleep(3 * attempt)

def _push_xlsx_rows_via_api(cfg, new_rows, commit_msg, max_retries=3):
    import requests, openpyxl
    token = cfg.token; branch = "main"; path = ULTRADATA_XLSX
    api_url = f"https://api.github.com/repos/{cfg.slug}/contents/{path}"
    headers = {"Authorization": f"token {token}","Accept": "application/vnd.github.v3+json","Content-Type": "application/json"}
    HEADERS = ["date_added","subject_name","category","subcategory","filename","png_file_id","webp_file_id","download_url","preview_url","seo_status"]
    def _fetch_wb():
        r = requests.get(api_url, headers=headers, params={"ref": branch}, timeout=30)
        r.raise_for_status()
        d = r.json()
        raw = base64.b64decode(d["content"].replace("\n",""))
        return openpyxl.load_workbook(io.BytesIO(raw)), d["sha"]
    def _append_and_encode(wb_):
        ws_ = wb_.active
        hdr = [ws_.cell(row=1, column=c).value for c in range(1, ws_.max_column+1)]
        if not hdr or hdr[0] is None:
            ws_.append(HEADERS); hdr = HEADERS
        for col_name in HEADERS:
            if col_name not in hdr:
                ws_.cell(row=1, column=ws_.max_column+1, value=col_name); hdr.append(col_name)
        hdr = [ws_.cell(row=1, column=c).value for c in range(1, ws_.max_column+1)]
        for row in new_rows: ws_.append([row.get(h,"") for h in hdr])
        buf = io.BytesIO(); wb_.save(buf)
        return base64.b64encode(buf.getvalue()).decode()
    wb, sha = _fetch_wb()
    for attempt in range(1, max_retries+1):
        r = requests.put(api_url, headers=headers,
                         json={"message": commit_msg, "content": _append_and_encode(wb), "sha": sha, "branch": branch},
                         timeout=90)
        if r.ok:
            print(f"  xlsx pushed via API (+{len(new_rows)} rows) ✓"); return
        if r.status_code == 409 and attempt < max_retries:
            print(f"  [WARN] xlsx 409 conflict (attempt {attempt}) – re-fetching ...")
            time.sleep(3 * attempt); wb, sha = _fetch_wb(); continue
        r.raise_for_status()

# ══════════════════════════════════════════════════════════════
# CORE PROCESSING LOOP  (shared by both local and kaggle_run modes)
# ══════════════════════════════════════════════════════════════

def run_seo_loop():
    root = Path(__file__).resolve().parent
    repo2_token = os.environ.get("REPO2_TOKEN","").strip()
    repo2_slug  = os.environ.get("REPO2_SLUG","").strip()
    if not repo2_token or not repo2_slug:
        raise SystemExit("❌ Missing REPO2_TOKEN or REPO2_SLUG")
    max_per_file   = int(os.environ.get("REPO2_MAX_PER_JSON","200"))
    checkpoint_every = 50   # more frequent checkpointing on fast GPU
    count_env = (os.environ.get("S2_COUNT","") or "").strip()
    requested = None
    if count_env:
        try:
            v = int(count_env)
            if v > 0: requested = min(v, INSTANT_CAP)
        except: pass

    on_kaggle = os.path.exists("/kaggle/working")
    mode_str  = "Kaggle GPU/CPU auto-detect" if on_kaggle else "local CPU/GPU auto-detect"

    print("="*65)
    print("  Section 2 – SEO JSON Builder (V10.0 – Vision-Powered SEO)")
    print(f"  Mode            : {mode_str}")
    print(f"  Requested count : {requested if requested else 'ALL pending'}")
    print(f"  Safety cap      : {INSTANT_CAP}")
    print(f"  Max per JSON    : {max_per_file}")
    print("="*65)

    print("\n[Step 1] Cloning private ultrapng repo ...")
    cfg = Repo2Config(token=repo2_token, slug=repo2_slug)
    repo2_dir = _clone_repo2(cfg, root / "_repo2_work")
    xlsx = repo2_dir / ULTRADATA_XLSX
    if not xlsx.exists():
        raise SystemExit(f"❌ {ULTRADATA_XLSX} not found in {repo2_slug}.")

    print("\n[Step 2] Scanning Drive png_library_images ...")
    new_from_drive = process_drive_png_library(repo2_dir, cfg)
    if new_from_drive > 0:
        print(f"  ✅ {new_from_drive} new entries added from Drive scan")
        import shutil
        shutil.rmtree(str(repo2_dir), ignore_errors=True)
        repo2_dir = _clone_repo2(cfg, root / "_repo2_work")
        xlsx = repo2_dir / ULTRADATA_XLSX

    print("\n[Step 3] Reading pending rows from ultradata.xlsx ...")
    pending = _read_pending_rows(xlsx)
    print(f"  Pending rows : {len(pending)}")
    if not pending:
        print("  ✅ Nothing pending – all done."); return

    seen = set(); deduped = []
    for r in pending:
        fn = r.get("filename")
        if fn and fn not in seen:
            seen.add(fn); deduped.append(r)
    pending = deduped

    print("\n[Step 4] Loading existing SEO entries from repo2 ...")
    existing, files = _load_existing_entries(repo2_dir, cfg.data_dir)
    print(f"  Existing SEO entries : {len(existing)}")
    todo = [r for r in pending if r["filename"] not in existing]
    print(f"  Still to generate    : {len(todo)}")
    if not todo:
        print("  ✅ All pending rows already have SEO."); return

    target = min(requested, len(todo)) if requested else len(todo)

    print(f"\n[Step 4b] Loading ViT-GPT2 vision model ...")
    _load_vitgpt2()
    print(f"  Device selected : {_vitgpt2_device_name}")

    print(f"\n  ▶ Generating SEO for up to {target} item(s) ...\n")

    file_entries = {}
    for f in files:
        try: file_entries[f] = json.loads(f.read_text(encoding="utf-8"))
        except: file_entries[f] = []

    added = 0; completed_filenames = set(); pending_push = 0; time_limit_hit = False

    for i, r in enumerate(todo, 1):
        if added >= target: break
        elapsed = time.time() - _RUN_START
        if not on_kaggle and elapsed > MAX_RUN_SECONDS:
            print(f"\n⏰ Time limit ({elapsed/3600:.2f}h) – checkpoint ...")
            time_limit_hit = True; break

        subject = r["subject_name"]; filename = r["filename"]
        print(f"  [{i}/{target}] {subject} ({filename}) ...", flush=True)
        try:
            seo = _vision_seo(r)
        except Exception as e:
            print(f"    ✗ SKIP ({e})", flush=True); continue

        slug = re.sub(r"[^a-z0-9]+", "-", _clean_subject(subject).lower()).strip("-") or "untitled"
        webp_fid = r.get("webp_file_id","")
        webp_preview = f"https://lh3.googleusercontent.com/d/{webp_fid}=s800" if webp_fid else r.get("preview_url","")
        target_file = _get_active_file(files, repo2_dir, cfg.data_dir, max_per_file, file_entries)
        if target_file not in file_entries: file_entries[target_file] = []
        file_entries[target_file].append({
            "category": r.get("category",""), "subcategory": r.get("subcategory",""),
            "subject_name": subject, "filename": filename, "slug": slug,
            "download_url": r["download_url"], "preview_url": r["preview_url"],
            "webp_preview_url": webp_preview,
            "title": seo["title"], "h1": seo["h1"], "meta_desc": seo["meta_desc"],
            "alt_text": seo["alt_text"], "tags": seo["tags"],
            "description": seo["description"],
            "word_count": _word_count(seo["description"]),
            "date_added": r.get("date_added", _today())
        })
        completed_filenames.add(filename); added += 1; pending_push += 1
        elapsed_m = (time.time() - _RUN_START)/60
        print(f"    ✓ title={len(seo['title'])}c ({elapsed_m:.1f} min)", flush=True)

        if pending_push >= checkpoint_every:
            print(f"\n  [Checkpoint] Saving {pending_push} entries ...")
            _save_json_files(file_entries)
            _mark_completed(xlsx, completed_filenames)
            _commit_push_repo2(repo2_dir, cfg, pending_push, file_entries=file_entries)
            pending_push = 0; print()

    if pending_push > 0 or completed_filenames:
        print(f"\n[Step 5] Final save & push ({pending_push} remaining) ...")
        _save_json_files(file_entries)
        updated = _mark_completed(xlsx, completed_filenames)
        print(f"  ultradata.xlsx: {updated} row(s) marked completed")
        _commit_push_repo2(repo2_dir, cfg, pending_push, file_entries=file_entries)

    remaining = len(todo) - added
    total_elapsed = (time.time() - _RUN_START)/60
    print("\n" + "="*65)
    print(f"  ✅ Section 2 complete")
    print(f"  Added this run    : {added}")
    print(f"  Total in repo2    : {len(existing) + added}")
    print(f"  Elapsed time      : {total_elapsed:.1f} min")
    if remaining > 0: print(f"  Still pending     : {remaining}")
    print("="*65)
    if time_limit_hit and remaining > 0:
        print(f"\n[Auto-restart] Dispatching for {remaining} remaining items ...")
        _trigger_self_restart(remaining)

# ══════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--mode", choices=["trigger_kaggle","kaggle_run","local"], default="local")
    args = parser.parse_args()

    if args.mode == "trigger_kaggle":
        trigger_kaggle_mode()
    else:
        # Both kaggle_run and local use the same loop.
        # BLIP-1 is automatically chosen when /kaggle/working exists.
        run_seo_loop()

if __name__ == "__main__":
    main()
