"""
╔══════════════════════════════════════════════════════════════╗
║   UltraPNG.com — Logo Z-Image Pipeline V2.0                 ║
╠══════════════════════════════════════════════════════════════╣
║  PHASE 1B → Z-Image-Turbo        1024x1024 logos           ║
║  PHASE 3  → RMBG-2.0 ONNX        Background removal        ║
║  PHASE 4  → Google Drive          Upload PNG + WebP         ║
║  PHASE 5  → ultradata.xlsx        Append rows in REPO1      ║
║  PHASE 6  → Save Run Logs → REPO1  (visible on GitHub!)    ║
╚══════════════════════════════════════════════════════════════╝

STANDALONE FILE — Does NOT import from main_pipeline.py
inject_creds.py prepends os.environ[] lines before this file.

MODEL CHANGE (V1 → V2):
  V1: FLUX.1-dev + Shakker-Labs Logo LoRA  (trigger: "wablogo, logo, ")
  V2: Z-Image-Turbo                        (no LoRA — direct generation)

  Why no LoRA?
    FLUX LoRAs target FLUX's MMDiT architecture.
    Z-Image uses S3-DiT (Lumina-based) — completely different weight shapes.
    Loading FLUX LoRA into Z-Image will crash. Not compatible.

STYLE PREFIX (replaces LoRA trigger word):
  "professional vector logo design, sharp clean lines, bold typography,
   isolated on white background, "
  → Natural language style guide — no special activation token needed.
  → "isolated on white background" helps Phase 3 RMBG removal.

VERIFIED FIXES IN THIS VERSION:
  [FIX 1] Generator: torch.Generator("cuda") when GPU available (not "cpu")
  [FIX 2] T4 (sm_75): also uses enable_model_cpu_offload() — 12GB+7GB > 16GB
  [FIX 3] All 16GB GPUs: enable_model_cpu_offload() (text encoder offloads first)
  [FIX 4] max_sequence_length=256 (was 1024 — logo prompts are ~50-80 tokens only!)
  [FIX 5] Updated LOGO_STYLE_PREFIX for Z-Image (descriptive, not LoRA token)
  [FIX 6] set_progress_bar_config: placed after CPU offload setup
  [FIX 7] Cache cleanup: uses correct HF folder pattern models--Tongyi-MAI--*

PERF FIXES V2.1 (P100 SPEED):
  [PERF 1] P100: 4-bit NF4 Qwen3-4B text encoder (7GB→2GB) + cpu_offload
           Result: enc offloads fast (2GB vs 7GB), DiT gets full VRAM for inference
  [PERF 2] VAE forced float32 on P100 → prevents NaN from float16 decode
           Root cause: NaN → blank images → triggered retry on EVERY image → 2x slow
  [PERF 3] Blank detection fixed: threshold tuned for white-bg logos, NaN check added
  [PERF 4] max_sequence_length 1024→256: logo prompts are <80 tokens, was wasting time
  [PERF 5] PNG compress_level 9→1: was adding ~3s per image with no quality benefit

OOM FIXES V2.2 (P100 16GB):
  [OOM 1] cpu_offload re-enabled WITH 4-bit encoder — previous code loaded 15.3GB
          on GPU leaving NO room for inference activations → OOM on EVERY image!
          With offload: encoder (2GB) moves to CPU after encode → DiT gets full VRAM.
  [OOM 2] OOM retry at 768x768: if 1024x1024 OOMs, auto-retry at 768 + upscale to 1024.
          This uses ~44% less activation memory.
  [OOM 3] torch.no_grad() wrapper around inference → prevents gradient accumulation.
  [OOM 4] Consecutive OOM threshold raised 3→5 (since we now have retry strategies).

MULTI-GPU V2.3 (T4 x2 — 2×16GB = 32GB total):
  [T4X2 1] Auto-detect num_gpus >= 2 → activate multi-GPU split mode.
  [T4X2 2] GPU:0 = transformer (12GB) + VAE (1GB) = 13GB on first GPU.
           GPU:1 = text_encoder Qwen3-4B full fp16 (7GB) on second GPU.
  [T4X2 3] No quantization needed (full precision encoder = better quality).
  [T4X2 4] No cpu_offload needed (no CPU↔GPU transfer = maximum speed).
  [T4X2 5] Expected: ~40-60s per image (vs ~2min on P100 with offload).

TARGET CATEGORIES (set LOGO_LORA_CATEGORY env var to pick one):
  admissions_enrollment | openings_launches | events
  education | announcements | business_promotions
  calls_to_action | (leave empty = run ALL logo categories)
"""

import os, sys, json, time, gc, re, io, shutil, base64, subprocess, math
from pathlib import Path
from datetime import datetime

# ── Force real-time log output in Kaggle ──────────────────────
try:
    sys.stdout.reconfigure(line_buffering=True)
    sys.stderr.reconfigure(line_buffering=True)
except AttributeError:
    pass

os.environ["PYTORCH_CUDA_ALLOC_CONF"] = "expandable_segments:True"
ULTRADATA_XLSX_NAME = "ultradata.xlsx"

# ── HuggingFace cache dir ──────────────────────────────────────
# Z-Image-Turbo: transformer ~12GB + text encoder ~7GB = ~19GB total disk
# BUT with enable_model_cpu_offload(), peak VRAM = 12GB (fits in 16GB!)
# /kaggle/tmp has ~57GB free — safe for downloads.
import shutil as _shutil_space
def _pick_hf_cache():
    for candidate in ["/kaggle/tmp/hf_cache", "/kaggle/working/hf_cache"]:
        p = Path(candidate)
        try:
            p.mkdir(parents=True, exist_ok=True)
            stat = _shutil_space.disk_usage(str(p))
            free_gb = stat.free / 1e9
            print(f"  [cache] {p}  free={free_gb:.1f}GB")
            if free_gb >= 20:   # Z-Image needs ~19GB disk (vs FLUX's ~24GB)
                return p
        except Exception as e:
            print(f"  [cache] {p} unavailable: {e}")
    p = Path("/kaggle/tmp/hf_cache")
    p.mkdir(parents=True, exist_ok=True)
    return p

HF_CACHE = _pick_hf_cache()
print(f"  [cache] Using HF_CACHE = {HF_CACHE}")
os.environ["HF_HOME"]               = str(HF_CACHE)
os.environ["HUGGINGFACE_HUB_CACHE"] = str(HF_CACHE)
os.environ["TRANSFORMERS_CACHE"]    = str(HF_CACHE)
os.environ["HF_HUB_ENABLE_HF_TRANSFER"] = "1"   # [FIX TIMEOUT] Rust downloader — no ReadTimeout on 19GB model
os.environ["HF_HUB_DOWNLOAD_TIMEOUT"]    = "600"  # 10 min safety net for slow links
_hf_token = os.environ.get("HF_TOKEN", "")
if _hf_token:
    os.environ["HUGGINGFACE_HUB_TOKEN"] = _hf_token
    from huggingface_hub import login as _hf_login
    _hf_login(token=_hf_token, add_to_git_credential=False)
    print(f"  HuggingFace login OK (token: {_hf_token[:8]}...)")
else:
    # Z-Image-Turbo is public — no HF_TOKEN needed
    print("  INFO: HF_TOKEN not set. Z-Image-Turbo is public — download will proceed.")

# ══════════════════════════════════════════════════════════════
# Z-IMAGE MODEL CONFIG
# ══════════════════════════════════════════════════════════════
ZIMAGE_HF_ID   = "Tongyi-MAI/Z-Image-Turbo"

ZIMAGE_STEPS   = 9      # 9 steps = 8 actual DiT forwards (distilled optimal)
                         # Do NOT go below 8 — quality degrades

ZIMAGE_GUIDANCE = 0.0   # MUST be 0.0 for distilled Turbo models
                         # CFG is disabled in distillation

ZIMAGE_MAX_SEQ = 256    # [PERF FIX] Logo prompts are ~50-80 tokens — 1024 was overkill!
                         # 256 is plenty for all logo category prompts.
                         # Reducing from 1024→256 cuts text-encoding time significantly.

# ── [FIX 5] Style prefix replacing LoRA trigger word ──────────
# V1 had: "wablogo, logo, "   (FLUX LoRA activation token)
# V2 uses: descriptive style guide Z-Image understands natively
# "isolated on white background" → critical for Phase 3 RMBG removal
LOGO_STYLE_PREFIX = (
    "professional vector logo design, sharp clean lines, "
    "bold typography, isolated on white background, "
)

RMBG_HF_ID = "ZhengPeng7/BiRefNet_HR"

# ── Categories this pipeline handles ──────────────────────────
LOGO_LORA_CATEGORIES = {
    "admissions_enrollment",
    "openings_launches",
    "events",
    "education",
    "announcements",
    "business_promotions",
    "calls_to_action",
}

# ══════════════════════════════════════════════════════════════
# CONFIG — injected by inject_creds.py
# ══════════════════════════════════════════════════════════════
GOOGLE_CLIENT_ID     = os.environ.get("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET", "")
GOOGLE_REFRESH_TOKEN = os.environ.get("GOOGLE_REFRESH_TOKEN", "")
GITHUB_TOKEN         = os.environ.get("GITHUB_TOKEN_REPO2", "")
GITHUB_REPO2         = os.environ.get("GITHUB_REPO2", "")
GITHUB_REPO1         = os.environ.get("GITHUB_REPO", "")
GITHUB_TOKEN_REPO1   = os.environ.get("GITHUB_TOKEN_REPO1", "")
START_INDEX          = int(float(os.environ.get("START_INDEX", "0").strip()))
END_INDEX            = int(float(os.environ.get("END_INDEX", "200").strip()))

LOGO_LORA_CATEGORY   = os.environ.get("LOGO_LORA_CATEGORY", "").strip().lower()

SITE_URL       = "https://www.ultrapng.com"
SITE_NAME      = "UltraPNG"
WATERMARK_TEXT = "www.ultrapng.com"

# ══════════════════════════════════════════════════════════════
# PATHS  (identical to logo_lora_pipeline.py)
# ══════════════════════════════════════════════════════════════
WORKING_DIR     = Path("/kaggle/working")
GENERATED_DIR   = WORKING_DIR / "lora_generated"
APPROVED_DIR    = WORKING_DIR / "lora_approved"
TRANSPARENT_DIR = WORKING_DIR / "lora_transparent"
PROJECT_DIR     = WORKING_DIR / "project"
REPO2_DIR       = WORKING_DIR / "repo2"
REPO1_DIR       = WORKING_DIR / "repo1"
CHECKPOINT_DIR  = WORKING_DIR / "lora_checkpoints"

for d in [GENERATED_DIR, APPROVED_DIR, TRANSPARENT_DIR, CHECKPOINT_DIR]:
    d.mkdir(parents=True, exist_ok=True)

# ══════════════════════════════════════════════════════════════
# INSTALL DEPS
# ══════════════════════════════════════════════════════════════
print("=" * 56)
print("Installing Z-Image Pipeline V2.0 dependencies...")

# ── PyTorch P100 (sm_60) FIX ──────────────────────────────────
import subprocess as _sp, sys as _sys
_cap_check = _sp.run(
    [_sys.executable, "-c",
     "import torch; m=torch.cuda.get_device_capability(0)[0] "
     "if torch.cuda.is_available() else 99; print(m)"],
    capture_output=True, text=True)
_cuda_major = int(_cap_check.stdout.strip()) if _cap_check.stdout.strip().isdigit() else 99

if _cuda_major < 7:
    print(f"  [FIX] sm_{_cuda_major}0 GPU (P100) — reinstalling PyTorch cu126...")
    _r = _sp.run([_sys.executable, "-m", "pip", "install", "-q",
        "torch==2.6.0", "torchvision==0.21.0",
        "--index-url", "https://download.pytorch.org/whl/cu126"])
    if _r.returncode == 0:
        print("  [FIX] PyTorch cu126 installed ✓ — P100/sm_60 CUDA kernels work")
    else:
        print("  [WARN] PyTorch reinstall failed — may crash at inference!")
else:
    print(f"  [OK] sm_{_cuda_major}0 GPU — current PyTorch compatible")

PKGS = [
    "git+https://github.com/huggingface/diffusers.git",  # ZImagePipeline is in latest diffusers
    "transformers>=4.47.0",
    "accelerate>=0.28.0", "sentencepiece",
    "huggingface_hub>=0.23.0",
    "hf-transfer",          # [FIX TIMEOUT] Rust-based downloader — no ReadTimeout on large models
    "Pillow>=10.0", "numpy", "requests",
    "onnxruntime-gpu", "torchvision",
    "opencv-python-headless", "piexif",
    "openpyxl>=3.1.2",
    "bitsandbytes>=0.43.0",  # [PERF] 4-bit text encoder quantization → fits all on GPU
]
for pkg in PKGS:
    r = subprocess.run(
        [sys.executable, "-m", "pip", "install", "-q", pkg],
        capture_output=True, text=True)
    print(f"  {'OK  ' if r.returncode == 0 else 'WARN'} {pkg}")
print("Done!\n")

import torch
import numpy as np
from PIL import Image, ImageDraw, ImageFont
import requests as req

# ══════════════════════════════════════════════════════════════
# GLOBAL LOG CAPTURE
# ══════════════════════════════════════════════════════════════
_LOG_LINES = []

def log(msg):
    line = f"[{datetime.now().strftime('%H:%M:%S')}] {msg}"
    print(line, flush=True)
    _LOG_LINES.append(line)

# ══════════════════════════════════════════════════════════════
# UTILITIES
# ══════════════════════════════════════════════════════════════
def free_memory():
    gc.collect()
    if torch.cuda.is_available():
        torch.cuda.empty_cache()
        torch.cuda.synchronize()
    used = torch.cuda.memory_allocated(0) / 1e9 if torch.cuda.is_available() else 0
    log(f"  GPU used: {used:.1f}GB")

def slugify(s, max_len=60):
    s = re.sub(r'[^a-z0-9]+', '-', str(s).lower()).strip('-')
    if len(s) > max_len:
        cut = s[:max_len]
        idx = cut.rfind('-')
        s   = cut[:idx] if idx > max_len // 2 else cut
    return s or 'untitled'

def preview_url(fid, size=800):
    return f"https://lh3.googleusercontent.com/d/{fid}=s{size}"

def download_url(fid):
    return f"https://drive.usercontent.google.com/download?id={fid}&export=download&authuser=0"

# ══════════════════════════════════════════════════════════════
# CHECKPOINT SYSTEM
# ══════════════════════════════════════════════════════════════
def save_checkpoint(name, data):
    path = CHECKPOINT_DIR / f"{name}.json"
    tmp  = str(path) + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    os.replace(tmp, str(path))
    log(f"  Checkpoint saved: {name} ({len(data)} items)")

def load_checkpoint(name):
    path = CHECKPOINT_DIR / f"{name}.json"
    if path.exists():
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        log(f"  Checkpoint loaded: {name} ({len(data)} items)")
        return data
    return None

# ══════════════════════════════════════════════════════════════
# CATEGORY ENHANCERS  (identical to logo_lora_pipeline.py)
# ══════════════════════════════════════════════════════════════
CATEGORY_ENHANCERS = {
    "admissions_enrollment": ", academic excellence, institutional prestige",
    "openings_launches":     ", grand opening excitement, vibrant energy",
    "events":                ", event atmosphere, celebratory design",
    "education":             ", knowledge and learning, academic theme",
    "announcements":         ", clear bold communication, attention-grabbing",
    "business_promotions":   ", professional branding, corporate appeal",
    "calls_to_action":       ", urgency, action-oriented, bold typography",
}

def enhance_prompt(raw_prompt, category):
    cat = (category or "").lower()
    for key, extra in CATEGORY_ENHANCERS.items():
        if cat.startswith(key) or key in cat:
            return raw_prompt + extra
    return raw_prompt

# ══════════════════════════════════════════════════════════════
# GOOGLE DRIVE API  (identical to logo_lora_pipeline.py)
# ══════════════════════════════════════════════════════════════
_token_cache = {"value": None, "expires": 0}

def get_drive_token():
    if _token_cache["value"] and time.time() < _token_cache["expires"]:
        return _token_cache["value"]
    r = req.post("https://oauth2.googleapis.com/token", data={
        "client_id":     GOOGLE_CLIENT_ID,
        "client_secret": GOOGLE_CLIENT_SECRET,
        "refresh_token": GOOGLE_REFRESH_TOKEN,
        "grant_type":    "refresh_token",
    }, timeout=30)
    d = r.json()
    if "access_token" not in d:
        raise Exception(f"Token error: {d}")
    _token_cache.update({"value": d["access_token"], "expires": time.time() + 3200})
    return _token_cache["value"]

def drive_folder(token, name, parent=None):
    h = {"Authorization": f"Bearer {token}"}
    q = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    if parent:
        q += f" and '{parent}' in parents"
    r = req.get("https://www.googleapis.com/drive/v3/files",
                headers=h, params={"q": q, "fields": "files(id)"}, timeout=30)
    files = r.json().get("files", [])
    if files:
        return files[0]["id"]
    meta = {"name": name, "mimeType": "application/vnd.google-apps.folder"}
    if parent:
        meta["parents"] = [parent]
    return req.post(
        "https://www.googleapis.com/drive/v3/files",
        headers={**h, "Content-Type": "application/json"},
        json=meta, timeout=30).json()["id"]

def drive_upload(token, folder_id, name, data, mime="image/png", retries=3):
    for attempt in range(1, retries + 1):
        try:
            h        = {"Authorization": f"Bearer {token}"}
            metadata = json.dumps({"name": name, "parents": [folder_id]})
            b        = "----UltraPNGLoRAPipe"
            body = (
                f"--{b}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n{metadata}\r\n"
                f"--{b}\r\nContent-Type: {mime}\r\n\r\n"
            ).encode() + data + f"\r\n--{b}--".encode()
            r = req.post(
                "https://www.googleapis.com/upload/drive/v3/files"
                "?uploadType=multipart&fields=id,name",
                headers={**h, "Content-Type": f'multipart/related; boundary="{b}"'},
                data=body, timeout=120)
            if r.ok:
                return r.json()
            raise Exception(f"HTTP {r.status_code}: {r.text[:150]}")
        except Exception as e:
            if attempt < retries:
                log(f"  Upload retry {attempt}/{retries}: {e}")
                time.sleep(5 * attempt)
            else:
                raise

def drive_share(token, fid):
    try:
        req.post(
            f"https://www.googleapis.com/drive/v3/files/{fid}/permissions",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json={"role": "reader", "type": "anyone"}, timeout=30)
    except Exception:
        pass

# ══════════════════════════════════════════════════════════════
# WATERMARK PREVIEWS  (identical to logo_lora_pipeline.py)
# ══════════════════════════════════════════════════════════════
def make_previews(png_path):
    # JPEG generation removed — jpg_file_id was always "" and JPEG was never
    # uploaded to Drive.  Generating JPEG + piexif EXIF was ~1-2s wasted/image.
    # Returns (webp_bytes, w, h) only.
    with Image.open(png_path).convert("RGBA") as img_rgba:
        w, h = img_rgba.size
        if max(w, h) > 800:
            r        = 800 / max(w, h)
            img_rgba = img_rgba.resize((int(w * r), int(h * r)), Image.LANCZOS)
        w, h = img_rgba.size
        bg   = Image.new("RGB", (w, h), (255, 255, 255))
        drw  = ImageDraw.Draw(bg)
        for ry in range(0, h, 20):
            for cx in range(0, w, 20):
                if (ry // 20 + cx // 20) % 2 == 1:
                    drw.rectangle([cx, ry, cx + 20, ry + 20], fill=(232, 232, 232))
        bg.paste(img_rgba.convert("RGB"), mask=img_rgba.split()[3])
        try:
            fnt = ImageFont.truetype(
                "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", 13)
        except Exception:
            fnt = ImageFont.load_default()
        wm_layer = Image.new("RGBA", (w, h), (0, 0, 0, 0))
        wm_draw  = ImageDraw.Draw(wm_layer)
        for ry in range(-h, h + 110, 110):
            for cx in range(-w, w + 110, 110):
                wm_draw.text((cx, ry), WATERMARK_TEXT, fill=(0, 0, 0, 42), font=fnt)
        wm_rot  = wm_layer.rotate(-30, expand=False)
        bg_rgba = bg.convert("RGBA")
        bg_rgba.alpha_composite(wm_rot)
        bg   = bg_rgba.convert("RGB")
        drw2 = ImageDraw.Draw(bg)
        drw2.rectangle([0, h - 36, w, h], fill=(13, 13, 20))
        try:
            fnt2 = ImageFont.truetype(
                "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", 15)
        except Exception:
            fnt2 = fnt
        drw2.text((w // 2 - 72, h - 26), WATERMARK_TEXT, fill=(245, 166, 35), font=fnt2)
        webp_buf = io.BytesIO()
        bg.save(webp_buf, "WEBP", quality=82, method=6)
    return webp_buf.getvalue(), w, h

# ══════════════════════════════════════════════════════════════
# SKIP SET  (identical to logo_lora_pipeline.py)
# ══════════════════════════════════════════════════════════════
def load_skip_set_from_json():
    log("Building skip-set from REPO2 JSON files...")
    skip_set = set()
    if not REPO2_DIR.exists():
        try:
            repo_url = f"https://x-access-token:{GITHUB_TOKEN}@github.com/{GITHUB_REPO2}.git"
            subprocess.run(
                ["git", "clone", "--depth", "1", "--filter=blob:none",
                 "--sparse", repo_url, str(REPO2_DIR)],
                capture_output=True, check=True)
            subprocess.run(["git", "sparse-checkout", "init", "--cone"],
                           cwd=str(REPO2_DIR), capture_output=True, check=True)
            subprocess.run(["git", "sparse-checkout", "set", "data"],
                           cwd=str(REPO2_DIR), capture_output=True, check=True)
        except Exception:
            log("  Skip-set: REPO2 empty — starting fresh")
            return skip_set
    data_dir = REPO2_DIR / "data"
    if not data_dir.exists():
        log("  No data dir — starting fresh")
        return skip_set
    for jf in data_dir.glob("*.json"):
        if jf.name.startswith("_"):
            continue
        try:
            for e in json.loads(jf.read_text("utf-8")):
                fn = e.get("filename", "")
                if fn:
                    skip_set.add(fn)
        except Exception:
            pass
    log(f"  Skip-set (REPO2): {len(skip_set)} already-done filenames")
    return skip_set

def load_skip_set_from_ultradata() -> set:
    skip_set = set()
    token = GITHUB_TOKEN_REPO1 or ""
    repo  = (GITHUB_REPO1 or "").strip()
    if not token or not repo:
        log("  Skip-set (ultradata): GITHUB_TOKEN_REPO1 not set — skipping")
        return skip_set
    repo_url  = f"https://x-access-token:{token}@github.com/{repo}.git"
    xrepo_dir = WORKING_DIR / "repo1_skipcheck"
    try:
        if xrepo_dir.exists() and (xrepo_dir / ".git").exists():
            subprocess.run(["git", "remote", "set-url", "origin", repo_url],
                           cwd=str(xrepo_dir), capture_output=True)
            subprocess.run(["git", "pull", "--rebase", "--autostash"],
                           cwd=str(xrepo_dir), capture_output=True)
        else:
            shutil.rmtree(str(xrepo_dir), ignore_errors=True)
            subprocess.run(["git", "clone", "--depth", "1", repo_url, str(xrepo_dir)],
                           capture_output=True, check=True)
    except Exception as e:
        log(f"  Skip-set (ultradata): clone failed ({e}) — skipping")
        return skip_set
    xlsx_path = xrepo_dir / ULTRADATA_XLSX_NAME
    if not xlsx_path.exists():
        return skip_set
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
        ws = wb.active
        if ws.max_row < 2:
            return skip_set
        headers = [str(c.value or "").strip().lower()
                   for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if "filename" not in headers:
            return skip_set
        col_idx = headers.index("filename")
        for row in ws.iter_rows(min_row=2, values_only=True):
            fn = row[col_idx] if col_idx < len(row) else None
            if fn:
                skip_set.add(str(fn).strip())
        wb.close()
    except Exception as e:
        log(f"  Skip-set (ultradata): read failed ({e})")
    log(f"  Skip-set (ultradata): {len(skip_set)} already-generated filenames")
    return skip_set

# ══════════════════════════════════════════════════════════════
# LOAD PROMPTS — logo categories only
# ══════════════════════════════════════════════════════════════
def load_logo_prompts():
    log("Loading logo prompts...")
    if GITHUB_REPO1 and not PROJECT_DIR.exists():
        subprocess.run(
            ["git", "clone", "--depth", "1",
             f"https://github.com/{GITHUB_REPO1}", str(PROJECT_DIR)],
            capture_output=True)
    if PROJECT_DIR.exists():
        sys.path.insert(0, str(PROJECT_DIR))
        try:
            from prompts.prompt_engine import load_all_prompts
            all_prompts = load_all_prompts(str(PROJECT_DIR / "prompts" / "splits"))
            target_cats = {LOGO_LORA_CATEGORY} if LOGO_LORA_CATEGORY else LOGO_LORA_CATEGORIES
            logo_prompts = [
                p for p in all_prompts
                if p.get("category_slug", "").lower() in target_cats
            ]
            log(f"  Total prompts: {len(all_prompts)} | Logo filtered: {len(logo_prompts)}")
            return logo_prompts
        except Exception as e:
            log(f"Prompt error: {e}")
    log("FATAL: No prompts loaded!")
    return []

# ══════════════════════════════════════════════════════════════
# PHASE 1B — Z-Image-Turbo  (replaces FLUX.1-dev + Logo LoRA)
# ══════════════════════════════════════════════════════════════
def phase1b_generate_logo(batch, skip_set):
    ckpt = load_checkpoint("phase1b_logo_generated")
    if ckpt:
        return ckpt

    log("=" * 56)
    log("PHASE 1B: Z-Image-Turbo — Logo Generation (No LoRA)")
    log(f"  Model      : {ZIMAGE_HF_ID}")
    log(f"  Steps      : {ZIMAGE_STEPS}  (= 8 actual DiT forwards)")
    log(f"  Guidance   : {ZIMAGE_GUIDANCE}  (0.0 = required for distilled Turbo)")
    log(f"  MaxSeqLen  : {ZIMAGE_MAX_SEQ}  (256 = sufficient for logo prompts <80 tokens)")
    log(f"  Style pfx  : '{LOGO_STYLE_PREFIX}'")
    log("=" * 56)

    from diffusers import ZImagePipeline

    # ── Detect GPU(s) ──────────────────────────────────────────
    has_cuda   = torch.cuda.is_available()
    num_gpus   = torch.cuda.device_count() if has_cuda else 0
    vram_gb    = torch.cuda.get_device_properties(0).total_memory / 1e9 if has_cuda else 0
    gpu_name   = torch.cuda.get_device_name(0) if has_cuda else "CPU"
    cuda_cap   = torch.cuda.get_device_capability(0) if has_cuda else (0, 0)
    cuda_major = cuda_cap[0]

    log(f"  GPU count  : {num_gpus}")
    for _gi in range(num_gpus):
        _gp = torch.cuda.get_device_properties(_gi)
        log(f"  GPU[{_gi}]     : {_gp.name} | VRAM: {_gp.total_memory/1e9:.1f}GB")
    log(f"  CUDA cap   : sm_{cuda_major}{cuda_cap[1]}")

    # ── [FIX 1] Generator device ───────────────────────────────
    gen_device = "cuda" if has_cuda else "cpu"
    # For multi-GPU with device_map, generator uses cuda:0 (transformer device)
    if has_cuda and num_gpus >= 2:
        gen_device = "cuda:0"

    # ── GPU strategy: dtype + offload + multi-GPU decision ─────
    #
    # Z-Image-Turbo memory breakdown:
    #   Transformer (DiT) : ~12 GB
    #   Text encoder (LLM): ~7 GB  (Qwen3-4B)
    #   VAE                : ~0.5-1 GB
    #   Total              : ~19-20 GB
    #
    # ══════════════════════════════════════════════════════════
    # STRATEGY TABLE:
    # ══════════════════════════════════════════════════════════
    #  T4 x2 (2×16GB, sm_75):  FASTEST — split across 2 GPUs
    #    GPU:0 = transformer (12GB) + VAE (1GB) = 13GB ✓
    #    GPU:1 = text encoder full fp16 (7GB) ✓
    #    No offload, no quantization → maximum speed!
    #
    #  A100/H100 (≥40GB, sm_80+): bfloat16, all on one GPU
    #
    #  T4 single (16GB, sm_75): bfloat16 + cpu_offload
    #
    #  P100 (16GB, sm_60): float16 + 4-bit encoder + cpu_offload
    # ══════════════════════════════════════════════════════════

    use_multi_gpu = False   # T4 x2 dual-GPU flag

    if num_gpus >= 2 and cuda_major >= 7:
        # ── T4 x2 PATH: Split model across 2 GPUs — FASTEST ──
        total_vram = sum(
            torch.cuda.get_device_properties(i).total_memory / 1e9
            for i in range(num_gpus)
        )
        log(f"  Strategy   : T4 x2 MULTI-GPU — split across {num_gpus} GPUs")
        log(f"  Total VRAM : {total_vram:.0f}GB ({num_gpus}×{vram_gb:.0f}GB)")
        log(f"  GPU:0 → transformer (12GB) + VAE (1GB)")
        log(f"  GPU:1 → text_encoder Qwen3-4B full fp16 (7GB)")
        log(f"  No quantization, no offload → MAXIMUM SPEED!")
        dtype = torch.bfloat16   # T4 sm_75 supports bfloat16
        use_cpu_offload = False
        use_4bit_enc    = False
        use_multi_gpu   = True
    elif cuda_major >= 8:
        log("  Strategy   : bfloat16, no CPU offload (A100/H100 ≥40GB)")
        dtype = torch.bfloat16
        use_cpu_offload = False
        use_4bit_enc    = False
    elif cuda_major >= 7:
        log("  Strategy   : bfloat16 + enable_model_cpu_offload (T4 single, 16GB)")
        dtype = torch.bfloat16
        use_cpu_offload = True
        use_4bit_enc    = False
    else:
        log("  Strategy   : float16 + 4-bit Qwen encoder + cpu_offload (P100 sm_60)")
        log("  FIX: Qwen3-4B 4-bit = ~2GB → fast offload (2GB vs 7GB transfer)")
        log("  FIX: VAE forced float32 → prevents NaN invalid-cast RuntimeWarning")
        log("  FIX: cpu_offload ON → enc offloads after encode, frees VRAM for DiT inference")
        dtype = torch.float16
        use_cpu_offload = True
        use_4bit_enc    = True

    # ── [FIX TIMEOUT] Robust model download with retry ─────────
    # Root cause of ReadTimeout: default HF timeout is ~10s — too short
    # for a 19GB model. hf-transfer (installed above) eliminates this by
    # using a Rust-based multi-part downloader.  This retry loop is a
    # belt-and-suspenders fallback in case hf-transfer is unavailable.
    from huggingface_hub import snapshot_download
    _max_retries = 3
    _model_cache = None
    for _attempt in range(1, _max_retries + 1):
        try:
            # First try: use cache if already downloaded (fast path)
            _model_cache = snapshot_download(
                ZIMAGE_HF_ID,
                cache_dir=str(HF_CACHE),
                local_files_only=(_attempt == 1),  # attempt 1 = cache-only
                ignore_patterns=["*.msgpack", "*.h5", "flax_model*"],
            )
            log(f"  Model cache hit ✓  ({_model_cache})")
            break
        except Exception as _e:
            # Attempt 1 uses local_files_only=True — any failure means cache miss,
            # fall through to network download immediately (no sleep needed).
            # huggingface_hub raises LocalEntryNotFoundError or EnvironmentError —
            # neither contains the string "local_files_only", so we just check attempt==1.
            if _attempt == 1:
                pass  # cache miss → fall through to network download below
            elif _attempt < _max_retries:
                _wait = 10 * _attempt
                log(f"  Download attempt {_attempt} failed: {_e} — retrying in {_wait}s...")
                time.sleep(_wait)
            else:
                log(f"  Download attempt {_attempt} failed: {_e}")

        if _attempt > 1 or _model_cache is None:
            try:
                log(f"  Downloading {ZIMAGE_HF_ID} (attempt {_attempt})...")
                _model_cache = snapshot_download(
                    ZIMAGE_HF_ID,
                    cache_dir=str(HF_CACHE),
                    local_files_only=False,
                    ignore_patterns=["*.msgpack", "*.h5", "flax_model*"],
                )
                log(f"  Download complete ✓")
                break
            except Exception as _e2:
                if _attempt < _max_retries:
                    _wait = 15 * _attempt
                    log(f"  Attempt {_attempt} failed: {_e2} — retry in {_wait}s...")
                    time.sleep(_wait)
                else:
                    raise RuntimeError(
                        f"Failed to download {ZIMAGE_HF_ID} after {_max_retries} attempts. "
                        f"Last error: {_e2}"
                    ) from _e2

    # ── Load pipeline from local cache (no network call) ───────
    if use_multi_gpu:
        # ══════════════════════════════════════════════════════════
        # T4 x2 PATH: Use BOTH GPUs (32GB total) — FASTEST
        #
        # Official diffusers method: device_map="balanced"
        #   → Automatically splits pipeline components across GPUs
        #   → text_encoder on one GPU, transformer on another
        #   → No quantization, no offload → MAXIMUM SPEED
        #
        # Reference: huggingface.co/docs/diffusers/tutorials/inference_with_big_models
        # ══════════════════════════════════════════════════════════
        log("  Loading Z-Image-Turbo — T4 x2 MULTI-GPU (device_map=balanced)...")

        pipe = ZImagePipeline.from_pretrained(
            _model_cache,
            torch_dtype=dtype,
            local_files_only=True,
            device_map="balanced",
            max_memory={0: "15GB", 1: "15GB"},
        )

        # ── Log device placement ──
        if hasattr(pipe, 'hf_device_map'):
            log(f"  Device map : {pipe.hf_device_map}")
        for _gi in range(num_gpus):
            _used = torch.cuda.memory_allocated(_gi) / 1e9
            _total = torch.cuda.get_device_properties(_gi).total_memory / 1e9
            log(f"  GPU[{_gi}] VRAM: {_used:.1f}GB / {_total:.0f}GB "
                f"(free: {_total - _used:.1f}GB)")

        log("  device_map=balanced → both GPUs active ✓")
        log("  No quantization, no offload → MAXIMUM SPEED ✓\n")

    elif use_4bit_enc:
        # ── P100 PATH: 4-bit Qwen3 text encoder + cpu_offload ──────
        # 4-bit encoder: 7GB → 2GB.  With cpu_offload, encoder offloads
        # to CPU after encoding (~2GB transfer, fast).  Transformer then
        # has full 16GB for inference activations.
        #
        # Previous bug: loading everything on GPU without offload used 15.3GB,
        # leaving NO room for inference activations → OOM on every image.
        log("  Loading 4-bit quantized text encoder (Qwen3 NF4)...")
        _loaded_4bit = False
        try:
            from transformers import AutoModelForCausalLM, BitsAndBytesConfig as BnbCfg
            _bnb = BnbCfg(
                load_in_4bit=True,
                bnb_4bit_compute_dtype=torch.float16,   # P100: float16 only
                bnb_4bit_quant_type="nf4",
                bnb_4bit_use_double_quant=True,          # extra ~0.5GB saving
            )
            _enc_path = os.path.join(_model_cache, "text_encoder")
            _q_enc = AutoModelForCausalLM.from_pretrained(
                _enc_path,
                quantization_config=_bnb,
                torch_dtype=torch.float16,
                device_map={"": 0},       # force CUDA:0
                trust_remote_code=True,
            )
            log(f"  4-bit text encoder loaded ✓  "
                f"({sum(p.numel() for p in _q_enc.parameters())/1e9:.1f}B params)")

            # ── [FIX DOUBLE-LOAD] Pass pre-loaded 4-bit encoder into from_pretrained.
            # low_cpu_mem_usage=True skips the CPU staging → only 2GB used. ✓
            pipe = ZImagePipeline.from_pretrained(
                _model_cache,
                text_encoder=_q_enc,
                torch_dtype=dtype,
                local_files_only=True,
                low_cpu_mem_usage=True,    # [FIX] prevents 7GB CPU staging of Qwen
            )
            pipe.text_encoder = _q_enc
            log("  [FIX] pipe.text_encoder = 4-bit _q_enc (no double-load) ✓")

            # ── CRITICAL: VAE in float32 prevents NaN ─────────────────
            pipe.vae = pipe.vae.to(torch.float32)
            log("  VAE set to float32 → NaN-safe decode ✓")

            # ── [FIX OOM] Use cpu_offload even with 4-bit encoder ─────
            # Previous code loaded all on GPU (15.3GB) → no room for activations.
            # With cpu_offload: encoder (2GB) offloads after encode step,
            # transformer (12GB) gets full VRAM for inference.
            # 4-bit offload is fast: only 2GB moves vs 7GB with full encoder.
            pipe.enable_model_cpu_offload()
            log("  enable_model_cpu_offload() applied ✓ (4-bit = fast 2GB offload)")

            _loaded_4bit = True
            vram_used = torch.cuda.memory_allocated(0)/1e9
            vram_total = torch.cuda.get_device_properties(0).total_memory/1e9
            log(f"  VRAM after load: {vram_used:.1f}GB / {vram_total:.0f}GB")
            log(f"  Free for inference: ~{vram_total - vram_used:.1f}GB ✓")

        except Exception as _qe:
            log(f"  4-bit load failed ({_qe}) — falling back to full cpu_offload")
            _loaded_4bit = False

        if not _loaded_4bit:
            # Fallback: original cpu_offload path (full 7GB encoder)
            pipe = ZImagePipeline.from_pretrained(
                _model_cache, torch_dtype=dtype, local_files_only=True)
            pipe.enable_model_cpu_offload()
            log("  enable_model_cpu_offload() applied (fallback) ✓")
    else:
        pipe = ZImagePipeline.from_pretrained(
            _model_cache,          # local path — guaranteed no timeout
            torch_dtype=dtype,
            local_files_only=True,
        )
        if cuda_major >= 8 and not use_cpu_offload:
            pipe.to("cuda")

        # Apply CPU offload BEFORE set_progress_bar_config
        # [FIX 6] Order matters: offload must be set before any pipe call
        if use_cpu_offload:
            pipe.enable_model_cpu_offload()
            log("  enable_model_cpu_offload() applied ✓")

    pipe.set_progress_bar_config(disable=True)

    log(f"  Ready! Batch: {len(batch)} | Skip: {len(skip_set)}\n")

    generated, skipped, t0 = [], 0, time.time()
    _consecutive_oom = 0   # OOM counter — abort if GPU is stuck
    _oom_downscaled  = False  # Track if we had to drop to 768x768

    for i, item in enumerate(batch):
        fname = item["filename"]

        if fname in skip_set:
            skipped += 1
            if skipped <= 3 or skipped % 50 == 0:
                log(f"  SKIP [{i+1}/{len(batch)}] {fname}")
            continue

        try:
            out_dir = GENERATED_DIR / item["category"] / item.get("subcategory", "general")
            out_dir.mkdir(parents=True, exist_ok=True)
            out = out_dir / fname

            if out.exists():
                generated.append({"path": str(out), "item": item})
                _consecutive_oom = 0
                continue

            # [FIX 1] Use cuda generator (Z-Image official code uses "cuda")
            gen = torch.Generator(gen_device).manual_seed(item["seed"])

            # ── Style prefix prepended (replaces V1 LoRA trigger word) ──
            raw    = enhance_prompt(item["prompt"], item.get("category", ""))
            prompt = LOGO_STYLE_PREFIX + raw

            # ── [FIX OOM] Determine resolution ──────────────────────
            # Start at 1024. If previous OOM forced downscale, stay at 768.
            img_size = 768 if _oom_downscaled else 1024

            # ── [FIX] Wrap in no_grad to reduce activation memory ───
            with torch.no_grad():
                try:
                    output = pipe(
                        prompt=prompt,
                        num_inference_steps=ZIMAGE_STEPS,
                        guidance_scale=ZIMAGE_GUIDANCE,
                        height=img_size, width=img_size,
                        generator=gen,
                        max_sequence_length=ZIMAGE_MAX_SEQ,
                    )
                    img = output.images[0]
                except torch.cuda.OutOfMemoryError:
                    # ── OOM at current size → retry at 768x768 ────────
                    torch.cuda.empty_cache()
                    gc.collect()
                    if img_size > 768:
                        log(f"  OOM at {img_size}x{img_size} → retrying at 768x768: {fname}")
                        _oom_downscaled = True
                        img_size = 768
                        gen = torch.Generator(gen_device).manual_seed(item["seed"])
                        output = pipe(
                            prompt=prompt,
                            num_inference_steps=ZIMAGE_STEPS,
                            guidance_scale=ZIMAGE_GUIDANCE,
                            height=768, width=768,
                            generator=gen,
                            max_sequence_length=ZIMAGE_MAX_SEQ,
                        )
                        img = output.images[0]
                    else:
                        raise  # already at 768, re-raise OOM

            # ── Upscale 768→1024 if downscaled (consistent output size) ──
            if img_size < 1024:
                img = img.resize((1024, 1024), Image.LANCZOS)

            # ── Blank / NaN detection ────────────────────────────────
            arr    = np.array(img, dtype=np.float32)
            n_px   = arr.shape[0] * arr.shape[1]
            thresh = int(n_px * 0.005)
            is_blank = (
                arr.std() < 3.0 or
                arr.mean() < 5.0 or
                (arr < 240).sum() < thresh
            )
            if is_blank:
                log(f"  Retry (blank): {fname}  std={arr.std():.1f}  "
                    f"mean={arr.mean():.0f}  non-white={(arr<240).sum()}")
                with torch.no_grad():
                    gen2 = torch.Generator(gen_device).manual_seed(item["seed"] + 99)
                    img  = pipe(
                        prompt=prompt,
                        num_inference_steps=ZIMAGE_STEPS,
                        guidance_scale=ZIMAGE_GUIDANCE,
                        height=img_size, width=img_size,
                        generator=gen2,
                        max_sequence_length=ZIMAGE_MAX_SEQ,
                    ).images[0]
                    if img_size < 1024:
                        img = img.resize((1024, 1024), Image.LANCZOS)

            img.save(str(out), "PNG", compress_level=1)
            generated.append({"path": str(out), "item": item})
            _consecutive_oom = 0   # reset on success

            done = len(generated)
            rate = done / (time.time() - t0)
            eta  = (len(batch) - i - 1) / rate / 60 if rate > 0 else 0
            res_tag = f" [768→1024]" if img_size < 1024 else ""
            log(f"  [{i+1}/{len(batch)}] OK {fname} | {item['category']}{res_tag} | ETA {eta:.0f}min")

        except torch.cuda.OutOfMemoryError:
            torch.cuda.empty_cache()
            gc.collect()
            _consecutive_oom += 1
            vram_used = torch.cuda.memory_allocated(0) / 1e9 if torch.cuda.is_available() else 0
            log(f"  OOM [{_consecutive_oom}]: {fname} — skipping  "
                f"(VRAM used after clear: {vram_used:.1f}GB)")
            if _consecutive_oom >= 5:
                log("  5 consecutive OOMs — GPU stuck. Saving checkpoint and aborting phase.")
                break   # save what we have rather than hanging forever
        except Exception as e:
            log(f"  FAIL {fname}: {e}")

    log("\n  Deleting Z-Image-Turbo → freeing VRAM + disk cache...")
    del pipe
    free_memory()

    # [FIX 7] Correct HF cache folder pattern for Tongyi-MAI/Z-Image-Turbo
    # HF saves as: models--Tongyi-MAI--Z-Image-Turbo  (lowercased on disk)
    import shutil as _shutil
    for _cache_sub in HF_CACHE.iterdir():
        if _cache_sub.is_dir():
            _name = _cache_sub.name.lower()
            if ("tongyi" in _name or "z-image" in _name or
                    "z_image" in _name or "zimage" in _name):
                _shutil.rmtree(str(_cache_sub), ignore_errors=True)
                log(f"  Deleted cache: {_cache_sub.name}")
    _used = sum(f.stat().st_size for f in HF_CACHE.rglob("*") if f.is_file()) / 1e9
    log(f"  HF cache after cleanup: {_used:.1f}GB")

    save_checkpoint("phase1b_logo_generated", generated)
    log(f"PHASE 1B DONE — Generated: {len(generated)} | Skipped: {skipped}\n")
    return generated

# ══════════════════════════════════════════════════════════════
# PHASE 3 + 4 — BG Removal + Drive Upload (batched every 50)
#
# WHY COMBINED?
#   Old flow: generate ALL → remove BG all → upload all
#   New flow: generate ALL → for every 50: remove BG + upload immediately
#
# BENEFITS:
#   1. Images appear on Drive after every 50 — no need to wait till end
#   2. Kaggle session crash → only lose last partial batch (not everything)
#   3. RMBG model stays loaded the whole time (no reload per batch)
#   4. Token refresh happens per-batch → no OAuth expiry on large runs
#
# CHECKPOINT STRATEGY:
#   "phase3_4_uploaded" checkpoint stores all uploaded posts so far.
#   On resume: already-uploaded filenames are skipped in RMBG + upload.
# ══════════════════════════════════════════════════════════════
DRIVE_BATCH_SIZE = 50   # Push to Drive after every N images

def phase3_4_batched(posts):
    # ── Resume: load already-uploaded posts from checkpoint ────
    ckpt = load_checkpoint("phase3_4_uploaded")
    if ckpt is not None:
        # Check if everything is already done
        uploaded_fnames = {p["filename"] for p in ckpt}
        remaining = [p for p in posts if p["filename"] not in uploaded_fnames]
        if not remaining:
            log(f"PHASE 3+4: All {len(ckpt)} already uploaded — skipping.")
            return ckpt
        log(f"PHASE 3+4: Resuming — {len(ckpt)} done, {len(remaining)} remaining")
        all_uploaded = list(ckpt)
        posts_to_do  = remaining
    else:
        all_uploaded = []
        posts_to_do  = posts

    if not posts_to_do:
        return all_uploaded

    log("=" * 56)
    log("PHASE 3+4: BiRefNet_HR BG Removal → Drive Upload")
    log(f"  Model      : {RMBG_HF_ID}")
    log(f"  Total      : {len(posts_to_do)} images")
    log(f"  Batch size : {DRIVE_BATCH_SIZE} (Drive push every {DRIVE_BATCH_SIZE} images)")
    log("=" * 56)

    # ── Load RMBG model ONCE for the whole run ─────────────────
    from torchvision import transforms
    from transformers import AutoModelForImageSegmentation

    rmbg_model = AutoModelForImageSegmentation.from_pretrained(
        RMBG_HF_ID,
        trust_remote_code=True,
        cache_dir=str(HF_CACHE),
    )
    device = "cuda" if torch.cuda.is_available() else "cpu"
    torch.set_float32_matmul_precision("high")
    rmbg_model = rmbg_model.to(device).eval().half()
    log(f"  BiRefNet_HR on {device.upper()} | FP16 | 2048x2048\n")

    transform_img = transforms.Compose([
        transforms.Resize((2048, 2048)),
        transforms.ToTensor(),
        transforms.Normalize([0.485, 0.456, 0.406],
                             [0.229, 0.224, 0.225]),
    ])

    def remove_bg(pil_img):
        ow, oh   = pil_img.size
        inp      = transform_img(pil_img.convert("RGB")).unsqueeze(0).to(device).half()
        with torch.no_grad():
            preds = rmbg_model(inp)[-1].sigmoid().cpu()
        pred     = preds[0].squeeze()
        mask_pil = transforms.ToPILImage()(pred).resize((ow, oh), Image.LANCZOS)
        result   = pil_img.convert("RGBA")
        result.putalpha(mask_pil)
        return result

    # ── Pre-fetch Drive folder IDs once — avoid repeated API calls ──
    token    = get_drive_token()
    png_root = drive_folder(token, "png_library_images")
    prv_root = drive_folder(token, "png_library_previews")
    # Cache folder IDs: (root_id, cat, sub) → folder_id
    _folder_cache: dict = {}

    def get_folders(cat, sub):
        key = (cat, sub)
        if key not in _folder_cache:
            _folder_cache[key] = {
                "png": drive_folder(token, sub,
                         drive_folder(token, cat, png_root)),
                "prv": drive_folder(token, sub,
                         drive_folder(token, cat, prv_root)),
            }
        return _folder_cache[key]["png"], _folder_cache[key]["prv"]

    def upload_one(post):
        """BG-remove + upload one post. Returns enriched post dict or None."""
        path  = Path(post["approved_path"])
        rel   = path.relative_to(APPROVED_DIR)
        out   = TRANSPARENT_DIR / rel
        out.parent.mkdir(parents=True, exist_ok=True)

        # ── BG removal ──
        if not out.exists():
            img    = Image.open(str(path)).convert("RGB")
            result = remove_bg(img)
            result.save(str(out), "PNG", compress_level=9)

        # ── Drive upload ──
        cat   = post.get("category",    "general")
        sub   = post.get("subcategory", "general")
        fname = post["filename"]

        png_folder, prv_folder = get_folders(cat, sub)

        png_data  = out.read_bytes()
        png_resp  = drive_upload(token, png_folder, fname, png_data, "image/png")
        png_fid   = png_resp["id"]
        drive_share(token, png_fid)

        webp_bytes, pw, ph = make_previews(out)
        webp_name = Path(fname).stem + ".webp"
        prv_resp  = drive_upload(token, prv_folder, webp_name, webp_bytes, "image/webp")
        webp_fid  = prv_resp["id"]
        drive_share(token, webp_fid)

        return {
            **post,
            "transparent_path":  str(out),
            "png_file_id":       png_fid,
            "jpg_file_id":       "",
            "webp_file_id":      webp_fid,
            "download_url":      download_url(png_fid),
            "preview_url":       preview_url(webp_fid, 800),
            "preview_url_small": preview_url(webp_fid, 400),
            "webp_preview_url":  preview_url(webp_fid, 800),
            "preview_w":         pw,
            "preview_h":         ph,
        }

    # ── Main loop: process + upload in batches of DRIVE_BATCH_SIZE ──
    t0          = time.time()
    batch_buf   = []    # accumulates uploaded posts in current batch
    batch_start = 0

    for i, post in enumerate(posts_to_do):
        try:
            enriched = upload_one(post)
            if enriched:
                all_uploaded.append(enriched)
                batch_buf.append(enriched)

            rate = (i + 1) / (time.time() - t0)
            eta  = (len(posts_to_do) - i - 1) / rate / 60 if rate > 0 else 0
            log(f"  [{i+1}/{len(posts_to_do)}] ✓ {post['filename']} | ETA {eta:.0f}min")

        except Exception as e:
            log(f"  FAIL {post.get('filename','?')}: {e}")

        # ── Every DRIVE_BATCH_SIZE images → checkpoint + xlsx push + token refresh ──
        if len(batch_buf) >= DRIVE_BATCH_SIZE:
            batch_end = batch_start + len(batch_buf)
            save_checkpoint("phase3_4_uploaded", all_uploaded)
            log(f"\n  ── BATCH CHECKPOINT [{batch_start}–{batch_end}]: "
                f"{len(all_uploaded)} total uploaded to Drive ──")

            # ── [DATA SAFETY] Push xlsx every batch ──────────────────
            # If Kaggle session dies before main() xlsx push, only the last
            # partial batch is missing from xlsx.  Skip set (built from xlsx)
            # on next session will correctly avoid re-generating already-done images.
            try:
                log(f"  Appending {len(batch_buf)} rows to ultradata.xlsx...")
                _append_ultradata_and_push(batch_buf)
                log(f"  xlsx batch push OK ✓")
            except Exception as _xe:
                log(f"  xlsx batch push WARN (non-fatal): {_xe}")

            batch_buf   = []
            batch_start = batch_end
            log("")
            # Refresh OAuth token to prevent expiry on long runs (token lasts ~55min)
            try:
                token = get_drive_token()
            except Exception as _te:
                log(f"  Token refresh warning: {_te}")

    # ── Final checkpoint for any remaining images ───────────────
    if batch_buf:
        save_checkpoint("phase3_4_uploaded", all_uploaded)
        log(f"\n  ── FINAL BATCH CHECKPOINT: {len(all_uploaded)} total uploaded ──")
        try:
            log(f"  Appending final {len(batch_buf)} rows to ultradata.xlsx...")
            _append_ultradata_and_push(batch_buf)
            log(f"  xlsx final batch push OK ✓\n")
        except Exception as _xe:
            log(f"  xlsx final batch push WARN (non-fatal): {_xe}\n")

    # ── Unload RMBG + cleanup ───────────────────────────────────
    log("  Deleting BiRefNet_HR → freeing VRAM...")
    del rmbg_model
    free_memory()

    import shutil as _shutil
    for _cache_sub in HF_CACHE.iterdir():
        if _cache_sub.is_dir():
            _name = _cache_sub.name.lower()
            if "rmbg" in _name or "briaai" in _name or "birefnet" in _name:
                _shutil.rmtree(str(_cache_sub), ignore_errors=True)
                log(f"  Deleted BiRefNet cache: {_cache_sub.name}")

    if APPROVED_DIR.exists():
        _shutil.rmtree(str(APPROVED_DIR), ignore_errors=True)
        APPROVED_DIR.mkdir(parents=True, exist_ok=True)
        log("  Deleted lora_approved/ images (transparent copies kept)")

    log(f"PHASE 3+4 DONE — BG removed + Uploaded: {len(all_uploaded)}\n")
    return all_uploaded


# ── Backwards-compat shims (old checkpoint names still work on resume) ──
def phase3_bg_remove(posts):
    """Legacy shim — redirects to phase3_4_batched."""
    return phase3_4_batched(posts)

def phase4_upload(posts):
    """Legacy shim — phase4 is now merged into phase3_4_batched."""
    log("phase4_upload: merged into phase3_4_batched — no-op")
    return posts

# ══════════════════════════════════════════════════════════════
# PHASE 5 — Append ultradata.xlsx → push REPO1  [IDENTICAL]
# ══════════════════════════════════════════════════════════════
def _append_ultradata_and_push(items):
    token = GITHUB_TOKEN_REPO1 or ""
    repo  = (GITHUB_REPO1 or "").strip()
    if not token or not repo:
        raise Exception("Missing GITHUB_TOKEN_REPO1 or GITHUB_REPO (repo1)")

    repo_url  = f"https://x-access-token:{token}@github.com/{repo}.git"
    xrepo_dir = WORKING_DIR / "repo1_xlsx"

    if xrepo_dir.exists() and (xrepo_dir / ".git").exists():
        subprocess.run(["git", "remote", "set-url", "origin", repo_url],
                       cwd=str(xrepo_dir), capture_output=True)
        subprocess.run(["git", "pull", "--rebase", "--autostash"],
                       cwd=str(xrepo_dir), capture_output=True)
    else:
        shutil.rmtree(str(xrepo_dir), ignore_errors=True)
        subprocess.run(["git", "clone", "--depth", "1", repo_url, str(xrepo_dir)],
                       capture_output=True, check=True)

    try:
        import openpyxl
        from openpyxl import Workbook
    except Exception:
        subprocess.run([sys.executable, "-m", "pip", "install", "-q", "openpyxl>=3.1.2"],
                       capture_output=True, check=True)
        import openpyxl
        from openpyxl import Workbook

    xlsx_path = xrepo_dir / ULTRADATA_XLSX_NAME
    headers   = [
        "date_added", "subject_name", "category", "subcategory", "filename",
        "png_file_id", "webp_file_id", "download_url", "preview_url"
    ]

    if xlsx_path.exists():
        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)

    def _row(it):
        return [
            it.get("date_added", ""),
            it.get("subject_name", ""),
            it.get("category", ""),
            it.get("subcategory", ""),
            it.get("filename", ""),
            it.get("png_file_id", ""),
            it.get("webp_file_id", ""),
            it.get("download_url", ""),
            it.get("preview_url", ""),
        ]

    appended = 0
    for it in items:
        if not it.get("subject_name") or not it.get("download_url") or not it.get("preview_url"):
            continue
        ws.append(_row(it))
        appended += 1

    wb.save(str(xlsx_path))

    if appended == 0:
        log("ultradata.xlsx: nothing new to append")
        return

    orig_dir = os.getcwd()
    try:
        os.chdir(str(xrepo_dir))
        subprocess.run(["git", "config", "user.name", "github-actions[bot]"],
                       check=True, capture_output=True)
        subprocess.run(["git", "config", "user.email",
                        "github-actions[bot]@users.noreply.github.com"],
                       check=True, capture_output=True)
        subprocess.run(["git", "add", ULTRADATA_XLSX_NAME], check=True, capture_output=True)
        diff = subprocess.run(["git", "diff", "--staged", "--quiet"], capture_output=True)
        if diff.returncode != 0:
            msg = f"ultradata: zimage-logo append {appended} rows [{datetime.now().strftime('%Y-%m-%d')}]"
            subprocess.run(["git", "commit", "-m", msg], check=True, capture_output=True)
            pushed = False
            for push_attempt in range(1, 5):
                result = subprocess.run(["git", "push"], capture_output=True, text=True)
                if result.returncode == 0:
                    pushed = True
                    break
                log(f"  Push attempt {push_attempt}/4 failed — pulling rebase...")
                subprocess.run(["git", "pull", "--rebase", "--autostash"],
                               capture_output=True, cwd=str(xrepo_dir))
                time.sleep(5 * push_attempt)
            if not pushed:
                raise Exception(f"git push failed after 4 attempts:\n{result.stderr.strip()}")
            log(f"ultradata.xlsx pushed: +{appended} rows")
        else:
            log("ultradata.xlsx: no staged changes")
    finally:
        os.chdir(orig_dir)

# ══════════════════════════════════════════════════════════════
# PHASE 6 — Save Logs → REPO1  [IDENTICAL]
# ══════════════════════════════════════════════════════════════
def phase6_save_logs(stats: dict):
    token = GITHUB_TOKEN_REPO1 or ""
    repo  = (GITHUB_REPO1 or "").strip()
    if not token or not repo:
        log("  Logs: GITHUB_TOKEN_REPO1 not set — skipping log push")
        return

    ts        = datetime.now().strftime("%Y-%m-%d_%H-%M")
    log_name  = f"lora_{ts}.log"
    repo_url  = f"https://x-access-token:{token}@github.com/{repo}.git"
    xrepo_dir = WORKING_DIR / "repo1_logs"

    if xrepo_dir.exists() and (xrepo_dir / ".git").exists():
        subprocess.run(["git", "remote", "set-url", "origin", repo_url],
                       cwd=str(xrepo_dir), capture_output=True)
        subprocess.run(["git", "pull", "--rebase", "--autostash"],
                       cwd=str(xrepo_dir), capture_output=True)
    else:
        shutil.rmtree(str(xrepo_dir), ignore_errors=True)
        subprocess.run(["git", "clone", "--depth", "1", repo_url, str(xrepo_dir)],
                       capture_output=True, check=True)

    logs_dir = xrepo_dir / "logs"
    logs_dir.mkdir(exist_ok=True)

    summary = "\n".join([
        "=" * 56,
        "Z-IMAGE LOGO PIPELINE V2.0 — RUN SUMMARY",
        f"  Date       : {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"  Status     : {stats.get('status', '?')}",
        f"  Duration   : {stats.get('duration', '?')}",
        f"  Generated  : {stats.get('generated', 0)}",
        f"  Transparent: {stats.get('transparent', 0)}",
        f"  Uploaded   : {stats.get('uploaded', 0)}",
        f"  Style pfx  : {LOGO_STYLE_PREFIX}",
        f"  Model      : {ZIMAGE_HF_ID}",
        f"  Steps      : {ZIMAGE_STEPS}",
        f"  Guidance   : {ZIMAGE_GUIDANCE}",
        f"  MaxSeqLen  : {ZIMAGE_MAX_SEQ}",
        "=" * 56,
        "",
    ] + _LOG_LINES)

    (logs_dir / log_name).write_text(summary, encoding="utf-8")
    (logs_dir / "latest_lora.log").write_text(summary, encoding="utf-8")

    orig_dir = os.getcwd()
    try:
        os.chdir(str(xrepo_dir))
        subprocess.run(["git", "config", "user.name", "github-actions[bot]"],
                       capture_output=True)
        subprocess.run(["git", "config", "user.email",
                        "github-actions[bot]@users.noreply.github.com"],
                       capture_output=True)
        subprocess.run(["git", "add", "logs/"], capture_output=True)
        subprocess.run(["git", "commit", "-m", f"zimage-logo log: {ts}"],
                       capture_output=True)
        subprocess.run(["git", "push"], capture_output=True)
        log(f"  Log saved: logs/{log_name}")
    except Exception as e:
        log(f"  Log push error (non-fatal): {e}")
    finally:
        os.chdir(orig_dir)

# ══════════════════════════════════════════════════════════════
# MAIN  [IDENTICAL to logo_lora_pipeline.py]
# ══════════════════════════════════════════════════════════════
def main():
    t0    = time.time()
    stats = {"status": "running", "generated": 0,
             "transparent": 0, "uploaded": 0, "duration": "?"}

    print("╔══════════════════════════════════════════════════════╗")
    print("║  UltraPNG — Z-Image Logo Pipeline V2.0              ║")
    print("║  Z-Image-Turbo → RMBG → Drive → xlsx               ║")
    print("╚══════════════════════════════════════════════════════╝")
    print(f"  Batch    : {START_INDEX} → {END_INDEX} ({END_INDEX-START_INDEX} prompts)")
    print(f"  StylePfx : {LOGO_STYLE_PREFIX}")
    print(f"  Steps    : {ZIMAGE_STEPS}")
    print(f"  Filter   : {LOGO_LORA_CATEGORY or 'ALL logo categories'}")
    if torch.cuda.is_available():
        _ngpu = torch.cuda.device_count()
        for _gi in range(_ngpu):
            p = torch.cuda.get_device_properties(_gi)
            print(f"  GPU[{_gi}]    : {p.name} | VRAM: {p.total_memory/1e9:.0f}GB")
        if _ngpu >= 2:
            print(f"  Mode     : T4 x2 MULTI-GPU (split across {_ngpu} GPUs)")
    print()

    try:
        prompts = load_logo_prompts()
        if not prompts:
            raise Exception("No logo prompts loaded!")

        skip_set  = load_skip_set_from_json()
        skip_set |= load_skip_set_from_ultradata()
        log(f"  Combined skip-set: {len(skip_set)} filenames\n")

        # ── AUTO-ADVANCE (identical to logo_lora_pipeline.py) ──
        BATCH_SIZE = END_INDEX - START_INDEX
        real_start = START_INDEX
        real_end   = END_INDEX
        MAX_SCAN   = min(len(prompts), START_INDEX + max(BATCH_SIZE * 50, 5000))

        scan_start = START_INDEX
        while scan_start < MAX_SCAN:
            window = prompts[scan_start:scan_start + BATCH_SIZE]
            fresh  = [p for p in window if p["filename"] not in skip_set]
            if fresh:
                if scan_start != START_INDEX:
                    log(f"  AUTO-ADVANCE: {START_INDEX} → {scan_start}")
                real_start = scan_start
                real_end   = scan_start + BATCH_SIZE
                break
            scan_start += BATCH_SIZE
        else:
            log("  All logo prompts already done — nothing to generate.")
            return

        batch = prompts[real_start:real_end]
        log(f"Batch: {real_start} → {real_end} | "
            f"{len([p for p in batch if p['filename'] not in skip_set])} fresh\n")

        os.environ["START_INDEX"] = str(real_start)
        os.environ["END_INDEX"]   = str(real_end)

        # ── PHASE 1B ──
        generated = phase1b_generate_logo(batch, skip_set)
        stats["generated"] = len(generated)
        if not generated:
            log("No logo images generated."); return

        # ── Build posts list (identical schema to logo_lora_pipeline.py) ──
        posts = []
        for gd in generated:
            item    = gd["item"]
            src     = Path(gd["path"])
            rel     = src.relative_to(GENERATED_DIR)
            dst     = APPROVED_DIR / rel
            dst.parent.mkdir(parents=True, exist_ok=True)
            if not dst.exists():
                shutil.copy2(str(src), str(dst))
            subject = (item.get("subject_name") or item.get("subject") or
                       item.get("name") or item.get("title") or
                       item.get("subcategory") or item.get("category") or "Untitled")
            subject = str(subject).replace("_", " ").replace("-", " ").title()
            posts.append({
                "category":       item.get("category", "general"),
                "subcategory":    item.get("subcategory", "general"),
                "subject_name":   subject,
                "filename":       item.get("filename", src.name),
                "original_prompt": item.get("prompt", ""),
                "approved_path":  str(dst),
                "slug":           slugify(subject),
                "title": "", "h1": "", "meta_desc": "", "alt_text": "",
                "tags": "", "description": "", "word_count": 0,
                "ai_generated": False, "qwen_confidence": 0, "qwen_group": "",
                "png_file_id": "", "jpg_file_id": "", "webp_file_id": "",
                "download_url": "", "preview_url": "", "preview_url_small": "",
                "webp_preview_url": "", "preview_w": 800, "preview_h": 800,
                "date_added": datetime.now().strftime("%Y-%m-%d"),
            })
        if not posts:
            log("No posts to process."); return

        # ── PHASE 3 + 4 (combined — BG removal + Drive upload every 50) ──
        uploaded = phase3_4_batched(posts)
        stats["transparent"] = len(uploaded)
        stats["uploaded"]    = len(uploaded)
        if not uploaded:
            log("BG removal / Drive upload produced no results."); return

        # ── PHASE 5 — Final xlsx push (catches any remainder not pushed in batches) ──
        # Note: phase3_4_batched already pushes xlsx every 50 images.
        # This final call is a safety net for the last partial batch if it
        # wasn't pushed (e.g. batch_buf had items but batch checkpoint didn't fire).
        try:
            _append_ultradata_and_push(uploaded)
        except Exception as _xe:
            log(f"  Final xlsx push warning (data already pushed in batches): {_xe}")

        for ck in CHECKPOINT_DIR.glob("*.json"):
            ck.unlink()

        eff_end = int(os.environ.get("END_INDEX", str(END_INDEX)))
        (WORKING_DIR / "effective_end.txt").write_text(str(eff_end))
        log(f"  effective_end.txt written: {eff_end}")

        hrs = (time.time() - t0) / 3600
        stats.update({"duration": f"{hrs:.1f}h", "status": "SUCCESS"})

        print(f"\n╔══════════════════════════════════════════════════════╗")
        print(f"║  Z-IMAGE LOGO DONE in {hrs:.1f}h")
        print(f"║  Gen:{len(generated)}  BG+Up:{len(uploaded)}")
        print(f"╚══════════════════════════════════════════════════════╝")

    except Exception as e:
        stats["status"] = f"FAILED: {e}"
        log(f"FATAL: {e}")
        import traceback; traceback.print_exc()
        raise

    finally:
        phase6_save_logs(stats)


if __name__ == "__main__":
    main()
