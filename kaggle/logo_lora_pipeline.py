"""
╔══════════════════════════════════════════════════════════════╗
║   UltraPNG.com — Logo LoRA Pipeline V1.1                    ║
╠══════════════════════════════════════════════════════════════╣
║  PHASE 1B → FLUX.1-dev + Logo LoRA   1024x1024 logos       ║
║  GPU AUTO-DETECT: NF4 quant on T4/P100 | fp16 on A100      ║
║  PHASE 3  → RMBG-2.0 ONNX           Background removal     ║
║  PHASE 4  → Google Drive             Upload PNG + WebP      ║
║  PHASE 5  → ultradata.xlsx           Append rows in REPO1   ║
║  PHASE 6  → Save Run Logs → REPO1   (visible on GitHub!)   ║
╚══════════════════════════════════════════════════════════════╝

STANDALONE FILE — Does NOT import from main_pipeline.py
inject_creds.py prepends os.environ[] lines before this file.

TRIGGER WORD:  "wablogo, logo, "  (auto-prepended to every prompt)

TARGET CATEGORIES (set LOGO_LORA_CATEGORY env var to pick one):
  admissions_enrollment | openings_launches | events
  education | announcements | business_promotions
  calls_to_action | (leave empty = run ALL logo categories)
"""

import os, sys, json, time, gc, re, io, shutil, base64, subprocess, math
from pathlib import Path
from datetime import datetime

# ── Force real-time log output in Kaggle ──────────────────────
try:
    sys.stdout.reconfigure(line_buffering=True)
    sys.stderr.reconfigure(line_buffering=True)
except AttributeError:
    pass

os.environ["PYTORCH_CUDA_ALLOC_CONF"] = "expandable_segments:True"
ULTRADATA_XLSX_NAME = "ultradata.xlsx"

# ── HuggingFace cache dir ──────────────────────────────────────
HF_CACHE = Path("/kaggle/working/hf_cache")
HF_CACHE.mkdir(parents=True, exist_ok=True)
os.environ["HF_HOME"]               = str(HF_CACHE)
os.environ["HUGGINGFACE_HUB_CACHE"] = str(HF_CACHE)
os.environ["TRANSFORMERS_CACHE"]    = str(HF_CACHE)
_hf_token = os.environ.get("HF_TOKEN", "")
if _hf_token:
    os.environ["HUGGINGFACE_HUB_TOKEN"] = _hf_token

# ══════════════════════════════════════════════════════════════
# LOGO LoRA MODEL CONFIG
# ══════════════════════════════════════════════════════════════
FLUX_DEV_HF_ID    = "black-forest-labs/FLUX.1-dev"
LOGO_LORA_HF_ID   = "Shakker-Labs/FLUX.1-dev-LoRA-Logo-Design"
LOGO_LORA_TRIGGER = "wablogo, logo, "   # ← Must prefix every prompt!
LOGO_LORA_STEPS   = 24                  # More steps = better text accuracy
LOGO_LORA_GUIDANCE = 3.5

RMBG_HF_ID = "ZhengPeng7/BiRefNet_HR"

# ── Categories this pipeline handles ──────────────────────────
LOGO_LORA_CATEGORIES = {
    "admissions_enrollment",
    "openings_launches",
    "events",
    "education",
    "announcements",
    "business_promotions",
    "calls_to_action",
}

# ══════════════════════════════════════════════════════════════
# CONFIG — injected by inject_creds.py
# ══════════════════════════════════════════════════════════════
GOOGLE_CLIENT_ID     = os.environ.get("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET", "")
GOOGLE_REFRESH_TOKEN = os.environ.get("GOOGLE_REFRESH_TOKEN", "")
GITHUB_TOKEN         = os.environ.get("GITHUB_TOKEN_REPO2", "")
GITHUB_REPO2         = os.environ.get("GITHUB_REPO2", "")
GITHUB_REPO1         = os.environ.get("GITHUB_REPO", "")
GITHUB_TOKEN_REPO1   = os.environ.get("GITHUB_TOKEN_REPO1", "")
START_INDEX          = int(float(os.environ.get("START_INDEX", "0").strip()))
END_INDEX            = int(float(os.environ.get("END_INDEX", "200").strip()))

# Optional: filter to one category only (set in Kaggle env vars)
LOGO_LORA_CATEGORY   = os.environ.get("LOGO_LORA_CATEGORY", "").strip().lower()

SITE_URL       = "https://www.ultrapng.com"
SITE_NAME      = "UltraPNG"
WATERMARK_TEXT = "www.ultrapng.com"

# ══════════════════════════════════════════════════════════════
# PATHS  (same as main_pipeline.py — shared working dir)
# ══════════════════════════════════════════════════════════════
WORKING_DIR     = Path("/kaggle/working")
GENERATED_DIR   = WORKING_DIR / "lora_generated"    # ← separate folder!
APPROVED_DIR    = WORKING_DIR / "lora_approved"
TRANSPARENT_DIR = WORKING_DIR / "lora_transparent"
PROJECT_DIR     = WORKING_DIR / "project"
REPO2_DIR       = WORKING_DIR / "repo2"
REPO1_DIR       = WORKING_DIR / "repo1"
CHECKPOINT_DIR  = WORKING_DIR / "lora_checkpoints"  # ← separate checkpoints!

for d in [GENERATED_DIR, APPROVED_DIR, TRANSPARENT_DIR, CHECKPOINT_DIR]:
    d.mkdir(parents=True, exist_ok=True)

# ══════════════════════════════════════════════════════════════
# INSTALL DEPS
# ══════════════════════════════════════════════════════════════
print("=" * 56)
print("Installing Logo LoRA Pipeline dependencies...")
PKGS = [
    "git+https://github.com/huggingface/diffusers.git",
    "transformers>=4.47.0",
    "accelerate>=0.28.0", "sentencepiece",
    "huggingface_hub>=0.23.0",
    "bitsandbytes>=0.43.0",          # ← NF4 quantization for T4/P100
    "Pillow>=10.0", "numpy", "requests",
    "onnxruntime-gpu", "torchvision",
    "opencv-python-headless", "piexif",
    "openpyxl>=3.1.2",
]
for pkg in PKGS:
    r = subprocess.run(
        [sys.executable, "-m", "pip", "install", "-q", pkg],
        capture_output=True, text=True)
    print(f"  {'OK  ' if r.returncode == 0 else 'WARN'} {pkg}")
print("Done!\n")

import torch
import numpy as np
from PIL import Image, ImageDraw, ImageFont
import requests as req

# ══════════════════════════════════════════════════════════════
# GLOBAL LOG CAPTURE
# ══════════════════════════════════════════════════════════════
_LOG_LINES = []

def log(msg):
    line = f"[{datetime.now().strftime('%H:%M:%S')}] {msg}"
    print(line, flush=True)
    _LOG_LINES.append(line)

# ══════════════════════════════════════════════════════════════
# UTILITIES
# ══════════════════════════════════════════════════════════════
def free_memory():
    gc.collect()
    if torch.cuda.is_available():
        torch.cuda.empty_cache()
        torch.cuda.synchronize()
    used = torch.cuda.memory_allocated(0) / 1e9 if torch.cuda.is_available() else 0
    log(f"  GPU used: {used:.1f}GB")

def slugify(s, max_len=60):
    s = re.sub(r'[^a-z0-9]+', '-', str(s).lower()).strip('-')
    if len(s) > max_len:
        cut = s[:max_len]
        idx = cut.rfind('-')
        s   = cut[:idx] if idx > max_len // 2 else cut
    return s or 'untitled'

def preview_url(fid, size=800):
    return f"https://lh3.googleusercontent.com/d/{fid}=s{size}"

def download_url(fid):
    return f"https://drive.usercontent.google.com/download?id={fid}&export=download&authuser=0"

# ══════════════════════════════════════════════════════════════
# CHECKPOINT SYSTEM  (uses lora_checkpoints/ — separate!)
# ══════════════════════════════════════════════════════════════
def save_checkpoint(name, data):
    path = CHECKPOINT_DIR / f"{name}.json"
    tmp  = str(path) + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    os.replace(tmp, str(path))
    log(f"  Checkpoint saved: {name} ({len(data)} items)")

def load_checkpoint(name):
    path = CHECKPOINT_DIR / f"{name}.json"
    if path.exists():
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        log(f"  Checkpoint loaded: {name} ({len(data)} items)")
        return data
    return None

# ══════════════════════════════════════════════════════════════
# CATEGORY ENHANCERS  (same as main_pipeline.py)
# ══════════════════════════════════════════════════════════════
CATEGORY_ENHANCERS = {
    "admissions_enrollment": ", academic excellence, institutional prestige",
    "openings_launches":     ", grand opening excitement, vibrant energy",
    "events":                ", event atmosphere, celebratory design",
    "education":             ", knowledge and learning, academic theme",
    "announcements":         ", clear bold communication, attention-grabbing",
    "business_promotions":   ", professional branding, corporate appeal",
    "calls_to_action":       ", urgency, action-oriented, bold typography",
}

def enhance_prompt(raw_prompt, category):
    cat = (category or "").lower()
    for key, extra in CATEGORY_ENHANCERS.items():
        if cat.startswith(key) or key in cat:
            return raw_prompt + extra
    return raw_prompt

# ══════════════════════════════════════════════════════════════
# GOOGLE DRIVE API  (identical to main_pipeline.py)
# ══════════════════════════════════════════════════════════════
_token_cache = {"value": None, "expires": 0}

def get_drive_token():
    if _token_cache["value"] and time.time() < _token_cache["expires"]:
        return _token_cache["value"]
    r = req.post("https://oauth2.googleapis.com/token", data={
        "client_id":     GOOGLE_CLIENT_ID,
        "client_secret": GOOGLE_CLIENT_SECRET,
        "refresh_token": GOOGLE_REFRESH_TOKEN,
        "grant_type":    "refresh_token",
    }, timeout=30)
    d = r.json()
    if "access_token" not in d:
        raise Exception(f"Token error: {d}")
    _token_cache.update({"value": d["access_token"], "expires": time.time() + 3200})
    return _token_cache["value"]

def drive_folder(token, name, parent=None):
    h = {"Authorization": f"Bearer {token}"}
    q = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    if parent:
        q += f" and '{parent}' in parents"
    r = req.get("https://www.googleapis.com/drive/v3/files",
                headers=h, params={"q": q, "fields": "files(id)"}, timeout=30)
    files = r.json().get("files", [])
    if files:
        return files[0]["id"]
    meta = {"name": name, "mimeType": "application/vnd.google-apps.folder"}
    if parent:
        meta["parents"] = [parent]
    return req.post(
        "https://www.googleapis.com/drive/v3/files",
        headers={**h, "Content-Type": "application/json"},
        json=meta, timeout=30).json()["id"]

def drive_upload(token, folder_id, name, data, mime="image/png", retries=3):
    for attempt in range(1, retries + 1):
        try:
            h        = {"Authorization": f"Bearer {token}"}
            metadata = json.dumps({"name": name, "parents": [folder_id]})
            b        = "----UltraPNGLoRAPipe"
            body = (
                f"--{b}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n{metadata}\r\n"
                f"--{b}\r\nContent-Type: {mime}\r\n\r\n"
            ).encode() + data + f"\r\n--{b}--".encode()
            r = req.post(
                "https://www.googleapis.com/upload/drive/v3/files"
                "?uploadType=multipart&fields=id,name",
                headers={**h, "Content-Type": f'multipart/related; boundary="{b}"'},
                data=body, timeout=120)
            if r.ok:
                return r.json()
            raise Exception(f"HTTP {r.status_code}: {r.text[:150]}")
        except Exception as e:
            if attempt < retries:
                log(f"  Upload retry {attempt}/{retries}: {e}")
                time.sleep(5 * attempt)
            else:
                raise

def drive_share(token, fid):
    try:
        req.post(
            f"https://www.googleapis.com/drive/v3/files/{fid}/permissions",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json={"role": "reader", "type": "anyone"}, timeout=30)
    except Exception:
        pass

# ══════════════════════════════════════════════════════════════
# WATERMARK PREVIEWS  (identical to main_pipeline.py)
# ══════════════════════════════════════════════════════════════
def make_previews(png_path):
    import piexif
    with Image.open(png_path).convert("RGBA") as img_rgba:
        w, h = img_rgba.size
        if max(w, h) > 800:
            r        = 800 / max(w, h)
            img_rgba = img_rgba.resize((int(w * r), int(h * r)), Image.LANCZOS)
        w, h = img_rgba.size
        bg   = Image.new("RGB", (w, h), (255, 255, 255))
        drw  = ImageDraw.Draw(bg)
        for ry in range(0, h, 20):
            for cx in range(0, w, 20):
                if (ry // 20 + cx // 20) % 2 == 1:
                    drw.rectangle([cx, ry, cx + 20, ry + 20], fill=(232, 232, 232))
        bg.paste(img_rgba.convert("RGB"), mask=img_rgba.split()[3])
        try:
            fnt = ImageFont.truetype(
                "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", 13)
        except Exception:
            fnt = ImageFont.load_default()
        wm_layer = Image.new("RGBA", (w, h), (0, 0, 0, 0))
        wm_draw  = ImageDraw.Draw(wm_layer)
        for ry in range(-h, h + 110, 110):
            for cx in range(-w, w + 110, 110):
                wm_draw.text((cx, ry), WATERMARK_TEXT, fill=(0, 0, 0, 42), font=fnt)
        wm_rot  = wm_layer.rotate(-30, expand=False)
        bg_rgba = bg.convert("RGBA")
        bg_rgba.alpha_composite(wm_rot)
        bg   = bg_rgba.convert("RGB")
        drw2 = ImageDraw.Draw(bg)
        drw2.rectangle([0, h - 36, w, h], fill=(13, 13, 20))
        try:
            fnt2 = ImageFont.truetype(
                "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", 15)
        except Exception:
            fnt2 = fnt
        drw2.text((w // 2 - 72, h - 26), WATERMARK_TEXT, fill=(245, 166, 35), font=fnt2)
        exif_bytes = piexif.dump({
            "0th": {
                piexif.ImageIFD.Copyright: WATERMARK_TEXT.encode(),
                piexif.ImageIFD.Artist:    SITE_NAME.encode(),
                piexif.ImageIFD.Software:  b"UltraPNG Logo LoRA Pipeline V1.0",
            }
        })
        jpg_buf  = io.BytesIO()
        bg.save(jpg_buf, "JPEG", quality=85, optimize=True, exif=exif_bytes)
        webp_buf = io.BytesIO()
        bg.save(webp_buf, "WEBP", quality=82, method=6)
    return jpg_buf.getvalue(), webp_buf.getvalue(), w, h

# ══════════════════════════════════════════════════════════════
# SKIP SET  (reads same REPO2 + ultradata.xlsx as main_pipeline)
# ══════════════════════════════════════════════════════════════
def load_skip_set_from_json():
    log("Building skip-set from REPO2 JSON files...")
    skip_set = set()
    if not REPO2_DIR.exists():
        try:
            repo_url = f"https://x-access-token:{GITHUB_TOKEN}@github.com/{GITHUB_REPO2}.git"
            subprocess.run(
                ["git", "clone", "--depth", "1", "--filter=blob:none",
                 "--sparse", repo_url, str(REPO2_DIR)],
                capture_output=True, check=True)
            subprocess.run(["git", "sparse-checkout", "init", "--cone"],
                           cwd=str(REPO2_DIR), capture_output=True, check=True)
            subprocess.run(["git", "sparse-checkout", "set", "data"],
                           cwd=str(REPO2_DIR), capture_output=True, check=True)
        except Exception:
            log("  Skip-set: REPO2 empty — starting fresh")
            return skip_set
    data_dir = REPO2_DIR / "data"
    if not data_dir.exists():
        log("  No data dir — starting fresh")
        return skip_set
    for jf in data_dir.glob("*.json"):
        if jf.name.startswith("_"):
            continue
        try:
            for e in json.loads(jf.read_text("utf-8")):
                fn = e.get("filename", "")
                if fn:
                    skip_set.add(fn)
        except Exception:
            pass
    log(f"  Skip-set (REPO2): {len(skip_set)} already-done filenames")
    return skip_set

def load_skip_set_from_ultradata() -> set:
    skip_set = set()
    token = GITHUB_TOKEN_REPO1 or ""
    repo  = (GITHUB_REPO1 or "").strip()
    if not token or not repo:
        log("  Skip-set (ultradata): GITHUB_TOKEN_REPO1 not set — skipping")
        return skip_set
    repo_url  = f"https://x-access-token:{token}@github.com/{repo}.git"
    xrepo_dir = WORKING_DIR / "repo1_skipcheck"
    try:
        if xrepo_dir.exists() and (xrepo_dir / ".git").exists():
            subprocess.run(["git", "remote", "set-url", "origin", repo_url],
                           cwd=str(xrepo_dir), capture_output=True)
            subprocess.run(["git", "pull", "--rebase", "--autostash"],
                           cwd=str(xrepo_dir), capture_output=True)
        else:
            shutil.rmtree(str(xrepo_dir), ignore_errors=True)
            subprocess.run(["git", "clone", "--depth", "1", repo_url, str(xrepo_dir)],
                           capture_output=True, check=True)
    except Exception as e:
        log(f"  Skip-set (ultradata): clone failed ({e}) — skipping")
        return skip_set
    xlsx_path = xrepo_dir / ULTRADATA_XLSX_NAME
    if not xlsx_path.exists():
        return skip_set
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
        ws = wb.active
        if ws.max_row < 2:
            return skip_set
        headers = [str(c.value or "").strip().lower()
                   for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if "filename" not in headers:
            return skip_set
        col_idx = headers.index("filename")
        for row in ws.iter_rows(min_row=2, values_only=True):
            fn = row[col_idx] if col_idx < len(row) else None
            if fn:
                skip_set.add(str(fn).strip())
        wb.close()
    except Exception as e:
        log(f"  Skip-set (ultradata): read failed ({e})")
    log(f"  Skip-set (ultradata): {len(skip_set)} already-generated filenames")
    return skip_set

# ══════════════════════════════════════════════════════════════
# LOAD PROMPTS — logo categories only
# ══════════════════════════════════════════════════════════════
def load_logo_prompts():
    log("Loading logo prompts...")
    if GITHUB_REPO1 and not PROJECT_DIR.exists():
        subprocess.run(
            ["git", "clone", "--depth", "1",
             f"https://github.com/{GITHUB_REPO1}", str(PROJECT_DIR)],
            capture_output=True)
    if PROJECT_DIR.exists():
        sys.path.insert(0, str(PROJECT_DIR))
        try:
            from prompts.prompt_engine import load_all_prompts
            all_prompts = load_all_prompts(str(PROJECT_DIR / "prompts" / "splits"))
            # Filter to logo categories only
            target_cats = {LOGO_LORA_CATEGORY} if LOGO_LORA_CATEGORY else LOGO_LORA_CATEGORIES
            logo_prompts = [
                p for p in all_prompts
                if p.get("category_slug", "").lower() in target_cats
            ]
            log(f"  Total prompts: {len(all_prompts)} | Logo filtered: {len(logo_prompts)}")
            return logo_prompts
        except Exception as e:
            log(f"Prompt error: {e}")
    log("FATAL: No prompts loaded!")
    return []

# ══════════════════════════════════════════════════════════════
# PHASE 1B — FLUX.1-dev + Logo LoRA
# ══════════════════════════════════════════════════════════════
def phase1b_generate_logo(batch, skip_set):
    ckpt = load_checkpoint("phase1b_logo_generated")
    if ckpt:
        return ckpt

    log("=" * 56)
    log("PHASE 1B: FLUX.1-dev + Logo LoRA — Logo Generation")
    log(f"  Base model : {FLUX_DEV_HF_ID}")
    log(f"  LoRA       : {LOGO_LORA_HF_ID}")
    log(f"  Trigger    : '{LOGO_LORA_TRIGGER}'")
    log(f"  Steps      : {LOGO_LORA_STEPS}")
    log(f"  Guidance   : {LOGO_LORA_GUIDANCE}")
    log("=" * 56)

    from diffusers import FluxPipeline
    from transformers import BitsAndBytesConfig

    # ── Detect GPU VRAM and pick loading strategy ──────────────
    vram_gb = torch.cuda.get_device_properties(0).total_memory / 1e9 if torch.cuda.is_available() else 0
    gpu_name = torch.cuda.get_device_name(0) if torch.cuda.is_available() else "CPU"
    log(f"  GPU       : {gpu_name}")
    log(f"  VRAM      : {vram_gb:.1f} GB")

    # T4 = 15GB, P100 = 16GB — both need quantization for FLUX.1-dev
    # A100 = 40GB+ — can run full fp16 safely
    USE_NF4 = vram_gb < 32.0

    log(f"  Strategy  : {'NF4 4-bit quantization (VRAM < 32GB)' if USE_NF4 else 'Full fp16 (VRAM >= 32GB)'}")
    log(f"  Loading FLUX.1-dev (first run: ~25GB download)...")

    if USE_NF4:
        # ── NF4 Quantized load (~10GB VRAM) ───────────────────
        # Works on T4 (15GB) and P100 (16GB)
        nf4_config = BitsAndBytesConfig(
            load_in_4bit=True,
            bnb_4bit_quant_type="nf4",
            bnb_4bit_compute_dtype=torch.bfloat16,
        )
        pipe = FluxPipeline.from_pretrained(
            FLUX_DEV_HF_ID,
            quantization_config=nf4_config,
            torch_dtype=torch.bfloat16,
            device_map="auto",
        )
    else:
        # ── Full fp16 load (A100 / H100) ──────────────────────
        pipe = FluxPipeline.from_pretrained(
            FLUX_DEV_HF_ID,
            torch_dtype=torch.float16,
        )
        pipe.to("cuda")

    log(f"  Loading Logo LoRA weights...")
    pipe.load_lora_weights(LOGO_LORA_HF_ID)

    # Sequential CPU offload only needed when NOT using device_map="auto"
    if not USE_NF4:
        pipe.enable_sequential_cpu_offload()
    pipe.set_progress_bar_config(disable=True)
    log(f"  Ready! Batch: {len(batch)} | Skip: {len(skip_set)}\n")

    generated, skipped, t0 = [], 0, time.time()

    for i, item in enumerate(batch):
        fname = item["filename"]

        if fname in skip_set:
            skipped += 1
            if skipped <= 3 or skipped % 50 == 0:
                log(f"  SKIP [{i+1}/{len(batch)}] {fname}")
            continue

        try:
            out_dir = GENERATED_DIR / item["category"] / item.get("subcategory", "general")
            out_dir.mkdir(parents=True, exist_ok=True)
            out = out_dir / fname

            if out.exists():
                generated.append({"path": str(out), "item": item})
                continue

            gen = torch.Generator("cpu").manual_seed(item["seed"])

            # ── TRIGGER WORD prepended here ──
            raw    = enhance_prompt(item["prompt"], item.get("category", ""))
            prompt = LOGO_LORA_TRIGGER + raw

            img = pipe(
                prompt=prompt,
                num_inference_steps=LOGO_LORA_STEPS,
                guidance_scale=LOGO_LORA_GUIDANCE,
                height=1024, width=1024,
                generator=gen,
            ).images[0]

            # Blank image check
            arr    = np.array(img)
            n_px   = arr.shape[0] * arr.shape[1]
            thresh = int(n_px * 0.003)
            if arr.std() < 5 or (arr < 250).sum() < thresh:
                log(f"  Retry (blank): {fname}")
                gen2 = torch.Generator("cpu").manual_seed(item["seed"] + 99)
                img  = pipe(
                    prompt=prompt,
                    num_inference_steps=LOGO_LORA_STEPS,
                    guidance_scale=LOGO_LORA_GUIDANCE,
                    height=1024, width=1024,
                    generator=gen2,
                ).images[0]

            img.save(str(out), "PNG", compress_level=9)
            generated.append({"path": str(out), "item": item})

            done = len(generated)
            rate = done / (time.time() - t0)
            eta  = (len(batch) - i - 1) / rate / 60 if rate > 0 else 0
            log(f"  [{i+1}/{len(batch)}] OK {fname} | {item['category']} | ETA {eta:.0f}min")

        except torch.cuda.OutOfMemoryError:
            torch.cuda.empty_cache()
            log(f"  OOM: {fname}")
        except Exception as e:
            log(f"  FAIL {fname}: {e}")

    log("\n  Deleting FLUX.1-dev + LoRA → freeing VRAM + disk cache...")
    del pipe
    free_memory()

    import shutil as _shutil
    for _cache_sub in HF_CACHE.iterdir():
        if _cache_sub.is_dir():
            _name = _cache_sub.name.lower()
            if "flux" in _name or "black-forest" in _name or "shakker" in _name:
                _shutil.rmtree(str(_cache_sub), ignore_errors=True)
                log(f"  Deleted cache: {_cache_sub.name}")
    _used = sum(f.stat().st_size for f in HF_CACHE.rglob("*") if f.is_file()) / 1e9
    log(f"  HF cache after cleanup: {_used:.1f}GB")

    save_checkpoint("phase1b_logo_generated", generated)
    log(f"PHASE 1B DONE — Generated: {len(generated)} | Skipped: {skipped}\n")
    return generated

# ══════════════════════════════════════════════════════════════
# PHASE 3 — Background Removal (BiRefNet_HR)
# ══════════════════════════════════════════════════════════════
def phase3_bg_remove(posts):
    ckpt = load_checkpoint("phase3_lora_transparent")
    if ckpt:
        return ckpt

    log("=" * 56)
    log("PHASE 3: BiRefNet_HR — Background Removal (GPU)")
    log(f"  Loading: {RMBG_HF_ID}")
    log("=" * 56)

    if not posts:
        return []

    from torchvision import transforms
    from transformers import AutoModelForImageSegmentation

    rmbg_model = AutoModelForImageSegmentation.from_pretrained(
        RMBG_HF_ID,
        trust_remote_code=True,
        cache_dir=str(HF_CACHE),
    )
    device = "cuda" if torch.cuda.is_available() else "cpu"
    torch.set_float32_matmul_precision("high")
    rmbg_model = rmbg_model.to(device).eval().half()
    log(f"  BiRefNet_HR loaded on {device.upper()} | FP16 | 2048x2048\n")

    transform_img = transforms.Compose([
        transforms.Resize((2048, 2048)),
        transforms.ToTensor(),
        transforms.Normalize([0.485, 0.456, 0.406],
                             [0.229, 0.224, 0.225]),
    ])

    def remove_bg(pil_img):
        ow, oh   = pil_img.size
        inp      = transform_img(pil_img.convert("RGB")).unsqueeze(0).to(device).half()
        with torch.no_grad():
            preds = rmbg_model(inp)[-1].sigmoid().cpu()
        pred     = preds[0].squeeze()
        mask_pil = transforms.ToPILImage()(pred).resize((ow, oh), Image.LANCZOS)
        result   = pil_img.convert("RGBA")
        result.putalpha(mask_pil)
        return result

    result_posts, t0 = [], time.time()

    for i, post in enumerate(posts):
        path = Path(post["approved_path"])
        try:
            rel = path.relative_to(APPROVED_DIR)
            out = TRANSPARENT_DIR / rel
            out.parent.mkdir(parents=True, exist_ok=True)

            if out.exists():
                result_posts.append({**post, "transparent_path": str(out)})
                continue

            img    = Image.open(str(path)).convert("RGB")
            result = remove_bg(img)
            result.save(str(out), "PNG", compress_level=9)
            result_posts.append({**post, "transparent_path": str(out)})

            if (i + 1) % 20 == 0:
                log(f"  BG done: {i+1}/{len(posts)} | {(i+1)/(time.time()-t0):.2f}/s")

        except Exception as e:
            log(f"  RMBG FAIL {path.name}: {e}")

    log("\n  Deleting BiRefNet_HR → freeing VRAM...")
    del rmbg_model, transform_img
    free_memory()

    import shutil as _shutil
    for _cache_sub in HF_CACHE.iterdir():
        if _cache_sub.is_dir():
            _name = _cache_sub.name.lower()
            if "rmbg" in _name or "briaai" in _name or "birefnet" in _name:
                _shutil.rmtree(str(_cache_sub), ignore_errors=True)
                log(f"  Deleted BiRefNet cache: {_cache_sub.name}")

    if APPROVED_DIR.exists():
        _shutil.rmtree(str(APPROVED_DIR), ignore_errors=True)
        APPROVED_DIR.mkdir(parents=True, exist_ok=True)
        log("  Deleted lora_approved/ images (transparent copies kept)")

    save_checkpoint("phase3_lora_transparent", result_posts)
    log(f"PHASE 3 DONE — Transparent PNGs: {len(result_posts)}\n")
    return result_posts

# ══════════════════════════════════════════════════════════════
# PHASE 4 — Google Drive Upload
# ══════════════════════════════════════════════════════════════
def phase4_upload(posts):
    ckpt = load_checkpoint("phase4_lora_uploaded")
    if ckpt:
        return ckpt

    log("=" * 56)
    log("PHASE 4: Google Drive Upload (PNG + WebP)")
    log("=" * 56)

    token    = get_drive_token()
    png_root = drive_folder(token, "png_library_images")
    prv_root = drive_folder(token, "png_library_previews")

    existing_png_ids = {}
    existing_prv_ids = {}

    result_posts, t0 = [], time.time()

    for i, post in enumerate(posts):
        try:
            cat  = post.get("category",    "general")
            sub  = post.get("subcategory", "general")
            fname = post["filename"]

            png_cat_folder = drive_folder(token, cat, png_root)
            png_sub_folder = drive_folder(token, sub, png_cat_folder)
            prv_cat_folder = drive_folder(token, cat, prv_root)
            prv_sub_folder = drive_folder(token, sub, prv_cat_folder)

            # PNG upload
            png_data  = Path(post["transparent_path"]).read_bytes()
            png_resp  = drive_upload(token, png_sub_folder, fname, png_data, "image/png")
            png_fid   = png_resp["id"]
            drive_share(token, png_fid)

            # WebP preview upload
            _jpg_bytes, webp_bytes, pw, ph = make_previews(Path(post["transparent_path"]))
            webp_name = Path(fname).stem + ".webp"
            prv_resp  = drive_upload(token, prv_sub_folder, webp_name, webp_bytes, "image/webp")
            webp_fid  = prv_resp["id"]
            drive_share(token, webp_fid)

            result_posts.append({
                **post,
                "png_file_id":       png_fid,
                "jpg_file_id":       "",
                "webp_file_id":      webp_fid,
                "download_url":      download_url(png_fid),
                "preview_url":       preview_url(webp_fid, 800),
                "preview_url_small": preview_url(webp_fid, 400),
                "webp_preview_url":  preview_url(webp_fid, 800),
                "preview_w":         pw,
                "preview_h":         ph,
            })

            rate = (i + 1) / (time.time() - t0)
            eta  = (len(posts) - i - 1) / rate / 60 if rate > 0 else 0
            log(f"  [{i+1}/{len(posts)}] OK {fname} | ETA {eta:.0f}min")

        except Exception as e:
            log(f"  UPLOAD FAIL {post.get('filename','?')}: {e}")

    save_checkpoint("phase4_lora_uploaded", result_posts)
    log(f"PHASE 4 DONE — Uploaded: {len(result_posts)}\n")
    return result_posts

# ══════════════════════════════════════════════════════════════
# PHASE 5 — Append ultradata.xlsx → push REPO1
# ══════════════════════════════════════════════════════════════
def _append_ultradata_and_push(items):
    token = GITHUB_TOKEN_REPO1 or ""
    repo  = (GITHUB_REPO1 or "").strip()
    if not token or not repo:
        raise Exception("Missing GITHUB_TOKEN_REPO1 or GITHUB_REPO (repo1)")

    repo_url  = f"https://x-access-token:{token}@github.com/{repo}.git"
    xrepo_dir = WORKING_DIR / "repo1_xlsx"

    if xrepo_dir.exists() and (xrepo_dir / ".git").exists():
        subprocess.run(["git", "remote", "set-url", "origin", repo_url],
                       cwd=str(xrepo_dir), capture_output=True)
        subprocess.run(["git", "pull", "--rebase", "--autostash"],
                       cwd=str(xrepo_dir), capture_output=True)
    else:
        shutil.rmtree(str(xrepo_dir), ignore_errors=True)
        subprocess.run(["git", "clone", "--depth", "1", repo_url, str(xrepo_dir)],
                       capture_output=True, check=True)

    try:
        import openpyxl
        from openpyxl import Workbook
    except Exception:
        subprocess.run([sys.executable, "-m", "pip", "install", "-q", "openpyxl>=3.1.2"],
                       capture_output=True, check=True)
        import openpyxl
        from openpyxl import Workbook

    xlsx_path = xrepo_dir / ULTRADATA_XLSX_NAME
    headers   = [
        "date_added", "subject_name", "category", "subcategory", "filename",
        "png_file_id", "webp_file_id", "download_url", "preview_url"
    ]

    if xlsx_path.exists():
        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)

    def _row(it):
        return [
            it.get("date_added", ""),
            it.get("subject_name", ""),
            it.get("category", ""),
            it.get("subcategory", ""),
            it.get("filename", ""),
            it.get("png_file_id", ""),
            it.get("webp_file_id", ""),
            it.get("download_url", ""),
            it.get("preview_url", ""),
        ]

    appended = 0
    for it in items:
        if not it.get("subject_name") or not it.get("download_url") or not it.get("preview_url"):
            continue
        ws.append(_row(it))
        appended += 1

    wb.save(str(xlsx_path))

    if appended == 0:
        log("ultradata.xlsx: nothing new to append")
        return

    orig_dir = os.getcwd()
    try:
        os.chdir(str(xrepo_dir))
        subprocess.run(["git", "config", "user.name", "github-actions[bot]"],
                       check=True, capture_output=True)
        subprocess.run(["git", "config", "user.email",
                        "github-actions[bot]@users.noreply.github.com"],
                       check=True, capture_output=True)
        subprocess.run(["git", "add", ULTRADATA_XLSX_NAME], check=True, capture_output=True)
        diff = subprocess.run(["git", "diff", "--staged", "--quiet"], capture_output=True)
        if diff.returncode != 0:
            msg = f"ultradata: logo-lora append {appended} rows [{datetime.now().strftime('%Y-%m-%d')}]"
            subprocess.run(["git", "commit", "-m", msg], check=True, capture_output=True)
            pushed = False
            for push_attempt in range(1, 5):
                result = subprocess.run(["git", "push"], capture_output=True, text=True)
                if result.returncode == 0:
                    pushed = True
                    break
                log(f"  Push attempt {push_attempt}/4 failed — pulling rebase...")
                subprocess.run(["git", "pull", "--rebase", "--autostash"],
                               capture_output=True, cwd=str(xrepo_dir))
                time.sleep(5 * push_attempt)
            if not pushed:
                raise Exception(f"git push failed after 4 attempts:\n{result.stderr.strip()}")
            log(f"ultradata.xlsx pushed: +{appended} rows")
        else:
            log("ultradata.xlsx: no staged changes")
    finally:
        os.chdir(orig_dir)

# ══════════════════════════════════════════════════════════════
# PHASE 6 — Save Logs → REPO1
# ══════════════════════════════════════════════════════════════
def phase6_save_logs(stats: dict):
    token = GITHUB_TOKEN_REPO1 or ""
    repo  = (GITHUB_REPO1 or "").strip()
    if not token or not repo:
        log("  Logs: GITHUB_TOKEN_REPO1 not set — skipping log push")
        return

    ts        = datetime.now().strftime("%Y-%m-%d_%H-%M")
    log_name  = f"lora_{ts}.log"
    repo_url  = f"https://x-access-token:{token}@github.com/{repo}.git"
    xrepo_dir = WORKING_DIR / "repo1_logs"

    if xrepo_dir.exists() and (xrepo_dir / ".git").exists():
        subprocess.run(["git", "remote", "set-url", "origin", repo_url],
                       cwd=str(xrepo_dir), capture_output=True)
        subprocess.run(["git", "pull", "--rebase", "--autostash"],
                       cwd=str(xrepo_dir), capture_output=True)
    else:
        shutil.rmtree(str(xrepo_dir), ignore_errors=True)
        subprocess.run(["git", "clone", "--depth", "1", repo_url, str(xrepo_dir)],
                       capture_output=True, check=True)

    logs_dir = xrepo_dir / "logs"
    logs_dir.mkdir(exist_ok=True)

    summary = "\n".join([
        "=" * 56,
        "LOGO LoRA PIPELINE — RUN SUMMARY",
        f"  Date      : {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"  Status    : {stats.get('status', '?')}",
        f"  Duration  : {stats.get('duration', '?')}",
        f"  Generated : {stats.get('generated', 0)}",
        f"  Transparent: {stats.get('transparent', 0)}",
        f"  Uploaded  : {stats.get('uploaded', 0)}",
        f"  Trigger   : {LOGO_LORA_TRIGGER}",
        f"  Model     : {FLUX_DEV_HF_ID}",
        f"  LoRA      : {LOGO_LORA_HF_ID}",
        "=" * 56,
        "",
    ] + _LOG_LINES)

    (logs_dir / log_name).write_text(summary, encoding="utf-8")
    (logs_dir / "latest_lora.log").write_text(summary, encoding="utf-8")

    orig_dir = os.getcwd()
    try:
        os.chdir(str(xrepo_dir))
        subprocess.run(["git", "config", "user.name", "github-actions[bot]"],
                       capture_output=True)
        subprocess.run(["git", "config", "user.email",
                        "github-actions[bot]@users.noreply.github.com"],
                       capture_output=True)
        subprocess.run(["git", "add", "logs/"], capture_output=True)
        subprocess.run(["git", "commit", "-m", f"logo-lora log: {ts}"],
                       capture_output=True)
        subprocess.run(["git", "push"], capture_output=True)
        log(f"  Log saved: logs/{log_name}")
    except Exception as e:
        log(f"  Log push error (non-fatal): {e}")
    finally:
        os.chdir(orig_dir)

# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════
def main():
    t0    = time.time()
    stats = {"status": "running", "generated": 0,
             "transparent": 0, "uploaded": 0, "duration": "?"}

    print("╔══════════════════════════════════════════════════════╗")
    print("║  UltraPNG — Logo LoRA Pipeline V1.0                 ║")
    print("║  FLUX.1-dev + Logo LoRA → RMBG → Drive → xlsx      ║")
    print("╚══════════════════════════════════════════════════════╝")
    print(f"  Batch   : {START_INDEX} → {END_INDEX} ({END_INDEX-START_INDEX} prompts)")
    print(f"  Trigger : {LOGO_LORA_TRIGGER}")
    print(f"  Steps   : {LOGO_LORA_STEPS}")
    print(f"  Filter  : {LOGO_LORA_CATEGORY or 'ALL logo categories'}")
    if torch.cuda.is_available():
        p = torch.cuda.get_device_properties(0)
        print(f"  GPU     : {p.name} | VRAM: {p.total_memory/1e9:.0f}GB")
    print()

    try:
        prompts = load_logo_prompts()
        if not prompts:
            raise Exception("No logo prompts loaded!")

        skip_set  = load_skip_set_from_json()
        skip_set |= load_skip_set_from_ultradata()
        log(f"  Combined skip-set: {len(skip_set)} filenames\n")

        # ── AUTO-ADVANCE (same logic as main_pipeline.py) ──
        BATCH_SIZE = END_INDEX - START_INDEX
        real_start = START_INDEX
        real_end   = END_INDEX
        MAX_SCAN   = min(len(prompts), START_INDEX + max(BATCH_SIZE * 50, 5000))

        scan_start = START_INDEX
        while scan_start < MAX_SCAN:
            window = prompts[scan_start:scan_start + BATCH_SIZE]
            fresh  = [p for p in window if p["filename"] not in skip_set]
            if fresh:
                if scan_start != START_INDEX:
                    log(f"  AUTO-ADVANCE: {START_INDEX} → {scan_start}")
                real_start = scan_start
                real_end   = scan_start + BATCH_SIZE
                break
            scan_start += BATCH_SIZE
        else:
            log("  All logo prompts already done — nothing to generate.")
            return

        batch = prompts[real_start:real_end]
        log(f"Batch: {real_start} → {real_end} | "
            f"{len([p for p in batch if p['filename'] not in skip_set])} fresh\n")

        os.environ["START_INDEX"] = str(real_start)
        os.environ["END_INDEX"]   = str(real_end)

        # ── PHASE 1B ──
        generated = phase1b_generate_logo(batch, skip_set)
        stats["generated"] = len(generated)
        if not generated:
            log("No logo images generated."); return

        # ── Build posts list (same schema as main_pipeline.py) ──
        posts = []
        for gd in generated:
            item    = gd["item"]
            src     = Path(gd["path"])
            rel     = src.relative_to(GENERATED_DIR)
            dst     = APPROVED_DIR / rel
            dst.parent.mkdir(parents=True, exist_ok=True)
            if not dst.exists():
                shutil.copy2(str(src), str(dst))
            subject = (item.get("subject_name") or item.get("subject") or
                       item.get("name") or item.get("title") or
                       item.get("subcategory") or item.get("category") or "Untitled")
            subject = str(subject).replace("_", " ").replace("-", " ").title()
            posts.append({
                "category":       item.get("category", "general"),
                "subcategory":    item.get("subcategory", "general"),
                "subject_name":   subject,
                "filename":       item.get("filename", src.name),
                "original_prompt": item.get("prompt", ""),
                "approved_path":  str(dst),
                "slug":           slugify(subject),
                "title": "", "h1": "", "meta_desc": "", "alt_text": "",
                "tags": "", "description": "", "word_count": 0,
                "ai_generated": False, "qwen_confidence": 0, "qwen_group": "",
                "png_file_id": "", "jpg_file_id": "", "webp_file_id": "",
                "download_url": "", "preview_url": "", "preview_url_small": "",
                "webp_preview_url": "", "preview_w": 800, "preview_h": 800,
                "date_added": datetime.now().strftime("%Y-%m-%d"),
            })
        if not posts:
            log("No posts to process."); return

        # ── PHASE 3 ──
        transparent = phase3_bg_remove(posts)
        stats["transparent"] = len(transparent)
        if not transparent:
            log("BG removal produced no results."); return

        # ── PHASE 4 ──
        uploaded = phase4_upload(transparent)
        stats["uploaded"] = len(uploaded)
        if not uploaded:
            log("Drive upload failed."); return

        # ── PHASE 5 ──
        _append_ultradata_and_push(uploaded)

        # Clear lora checkpoints only after xlsx pushed
        for ck in CHECKPOINT_DIR.glob("*.json"):
            ck.unlink()

        # Write effective_end.txt
        eff_end = int(os.environ.get("END_INDEX", str(END_INDEX)))
        (WORKING_DIR / "effective_end.txt").write_text(str(eff_end))
        log(f"  effective_end.txt written: {eff_end}")

        hrs = (time.time() - t0) / 3600
        stats.update({"duration": f"{hrs:.1f}h", "status": "SUCCESS"})

        print(f"\n╔══════════════════════════════════════════════════════╗")
        print(f"║  LOGO LoRA DONE in {hrs:.1f}h")
        print(f"║  Gen:{len(generated)}  Trans:{len(transparent)}  Up:{len(uploaded)}")
        print(f"╚══════════════════════════════════════════════════════╝")

    except Exception as e:
        stats["status"] = f"FAILED: {e}"
        log(f"FATAL: {e}")
        import traceback; traceback.print_exc()
        raise

    finally:
        phase6_save_logs(stats)


if __name__ == "__main__":
    main()
