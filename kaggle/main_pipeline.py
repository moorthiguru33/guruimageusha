"""
╔══════════════════════════════════════════════════════════════╗
║   UltraPNG.com — PNG Library Pipeline V7.0                  ║
╠══════════════════════════════════════════════════════════════╣
║  PHASE 1 → FLUX.2-Klein-4B   Generate 1024x1024 images     ║
║  PHASE 3 → RMBG-2.0 ONNX    Background removal (GPU)       ║
║  PHASE 4 → Google Drive      Upload PNG + WebP              ║
║  PHASE 5 → ultradata.xlsx     Append rows in REPO1          ║
║  PHASE 6 → Save Run Logs → REPO1 (visible on GitHub!)      ║
╚══════════════════════════════════════════════════════════════╝

inject_creds.py prepends os.environ[] lines before this file.
"""

import os, sys, json, time, gc, re, io, shutil, base64, subprocess, math
from pathlib import Path
from datetime import datetime

# ── Force real-time log output in Kaggle (safe — works in Jupyter & script) ──
try:
    sys.stdout.reconfigure(line_buffering=True)
    sys.stderr.reconfigure(line_buffering=True)
except AttributeError:
    pass  # Kaggle Jupyter IPython stream — reconfigure() not supported

os.environ["PYTORCH_CUDA_ALLOC_CONF"] = "expandable_segments:True"
ULTRADATA_XLSX_NAME = "ultradata.xlsx"

# ── HuggingFace cache dir (writable in Kaggle) ────────────────
HF_CACHE = Path("/kaggle/working/hf_cache")
HF_CACHE.mkdir(parents=True, exist_ok=True)
os.environ["HF_HOME"]               = str(HF_CACHE)
os.environ["HUGGINGFACE_HUB_CACHE"] = str(HF_CACHE)
os.environ["TRANSFORMERS_CACHE"]    = str(HF_CACHE)
# Set HF token if available (faster downloads, no rate limits)
_hf_token = os.environ.get("HF_TOKEN", "")
if _hf_token:
    os.environ["HUGGINGFACE_HUB_TOKEN"] = _hf_token

# ══════════════════════════════════════════════════════════════
# HuggingFace Model IDs  ← change here if repo name differs
# ══════════════════════════════════════════════════════════════
FLUX_HF_ID = "black-forest-labs/FLUX.2-klein-4B"
RMBG_HF_ID = "ZhengPeng7/BiRefNet_HR"  # Highest quality: 2048x2048 trained, public MIT, no gating, verified working

# ══════════════════════════════════════════════════════════════
# GLOBAL LOG CAPTURE
# ══════════════════════════════════════════════════════════════
_LOG_LINES = []

def log(msg):
    line = f"[{datetime.now().strftime('%H:%M:%S')}] {msg}"
    print(line, flush=True)
    _LOG_LINES.append(line)

# ── Install deps ──────────────────────────────────────────────
print("=" * 56)
print("Installing dependencies...")
PKGS = [
    "git+https://github.com/huggingface/diffusers.git",
    "transformers>=4.47.0",
    "accelerate>=0.28.0", "sentencepiece",
    "huggingface_hub>=0.23.0",
    "Pillow>=10.0", "numpy", "requests",
    "onnxruntime-gpu", "torchvision",
    "opencv-python-headless", "piexif",
]
for pkg in PKGS:
    r = subprocess.run(
        [sys.executable, "-m", "pip", "install", "-q", pkg],
        capture_output=True, text=True)
    print(f"  {'OK  ' if r.returncode == 0 else 'WARN'} {pkg}")
print("Done!\n")

import torch
import numpy as np
from PIL import Image, ImageDraw, ImageFont
import requests as req

# ══════════════════════════════════════════════════════════════
# PATHS
# ══════════════════════════════════════════════════════════════
WORKING_DIR     = Path("/kaggle/working")
GENERATED_DIR   = WORKING_DIR / "generated"
APPROVED_DIR    = WORKING_DIR / "approved"
TRANSPARENT_DIR = WORKING_DIR / "transparent"
PROJECT_DIR     = WORKING_DIR / "project"
REPO2_DIR       = WORKING_DIR / "repo2"
REPO1_DIR       = WORKING_DIR / "repo1"
CHECKPOINT_DIR  = WORKING_DIR / "checkpoints"

for d in [GENERATED_DIR, APPROVED_DIR, TRANSPARENT_DIR, CHECKPOINT_DIR]:
    d.mkdir(parents=True, exist_ok=True)

# ══════════════════════════════════════════════════════════════
# CONFIG — injected by inject_creds.py
# ══════════════════════════════════════════════════════════════
GOOGLE_CLIENT_ID     = os.environ.get("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET", "")
GOOGLE_REFRESH_TOKEN = os.environ.get("GOOGLE_REFRESH_TOKEN", "")
GITHUB_TOKEN         = os.environ.get("GITHUB_TOKEN_REPO2", "")
GITHUB_REPO2         = os.environ.get("GITHUB_REPO2", "")
GITHUB_REPO1         = os.environ.get("GITHUB_REPO", "")
GITHUB_TOKEN_REPO1   = os.environ.get("GITHUB_TOKEN_REPO1", "")
START_INDEX          = int(float(os.environ.get("START_INDEX", "0").strip()))
END_INDEX            = int(float(os.environ.get("END_INDEX", "200").strip()))

SITE_URL        = "https://www.ultrapng.com"
SITE_NAME       = "UltraPNG"
WATERMARK_TEXT  = "www.ultrapng.com"
ITEMS_PER_PAGE  = 24

SITEMAP_MAX_URL = 45000

# ══════════════════════════════════════════════════════════════
# CATEGORY ENHANCERS — prompt quality boost per category
# ══════════════════════════════════════════════════════════════
CATEGORY_ENHANCERS = {
    "indian_foods":     ", appetizing food styling, steam visible, glistening surface",
    "food_indian":      ", appetizing food styling, steam visible, glistening surface",
    "world_foods":      ", appetizing food styling, steam visible, glistening sauce",
    "food_world":       ", appetizing food styling, steam visible, glistening sauce",
    "fruits":           ", natural skin texture, juice droplets",
    "vegetables":       ", natural surface texture, fresh harvest quality",
    "flowers":          ", petal vein detail, natural color saturation",
    "jewellery_models": ", gem facet reflections, gold metal mirror finish",
    "jewellery":        ", gem facet reflections, gold metal mirror finish",
    "vehicles":         ", automotive paint reflection, chrome detail",
    "vehicles_cars":    ", automotive paint reflection, chrome detail",
    "vehicles_bikes":   ", automotive paint reflection, chrome detail",
    "poultry_animals":  ", fur and feather strand detail, catchlight in eyes",
    "animals":          ", fur and feather strand detail, catchlight in eyes",
    "raw_meat":         ", fresh meat texture, glistening moist surface",
    "cool_drinks":      ", condensation droplets, liquid transparency",
    "beverages":        ", condensation droplets, liquid transparency",
    "footwear":         ", leather grain, stitching detail",
    "shoes":            ", leather grain, stitching detail",
    "indian_dress":     ", fabric weave, embroidery thread detail",
    "clothing":         ", fabric weave, embroidery thread detail",
    "office_models":    ", professional portrait, sharp clothing detail",
}

def enhance_prompt(raw_prompt, category):
    cat = (category or "").lower()
    for key, extra in CATEGORY_ENHANCERS.items():
        if cat.startswith(key) or key in cat:
            return raw_prompt + extra
    return raw_prompt

# ══════════════════════════════════════════════════════════════
# UTILITIES
# ══════════════════════════════════════════════════════════════
def free_memory():
    gc.collect()
    if torch.cuda.is_available():
        torch.cuda.empty_cache()
        torch.cuda.synchronize()
    used = torch.cuda.memory_allocated(0) / 1e9 if torch.cuda.is_available() else 0
    log(f"  GPU used: {used:.1f}GB")

def slugify(s, max_len=60):
    s = re.sub(r'[^a-z0-9]+', '-', str(s).lower()).strip('-')
    if len(s) > max_len:
        cut = s[:max_len]
        idx = cut.rfind('-')
        s   = cut[:idx] if idx > max_len // 2 else cut
    return s or 'untitled'

def esc(s):
    return (str(s or '')
            .replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            .replace('"', '&quot;').replace("'", "&#39;"))

def preview_url(fid, size=800):
    return f"https://lh3.googleusercontent.com/d/{fid}=s{size}"

def download_url(fid):
    return f"https://drive.usercontent.google.com/download?id={fid}&export=download&authuser=0"

# ══════════════════════════════════════════════════════════════
# CHECKPOINT SYSTEM
# ══════════════════════════════════════════════════════════════
def save_checkpoint(name, data):
    path = CHECKPOINT_DIR / f"{name}.json"
    tmp  = str(path) + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False)
    os.replace(tmp, str(path))
    log(f"  Checkpoint saved: {name} ({len(data)} items)")

def load_checkpoint(name):
    path = CHECKPOINT_DIR / f"{name}.json"
    if path.exists():
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        log(f"  Checkpoint loaded: {name} ({len(data)} items)")
        return data
    return None

# ══════════════════════════════════════════════════════════════
# GOOGLE DRIVE API
# ══════════════════════════════════════════════════════════════
_token_cache = {"value": None, "expires": 0}

def get_drive_token():
    if _token_cache["value"] and time.time() < _token_cache["expires"]:
        return _token_cache["value"]
    r = req.post("https://oauth2.googleapis.com/token", data={
        "client_id":     GOOGLE_CLIENT_ID,
        "client_secret": GOOGLE_CLIENT_SECRET,
        "refresh_token": GOOGLE_REFRESH_TOKEN,
        "grant_type":    "refresh_token",
    }, timeout=30)
    d = r.json()
    if "access_token" not in d:
        raise Exception(f"Token error: {d}")
    _token_cache.update({"value": d["access_token"], "expires": time.time() + 3200})
    return _token_cache["value"]

def drive_folder(token, name, parent=None):
    h = {"Authorization": f"Bearer {token}"}
    q = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    if parent:
        q += f" and '{parent}' in parents"
    r = req.get("https://www.googleapis.com/drive/v3/files",
                headers=h, params={"q": q, "fields": "files(id)"}, timeout=30)
    files = r.json().get("files", [])
    if files:
        return files[0]["id"]
    meta = {"name": name, "mimeType": "application/vnd.google-apps.folder"}
    if parent:
        meta["parents"] = [parent]
    return req.post(
        "https://www.googleapis.com/drive/v3/files",
        headers={**h, "Content-Type": "application/json"},
        json=meta, timeout=30).json()["id"]

def drive_upload(token, folder_id, name, data, mime="image/png", retries=3):
    for attempt in range(1, retries + 1):
        try:
            h        = {"Authorization": f"Bearer {token}"}
            metadata = json.dumps({"name": name, "parents": [folder_id]})
            b        = "----UltraPNGPipe"
            body = (
                f"--{b}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n{metadata}\r\n"
                f"--{b}\r\nContent-Type: {mime}\r\n\r\n"
            ).encode() + data + f"\r\n--{b}--".encode()
            r = req.post(
                "https://www.googleapis.com/upload/drive/v3/files"
                "?uploadType=multipart&fields=id,name",
                headers={**h, "Content-Type": f'multipart/related; boundary="{b}"'},
                data=body, timeout=120)
            if r.ok:
                return r.json()
            raise Exception(f"HTTP {r.status_code}: {r.text[:150]}")
        except Exception as e:
            if attempt < retries:
                log(f"  Upload retry {attempt}/{retries}: {e}")
                time.sleep(5 * attempt)
            else:
                raise

def drive_share(token, fid):
    try:
        req.post(
            f"https://www.googleapis.com/drive/v3/files/{fid}/permissions",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json={"role": "reader", "type": "anyone"}, timeout=30)
    except Exception:
        pass

def drive_list_pngs(token, folder_id, page_size=1000):
    h = {"Authorization": f"Bearer {token}"}
    q = (
        f"'{folder_id}' in parents and trashed=false and "
        "(mimeType='image/png' or name contains '.png')"
    )
    r = req.get(
        "https://www.googleapis.com/drive/v3/files",
        headers=h,
        params={"q": q, "pageSize": page_size, "fields": "files(id,name,parents,mimeType)"},
        timeout=30,
    )
    r.raise_for_status()
    return r.json().get("files", [])

def drive_download_bytes(token, fid):
    h = {"Authorization": f"Bearer {token}"}
    r = req.get(
        f"https://www.googleapis.com/drive/v3/files/{fid}",
        headers=h,
        params={"alt": "media"},
        timeout=120,
    )
    r.raise_for_status()
    return r.content

def drive_move(token, fid, add_parent_id, remove_parent_id):
    h = {"Authorization": f"Bearer {token}"}
    r = req.patch(
        f"https://www.googleapis.com/drive/v3/files/{fid}",
        headers=h,
        params={"addParents": add_parent_id, "removeParents": remove_parent_id, "fields": "id,parents"},
        timeout=30,
    )
    r.raise_for_status()
    return r.json()

def process_manual_drop():
    """
    Poll a Google Drive folder named 'manual drop' for PNG uploads.
    - Move PNG into png_library_images/manual_drop/
    - Generate WEBP preview with watermark/footer and upload into png_library_previews/manual_drop/
    - Append ultradata.xlsx (GitHub repo1)
    """
    token = get_drive_token()
    manual_root = drive_folder(token, "manual drop")
    files = drive_list_pngs(token, manual_root)
    if not files:
        log("Manual drop: no PNG files found")
        return

    png_root  = drive_folder(token, "png_library_images")
    prev_root = drive_folder(token, "png_library_previews")
    dest_png_folder  = drive_folder(token, "manual_drop", png_root)
    dest_prev_folder = drive_folder(token, "manual_drop", prev_root)

    temp_dir = WORKING_DIR / "manual_drop_tmp"
    temp_dir.mkdir(parents=True, exist_ok=True)

    uploaded = []
    for f in files:
        try:
            fid = f["id"]
            name = f.get("name", "untitled.png")
            parents = (f.get("parents") or [])
            old_parent = parents[0] if parents else manual_root

            # Download bytes for preview creation
            png_bytes = drive_download_bytes(token, fid)
            local_png = temp_dir / name
            local_png.write_bytes(png_bytes)

            # Create WEBP preview (watermark+footer) from local PNG
            _jpg_bytes, webp_bytes, pw, ph = make_previews(local_png)
            wr = drive_upload(token, dest_prev_folder, Path(name).stem + ".webp", webp_bytes, "image/webp")
            drive_share(token, wr["id"])

            # Move original PNG inside Drive (no re-upload)
            drive_move(token, fid, dest_png_folder, old_parent)
            drive_share(token, fid)

            subject = Path(name).stem.replace("_", " ").replace("-", " ").title()
            uploaded.append({
                "category": "manual_drop",
                "subcategory": "manual_drop",
                "subject_name": subject,
                "filename": name,
                "png_file_id": fid,
                "jpg_file_id": "",
                "webp_file_id": wr["id"],
                "download_url": download_url(fid),
                "preview_url": preview_url(wr["id"], 800),
                "preview_url_small": preview_url(wr["id"], 400),
                "webp_preview_url": preview_url(wr["id"], 800),
                "preview_w": pw,
                "preview_h": ph,
                "date_added": datetime.now().strftime("%Y-%m-%d"),
            })

            log(f"Manual drop OK: {name}")
        except Exception as e:
            log(f"Manual drop FAIL: {f.get('name','?')}: {e}")

    if uploaded:
        _append_ultradata_and_push(uploaded)
        log(f"Manual drop appended: {len(uploaded)}")

def make_previews(png_path):
    """Diagonal watermarked JPG + WebP previews with EXIF copyright."""
    import piexif
    with Image.open(png_path).convert("RGBA") as img_rgba:
        w, h = img_rgba.size
        if max(w, h) > 800:
            r        = 800 / max(w, h)
            img_rgba = img_rgba.resize((int(w * r), int(h * r)), Image.LANCZOS)
        w, h = img_rgba.size
        bg  = Image.new("RGB", (w, h), (255, 255, 255))
        drw = ImageDraw.Draw(bg)
        for ry in range(0, h, 20):
            for cx in range(0, w, 20):
                if (ry // 20 + cx // 20) % 2 == 1:
                    drw.rectangle([cx, ry, cx + 20, ry + 20], fill=(232, 232, 232))
        bg.paste(img_rgba.convert("RGB"), mask=img_rgba.split()[3])
        try:
            fnt = ImageFont.truetype(
                "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", 13)
        except Exception:
            fnt = ImageFont.load_default()
        wm_layer = Image.new("RGBA", (w, h), (0, 0, 0, 0))
        wm_draw  = ImageDraw.Draw(wm_layer)
        for ry in range(-h, h + 110, 110):
            for cx in range(-w, w + 110, 110):
                wm_draw.text((cx, ry), WATERMARK_TEXT, fill=(0, 0, 0, 42), font=fnt)
        wm_rot  = wm_layer.rotate(-30, expand=False)
        bg_rgba = bg.convert("RGBA")
        bg_rgba.alpha_composite(wm_rot)
        bg = bg_rgba.convert("RGB")
        drw2 = ImageDraw.Draw(bg)
        drw2.rectangle([0, h - 36, w, h], fill=(13, 13, 20))
        try:
            fnt2 = ImageFont.truetype(
                "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", 15)
        except Exception:
            fnt2 = fnt
        drw2.text((w // 2 - 72, h - 26), WATERMARK_TEXT, fill=(245, 166, 35), font=fnt2)
        exif_bytes = piexif.dump({
            "0th": {
                piexif.ImageIFD.Copyright: WATERMARK_TEXT.encode(),
                piexif.ImageIFD.Artist:    SITE_NAME.encode(),
                piexif.ImageIFD.Software:  b"UltraPNG Pipeline V7.0",
            }
        })
        jpg_buf = io.BytesIO()
        bg.save(jpg_buf, "JPEG", quality=85, optimize=True, exif=exif_bytes)
        webp_buf = io.BytesIO()
        bg.save(webp_buf, "WEBP", quality=82, method=6)
    return jpg_buf.getvalue(), webp_buf.getvalue(), w, h

# ══════════════════════════════════════════════════════════════
# SKIP SET
# ══════════════════════════════════════════════════════════════
def load_skip_set_from_json():
    log("Building skip-set from REPO2 JSON files...")
    skip_set = set()
    if not REPO2_DIR.exists():
        try:
            repo_url = f"https://x-access-token:{GITHUB_TOKEN}@github.com/{GITHUB_REPO2}.git"
            subprocess.run(
                ["git", "clone", "--depth", "1", "--filter=blob:none",
                 "--sparse", repo_url, str(REPO2_DIR)],
                capture_output=True, check=True)
            subprocess.run(["git", "sparse-checkout", "init", "--cone"],
                           cwd=str(REPO2_DIR), capture_output=True, check=True)
            subprocess.run(["git", "sparse-checkout", "set", "data"],
                           cwd=str(REPO2_DIR), capture_output=True, check=True)
        except Exception:
            log("  Skip-set: REPO2 empty — starting fresh")
            return skip_set
    data_dir = REPO2_DIR / "data"
    if not data_dir.exists():
        log("  No data dir — starting fresh")
        return skip_set
    for jf in data_dir.glob("*.json"):
        if jf.name.startswith("_"):
            continue
        try:
            for e in json.loads(jf.read_text("utf-8")):
                fn = e.get("filename", "")
                if fn:
                    skip_set.add(fn)
        except Exception:
            pass
    log(f"  Skip-set (REPO2): {len(skip_set)} already-done filenames")
    return skip_set


def load_skip_set_from_ultradata() -> set:
    """
    Read ultradata.xlsx from REPO1 and return all filenames already generated.
    This catches images generated+uploaded but not yet SEO-processed in REPO2.
    Without this, Section 1 would re-generate those images on the next daily run.
    """
    skip_set = set()
    token = GITHUB_TOKEN_REPO1 or ""
    repo  = (GITHUB_REPO1 or "").strip()
    if not token or not repo:
        log("  Skip-set (ultradata): GITHUB_TOKEN_REPO1 not set — skipping")
        return skip_set

    repo_url  = f"https://x-access-token:{token}@github.com/{repo}.git"
    xrepo_dir = WORKING_DIR / "repo1_skipcheck"

    try:
        if xrepo_dir.exists() and (xrepo_dir / ".git").exists():
            subprocess.run(["git", "remote", "set-url", "origin", repo_url],
                           cwd=str(xrepo_dir), capture_output=True)
            subprocess.run(["git", "pull", "--rebase", "--autostash"],
                           cwd=str(xrepo_dir), capture_output=True)
        else:
            shutil.rmtree(str(xrepo_dir), ignore_errors=True)
            subprocess.run(["git", "clone", "--depth", "1", repo_url, str(xrepo_dir)],
                           capture_output=True, check=True)
    except Exception as e:
        log(f"  Skip-set (ultradata): clone failed ({e}) — skipping")
        return skip_set

    xlsx_path = xrepo_dir / ULTRADATA_XLSX_NAME
    if not xlsx_path.exists():
        log("  Skip-set (ultradata): ultradata.xlsx not found — starting fresh")
        return skip_set

    try:
        import openpyxl
    except Exception:
        subprocess.run([sys.executable, "-m", "pip", "install", "-q", "openpyxl>=3.1.2"],
                       capture_output=True, check=True)
        import openpyxl

    try:
        wb = openpyxl.load_workbook(str(xlsx_path), read_only=True)
        ws = wb.active
        if ws.max_row < 2:
            log("  Skip-set (ultradata): sheet empty")
            return skip_set
        headers = [str(c.value or "").strip().lower()
                   for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if "filename" not in headers:
            log("  Skip-set (ultradata): no filename column")
            return skip_set
        col_idx = headers.index("filename")
        for row in ws.iter_rows(min_row=2, values_only=True):
            fn = row[col_idx] if col_idx < len(row) else None
            if fn:
                skip_set.add(str(fn).strip())
        wb.close()
    except Exception as e:
        log(f"  Skip-set (ultradata): read failed ({e})")
        return skip_set

    log(f"  Skip-set (ultradata): {len(skip_set)} already-generated filenames")
    return skip_set


def load_prompts():
    log("Loading prompts...")
    if GITHUB_REPO1 and not PROJECT_DIR.exists():
        subprocess.run(
            ["git", "clone", "--depth", "1",
             f"https://github.com/{GITHUB_REPO1}", str(PROJECT_DIR)],
            capture_output=True)
    if PROJECT_DIR.exists():
        sys.path.insert(0, str(PROJECT_DIR))
        try:
            from prompts.prompt_engine import load_all_prompts
            prompts = load_all_prompts(str(PROJECT_DIR / "prompts" / "splits"))
            log(f"Loaded {len(prompts)} prompts")
            return prompts
        except Exception as e:
            log(f"Prompt error: {e}")
    log("FATAL: No prompts!")
    return []

# ══════════════════════════════════════════════════════════════
# PHASE 1 — FLUX.2-Klein-4B  (loads from HuggingFace)
# ══════════════════════════════════════════════════════════════
def phase1_generate(batch, skip_set):
    ckpt = load_checkpoint("phase1_generated")
    if ckpt:
        return ckpt

    log("=" * 56)
    log("PHASE 1: FLUX.2-Klein-4B — Image Generation")
    log(f"  Loading from HuggingFace: {FLUX_HF_ID}")
    log("=" * 56)

    from diffusers import Flux2KleinPipeline

    log(f"  Loading: {FLUX_HF_ID}")
    log("  (First run: downloads ~8GB — ~5 min)")
    pipe = Flux2KleinPipeline.from_pretrained(
        FLUX_HF_ID,
        torch_dtype=torch.bfloat16,
    )
    pipe.enable_model_cpu_offload(gpu_id=0)
    pipe.set_progress_bar_config(disable=True)
    log(f"FLUX.2 loaded | Batch: {len(batch)} | Skip: {len(skip_set)}\n")

    generated, skipped, t0 = [], 0, time.time()

    for i, item in enumerate(batch):
        fname = item["filename"]

        if fname in skip_set:
            skipped += 1
            if skipped <= 3 or skipped % 50 == 0:
                log(f"  SKIP [{i+1}/{len(batch)}] {fname}")
            continue

        try:
            out_dir = GENERATED_DIR / item["category"] / item.get("subcategory", "general")
            out_dir.mkdir(parents=True, exist_ok=True)
            out = out_dir / fname

            if out.exists():
                generated.append({"path": str(out), "item": item})
                continue

            gen    = torch.Generator("cpu").manual_seed(item["seed"])
            prompt = enhance_prompt(item["prompt"], item.get("category", ""))
            img = pipe(
                prompt=prompt,
                num_inference_steps=4,
                guidance_scale=1.0,
                height=1024, width=1024,
                generator=gen,
            ).images[0]

            # Blank image check
            arr    = np.array(img)
            n_px   = arr.shape[0] * arr.shape[1]
            thresh = int(n_px * 0.003)
            if arr.std() < 5 or (arr < 250).sum() < thresh:
                log(f"  Retry (blank): {fname}")
                gen2 = torch.Generator("cpu").manual_seed(item["seed"] + 99)
                img  = pipe(
                    prompt=prompt,
                    num_inference_steps=4,
                    guidance_scale=1.0,
                    height=1024, width=1024,
                    generator=gen2,
                ).images[0]

            img.save(str(out), "PNG", compress_level=9)
            generated.append({"path": str(out), "item": item})

            done = len(generated)
            rate = done / (time.time() - t0)
            eta  = (len(batch) - i - 1) / rate / 60 if rate > 0 else 0
            log(f"  [{i+1}/{len(batch)}] OK {fname} | {item['category']} | ETA {eta:.0f}min")

        except torch.cuda.OutOfMemoryError:
            torch.cuda.empty_cache()
            log(f"  OOM: {fname}")
        except Exception as e:
            log(f"  FAIL {fname}: {e}")

    log("\n  Deleting FLUX.2 -> freeing VRAM + disk cache...")
    del pipe
    free_memory()

    # Delete FLUX HF cache to free disk
    import shutil as _shutil
    for _cache_sub in HF_CACHE.iterdir():
        if _cache_sub.is_dir():
            _name = _cache_sub.name.lower()
            if "flux" in _name or "black-forest" in _name:
                _shutil.rmtree(str(_cache_sub), ignore_errors=True)
                log(f"  Deleted FLUX cache: {_cache_sub.name}")
    _used = sum(f.stat().st_size for f in HF_CACHE.rglob("*") if f.is_file()) / 1e9
    log(f"  HF cache after cleanup: {_used:.1f}GB")

    save_checkpoint("phase1_generated", generated)
    log(f"PHASE 1 DONE — Generated: {len(generated)} | Skipped: {skipped}\n")
    return generated

# ══════════════════════════════════════════════════════════════
# PHASE 3 — RMBG-2.0 ONNX  (downloads from HuggingFace)
# ══════════════════════════════════════════════════════════════
def phase3_bg_remove(posts):
    ckpt = load_checkpoint("phase3_transparent")
    if ckpt:
        return ckpt

    log("=" * 56)
    log("PHASE 3: BiRefNet_HR — 2048x2048 FP16 — Background Removal (GPU)")
    log(f"  Loading from HuggingFace: {RMBG_HF_ID}")
    log("=" * 56)

    if not posts:
        return []

    from torchvision import transforms
    from transformers import AutoModelForImageSegmentation

    rmbg_model = AutoModelForImageSegmentation.from_pretrained(
        RMBG_HF_ID,
        trust_remote_code=True,
        cache_dir=str(HF_CACHE),
    )
    device = "cuda" if torch.cuda.is_available() else "cpu"
    torch.set_float32_matmul_precision("high")
    rmbg_model = rmbg_model.to(device).eval().half()  # FP16 official recommendation
    log(f"  BiRefNet_HR loaded on {device.upper()} | FP16 | 2048x2048\n")

    # Official inference code from ZhengPeng7/BiRefNet_HR HuggingFace page
    transform_img = transforms.Compose([
        transforms.Resize((2048, 2048)),
        transforms.ToTensor(),
        transforms.Normalize([0.485, 0.456, 0.406],
                             [0.229, 0.224, 0.225]),
    ])

    def remove_bg(pil_img):
        ow, oh   = pil_img.size
        inp      = transform_img(pil_img.convert("RGB")).unsqueeze(0).to(device).half()
        with torch.no_grad():
            preds = rmbg_model(inp)[-1].sigmoid().cpu()
        pred     = preds[0].squeeze()
        mask_pil = transforms.ToPILImage()(pred).resize((ow, oh), Image.LANCZOS)
        result   = pil_img.convert("RGBA")
        result.putalpha(mask_pil)
        return result

    result_posts, t0 = [], time.time()

    for i, post in enumerate(posts):
        path = Path(post["approved_path"])
        try:
            rel = path.relative_to(APPROVED_DIR)
            out = TRANSPARENT_DIR / rel
            out.parent.mkdir(parents=True, exist_ok=True)

            if out.exists():
                result_posts.append({**post, "transparent_path": str(out)})
                continue

            img    = Image.open(str(path)).convert("RGB")
            result = remove_bg(img)
            result.save(str(out), "PNG", compress_level=9)
            result_posts.append({**post, "transparent_path": str(out)})

            if (i + 1) % 20 == 0:
                log(f"  BG done: {i+1}/{len(posts)} | {(i+1)/(time.time()-t0):.2f}/s")

        except Exception as e:
            log(f"  RMBG FAIL {path.name}: {e}")

    # ── DELETE model → free VRAM + disk for Phase 4 ──
    log("\n  Deleting RMBG-2.0 → freeing VRAM + disk cache...")
    del rmbg_model, transform_img
    free_memory()

    # Delete BiRefNet HF cache
    import shutil as _shutil
    for _cache_sub in HF_CACHE.iterdir():
        if _cache_sub.is_dir():
            _name = _cache_sub.name.lower()
            if "rmbg" in _name or "briaai" in _name or "birefnet" in _name:
                _shutil.rmtree(str(_cache_sub), ignore_errors=True)
                log(f"  Deleted BiRefNet cache: {_cache_sub.name}")

    # Delete APPROVED_DIR images — transparent/ has the final copies
    if APPROVED_DIR.exists():
        _shutil.rmtree(str(APPROVED_DIR), ignore_errors=True)
        APPROVED_DIR.mkdir(parents=True, exist_ok=True)
        log("  Deleted approved/ images (transparent/ copies kept)")

    _used = sum(f.stat().st_size for f in HF_CACHE.rglob("*") if f.is_file()) / 1e9
    log(f"  HF cache after cleanup: {_used:.1f}GB")

    save_checkpoint("phase3_transparent", result_posts)
    log(f"PHASE 3 DONE — Transparent PNGs: {len(result_posts)}\n")
    return result_posts


# ══════════════════════════════════════════════════════════════
# PHASE 4 — Google Drive Upload
# ══════════════════════════════════════════════════════════════
def _drive_list_folder_names(token, folder_id):
    """Return dict {filename: file_id} for ALL files in a Drive folder (handles pagination)."""
    h = {"Authorization": f"Bearer {token}"}
    q = f"'{folder_id}' in parents and trashed=false"
    result = {}
    page_token = None
    while True:
        params = {"q": q, "pageSize": 1000,
                  "fields": "nextPageToken,files(id,name)",
                  "pageToken": page_token} if page_token else                  {"q": q, "pageSize": 1000, "fields": "nextPageToken,files(id,name)"}
        r = req.get("https://www.googleapis.com/drive/v3/files",
                    headers=h, params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
        for f in data.get("files", []):
            result[f["name"]] = f["id"]
        page_token = data.get("nextPageToken")
        if not page_token:
            break
    return result


def phase4_upload(posts):
    ckpt = load_checkpoint("phase4_uploaded")
    if ckpt:
        return ckpt

    log("=" * 56)
    log("PHASE 4: Google Drive Upload + URL Capture (PNG + WEBP only)")
    log("=" * 56)

    if not posts:
        return []

    token     = get_drive_token()
    fcache    = {}          # folder IDs
    dcache    = {}          # existing filenames per folder_id  {folder_id: {name: fid}}
    png_root  = drive_folder(token, "png_library_images")
    prev_root = drive_folder(token, "png_library_previews")
    log(f"Drive ready. Uploading {len(posts)} images...\n")

    uploaded, skipped_drive, t0 = [], 0, time.time()

    for i, post in enumerate(posts):
        path = Path(post["transparent_path"])

        if i > 0 and i % 50 == 0:
            token = get_drive_token()

        try:
            cat = post["category"]
            sub = post.get("subcategory", "general")
            key = f"{cat}/{sub}"

            if key not in fcache:
                cp = drive_folder(token, cat, png_root)
                sp = drive_folder(token, sub, cp)
                cr = drive_folder(token, cat, prev_root)
                sr = drive_folder(token, sub, cr)
                fcache[f"p_{key}"] = sp
                fcache[f"r_{key}"] = sr
                # ── Pre-load existing filenames for dedup ──
                dcache[f"p_{key}"] = _drive_list_folder_names(token, sp)
                dcache[f"r_{key}"] = _drive_list_folder_names(token, sr)
                log(f"  Drive folder '{key}': {len(dcache[f'p_{key}'])} existing PNGs")

            png_folder_id  = fcache[f"p_{key}"]
            webp_folder_id = fcache[f"r_{key}"]
            existing_png   = dcache[f"p_{key}"]
            existing_webp  = dcache[f"r_{key}"]

            webp_name = path.stem + ".webp"

            # ── DRIVE-SIDE DEDUP ─────────────────────────────────────────
            if path.name in existing_png and webp_name in existing_webp:
                png_fid  = existing_png[path.name]
                webp_fid = existing_webp[webp_name]
                _jpg_bytes, webp_bytes, pw, ph = make_previews(path)
                skipped_drive += 1
                log(f"  DRIVE-SKIP [{i+1}] {path.name} (already exists)")
                uploaded.append({
                    **post,
                    "png_file_id":       png_fid,
                    "webp_file_id":      webp_fid,
                    "download_url":      download_url(png_fid),
                    "preview_url":       preview_url(webp_fid, 800),
                    "preview_url_small": preview_url(webp_fid, 400),
                    "webp_preview_url":  preview_url(webp_fid, 800),
                    "preview_w": pw, "preview_h": ph,
                    "date_added": datetime.now().strftime("%Y-%m-%d"),
                })
                continue

            # ── Fresh upload ─────────────────────────────────────────────
            png_bytes = path.read_bytes()
            pr = drive_upload(token, png_folder_id, path.name, png_bytes)
            drive_share(token, pr["id"])
            existing_png[path.name] = pr["id"]   # update local cache

            _jpg_bytes, webp_bytes, pw, ph = make_previews(path)
            wr = drive_upload(token, webp_folder_id, webp_name, webp_bytes, "image/webp")
            drive_share(token, wr["id"])
            existing_webp[webp_name] = wr["id"]  # update local cache

            uploaded.append({
                **post,
                "png_file_id":       pr["id"],
                "webp_file_id":      wr["id"],
                "download_url":      download_url(pr["id"]),
                "preview_url":       preview_url(wr["id"], 800),
                "preview_url_small": preview_url(wr["id"], 400),
                "webp_preview_url":  preview_url(wr["id"], 800),
                "preview_w": pw, "preview_h": ph,
                "date_added": datetime.now().strftime("%Y-%m-%d"),
            })

            if (i + 1) % 10 == 0:
                log(f"  Uploaded: {i+1}/{len(posts)} | skip={skipped_drive} | {(i+1)/(time.time()-t0):.2f}/s")
            time.sleep(0.05)

        except Exception as e:
            log(f"  Upload FAIL {path.name}: {e}")

    save_checkpoint("phase4_uploaded", uploaded)
    log(f"PHASE 4 DONE — Uploaded: {len(uploaded)} | Drive-skipped: {skipped_drive} | {(time.time()-t0)/60:.0f}min")

    # Delete transparent/ images — all uploaded to Drive already
    import shutil as _shutil
    if TRANSPARENT_DIR.exists():
        _shutil.rmtree(str(TRANSPARENT_DIR), ignore_errors=True)
        TRANSPARENT_DIR.mkdir(parents=True, exist_ok=True)
        log("  Deleted transparent/ images (uploaded to Drive)")
    log("")
    return uploaded

# ══════════════════════════════════════════════════════════════
# HTML BUILDER
# ══════════════════════════════════════════════════════════════
def md_to_html(md):
    if not md: return ""
    o = str(md)
    o = re.sub(r'^### (.+)$', r'<h3>\1</h3>', o, flags=re.MULTILINE)
    o = re.sub(r'^## (.+)$',  r'<h2>\1</h2>', o, flags=re.MULTILINE)
    o = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', o)
    def tbl(m):
        rows = [r for r in m.group(0).strip().split('\n')
                if not re.match(r'^\|[\s\-|:]+\|$', r)]
        html = ""
        for idx, row in enumerate(rows):
            cells = [c.strip() for c in row.strip('|').split('|')]
            tag   = 'th' if idx == 0 else 'td'
            html += '<tr>' + ''.join(f'<{tag}>{c}</{tag}>' for c in cells) + '</tr>'
        return f'<table class="spec-table">{html}</table>'
    o = re.sub(r'((?:^\|.+\|\n?)+)', tbl, o, flags=re.MULTILINE)
    o = re.sub(r'^- (.+)$', r'<li>\1</li>', o, flags=re.MULTILINE)
    o = re.sub(r'(<li>.*?</li>\n?)+', lambda m: f'<ul>{m.group(0)}</ul>', o)
    o = re.sub(r'^\d+\. (.+)$', r'<oli>\1</oli>', o, flags=re.MULTILINE)
    o = re.sub(r'(<oli>.*?</oli>\n?)+',
               lambda m: '<ol>' + m.group(0).replace('<oli>','<li>').replace('</oli>','</li>') + '</ol>', o)
    o = re.sub(r'\n{2,}', '</p><p>', o)
    o = f'<p>{o}</p>'
    o = re.sub(r'<p>\s*</p>', '', o)
    return o

def _head(title, desc, url, img="", kw="", webp_img=""):
    og_img = webp_img or img
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>{esc(title)}</title>
<meta name="description" content="{esc(desc)}"/>
{f'<meta name="keywords" content="{esc(kw)}"/>' if kw else ''}
<meta name="robots" content="index,follow,max-image-preview:large,max-snippet:-1"/>
<link rel="canonical" href="{url}"/>
<meta property="og:type" content="website"/>
<meta property="og:url" content="{url}"/>
<meta property="og:title" content="{esc(title)}"/>
<meta property="og:description" content="{esc(desc)}"/>
{f'<meta property="og:image" content="{esc(og_img)}"/>' if og_img else ''}
<meta property="og:site_name" content="UltraPNG"/>
<meta name="twitter:card" content="summary_large_image"/>
<link rel="icon" type="image/svg+xml" href="/favicon.svg"/>
<link rel="preconnect" href="https://fonts.googleapis.com"/>
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin/>
<link rel="stylesheet" href="https://fonts.googleapis.com/css2?family=Sora:wght@500;600;700;800&family=Outfit:wght@400;500;600;700&display=swap" media="print" onload="this.media='all'"/>
<link rel="stylesheet" href="/css/style.css?v=22"/>
<link rel="stylesheet" href="/css/png-library.css?v=10"/>"""

def _header():
    return ('<header><div class="header-inner">'
            '<a href="/" class="logo"><div class="logo-icon">U</div>'
            '<div class="logo-text"><span class="logo-name">Ultra<span>PNG</span></span>'
            '<span class="logo-sub">Free Transparent PNG Images</span></div></a>'
            '<nav><a href="/">Home</a><a href="/png-library/" class="active">PNG Library</a>'
            '<a href="/pages/about.html">About</a>'
            '<a href="/pages/contact.html" class="nav-contact">Contact</a></nav>'
            '</div></header>')

def _footer():
    return (f'<footer><div class="footer-inner">'
            f'<div class="footer-links">'
            f'<a href="/png-library/">PNG Library</a>'
            f'<a href="/pages/about.html">About</a>'
            f'<a href="/pages/contact.html">Contact</a>'
            f'<a href="/pages/terms.html">Terms</a>'
            f'<a href="/pages/privacy.html">Privacy</a></div>'
            f'<p>&copy; {datetime.now().year} UltraPNG.com — Free Transparent PNG Images.</p>'
            f'</div></footer>')

def build_item_page(post, related):
    url       = f"{SITE_URL}/png-library/{post['category']}/{post['slug']}/"
    img       = post.get("preview_url", "")
    webp_img  = post.get("webp_preview_url", "")
    desc_html = md_to_html(post.get("description", ""))
    enc_dl    = base64.b64encode(post.get("download_url", "").encode()).decode()
    tags      = [t.strip() for t in post.get("tags", "").split(",") if t.strip()]
    cat_label = post.get("subject_name", post["category"])
    tags_html = "".join(
        f'<a href="/png-library/?q={esc(t)}" class="png-tag">{esc(t)}</a>' for t in tags)
    share_text = json.dumps(f'{post.get("h1", "")} PNG Free Download — ')

    if webp_img:
        img_tag = (f'<picture>'
                   f'<source srcset="{esc(webp_img)}" type="image/webp"/>'
                   f'<img src="{esc(img)}" alt="{esc(post.get("alt_text",""))}" '
                   f'width="{post.get("preview_w",800)}" height="{post.get("preview_h",800)}" '
                   f'fetchpriority="high" onerror="this.src=\'/img/placeholder.png\'"/>'
                   f'</picture>')
    else:
        img_tag = (f'<img src="{esc(img)}" alt="{esc(post.get("alt_text",""))}" '
                   f'width="{post.get("preview_w",800)}" height="{post.get("preview_h",800)}" '
                   f'fetchpriority="high" onerror="this.src=\'/img/placeholder.png\'"/>')

    rel_html = "".join(
        f'<a href="/png-library/{r["category"]}/{r["slug"]}/" class="png-related-card">'
        f'<div class="png-related-thumb">'
        f'<img src="{esc(r.get("preview_url_small", r.get("preview_url","")))} " '
        f'alt="{esc(r.get("h1",""))}" loading="lazy" width="300" height="300" '
        f'onerror="this.parentNode.style.display=\'none\'"/></div>'
        f'<span class="png-related-name">{esc(r.get("h1",""))}</span></a>\n'
        for r in related[:24]
    )

    schema = json.dumps({
        "@context": "https://schema.org",
        "@graph": [
            {"@type": "ImageObject", "name": post.get("h1",""),
             "description": post.get("meta_desc",""),
             "contentUrl": post.get("download_url",""), "thumbnailUrl": img,
             "encodingFormat": "image/png", "isAccessibleForFree": True,
             "datePublished": post.get("date_added",""),
             "license": f"{SITE_URL}/pages/terms.html",
             "publisher": {"@type": "Organization", "name": "UltraPNG", "url": SITE_URL + "/"},
             "keywords": ", ".join(tags)},
            {"@type": "BreadcrumbList", "itemListElement": [
                {"@type": "ListItem", "position": 1, "name": "Home", "item": SITE_URL + "/"},
                {"@type": "ListItem", "position": 2, "name": "PNG Library", "item": f"{SITE_URL}/png-library/"},
                {"@type": "ListItem", "position": 3, "name": cat_label, "item": f"{SITE_URL}/png-library/{post['category']}/"},
                {"@type": "ListItem", "position": 4, "name": post.get("h1","")},
            ]},
            {"@type": "FAQPage", "mainEntity": [
                {"@type": "Question", "name": f"Is this {cat_label} PNG free?",
                 "acceptedAnswer": {"@type": "Answer", "text": "Yes — 100% free. No account required."}},
                {"@type": "Question", "name": "Can I use this PNG in Canva?",
                 "acceptedAnswer": {"@type": "Answer", "text": "Yes. Upload via My Files in Canva."}},
                {"@type": "Question", "name": "Does the downloaded PNG have a watermark?",
                 "acceptedAnswer": {"@type": "Answer", "text": "No. Downloaded file is completely clean."}},
                {"@type": "Question", "name": "Can I print this on flex banners?",
                 "acceptedAnswer": {"@type": "Answer", "text": "Yes. HD PNG is print-ready for flex/hoarding."}},
            ]},
        ]
    })

    return f"""{_head(post.get("title",""), post.get("meta_desc",""), url, img, post.get("tags",""), webp_img)}
{f'<link rel="preload" as="image" href="{esc(webp_img or img)}" fetchpriority="high"/>' if (webp_img or img) else ''}
<script type="application/ld+json">{schema}</script>
</head>
<body class="png-item-page">
{_header()}
<nav class="breadcrumb" aria-label="Breadcrumb">
  <a href="/">Home</a><span>&rsaquo;</span>
  <a href="/png-library/">PNG Library</a><span>&rsaquo;</span>
  <a href="/png-library/{post['category']}/">{esc(cat_label)}</a><span>&rsaquo;</span>
  <span aria-current="page">{esc(post.get("h1",""))}</span>
</nav>
<div class="png-item-wrap"><div class="png-item-layout"><div class="png-item-center">
<div class="png-item-top-row">
<div class="png-item-img-col"><div class="png-img-card">{img_tag}</div></div>
<div class="png-item-info-col">
  <h1 class="png-dl-title">{esc(post.get("h1",""))}</h1>
  <div class="png-dl-btn-wrap" id="bW">
    <button class="png-dl-btn" onclick="startCD()" aria-label="Download PNG Free">
      <svg width="20" height="20" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5">
        <path d="M21 15v4a2 2 0 01-2 2H5a2 2 0 01-2-2v-4M7 10l5 5 5-5M12 15V3"/>
      </svg>Download PNG Free
    </button>
  </div>
  <div class="png-timer-wrap" id="tW">
    <svg class="png-timer-circle" viewBox="0 0 120 120">
      <circle class="png-timer-ring-bg" cx="60" cy="60" r="52"/>
      <circle class="png-timer-ring" id="tR" cx="60" cy="60" r="52"
              stroke-dasharray="326.73" stroke-dashoffset="0"/>
      <text class="png-timer-text" x="60" y="68" text-anchor="middle" id="tN">15</text>
    </svg>
    <div class="png-timer-label">Preparing your download...</div>
  </div>
  <div class="png-dl-ready" id="rW">
    <button class="png-dl-ready-btn" onclick="triggerDL()">
      <svg width="22" height="22" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2.5">
        <path d="M22 11.08V12a10 10 0 11-5.93-9.14"/>
        <polyline points="22 4 12 14.01 9 11.01"/>
      </svg>Download Now!
    </button>
  </div>
  <div class="png-share-bar">
    <button class="png-share-btn png-share-wa" onclick="shareWA()">WhatsApp</button>
    <button class="png-share-btn png-share-cp" onclick="copyLink()" id="cpBtn">Copy Link</button>
  </div>
</div>
</div>
<div class="png-item-desc-col">
  <div class="png-content">{desc_html}</div>
  <div class="png-tags">{tags_html}</div>
  {f'<div class="png-related-section"><h2>More {esc(cat_label)} PNG Images</h2><div class="png-related-masonry">{rel_html}</div></div>' if rel_html else ''}
</div>
</div></div></div>
{_footer()}
<script>
var _d="{enc_dl}",_s=15,_pg="{esc(url)}";
function startCD(){{
  document.getElementById('bW').style.display='none';
  document.getElementById('tW').classList.add('active');
  var r=document.getElementById('tR'),n=document.getElementById('tN'),t=326.73,c=_s;
  var iv=setInterval(function(){{
    c--;n.textContent=c;r.style.strokeDashoffset=t*(1-c/_s);
    if(c<=0){{clearInterval(iv);document.getElementById('tW').classList.remove('active');
      document.getElementById('rW').classList.add('active');setTimeout(triggerDL,300);}}
  }},1000);
}}
function triggerDL(){{
  try{{var u=atob(_d);var a=document.createElement('a');a.href=u;a.download='';
    a.style.display='none';document.body.appendChild(a);a.click();
    setTimeout(function(){{document.body.removeChild(a);}},200);}}catch(e){{console.error(e);}}
}}
function shareWA(){{window.open('https://wa.me/?text='+encodeURIComponent({share_text}+_pg),'_blank','noopener');}}
function copyLink(){{navigator.clipboard&&navigator.clipboard.writeText(_pg).then(function(){{
  var b=document.getElementById('cpBtn');b.textContent='Copied!';
  setTimeout(function(){{b.textContent='Copy Link';}},2000);}});}}
</script>
</body></html>"""

def build_category_page(cat, items):
    label = items[0].get("subject_name", cat) if items else cat
    url   = f"{SITE_URL}/png-library/{cat}/"
    img   = items[0].get("preview_url", "") if items else ""
    total = len(items)

    def _card(idx, it):
        thumb = ""
        if it.get("preview_url_small") or it.get("preview_url"):
            sw = esc(it.get("webp_preview_url","") or it.get("preview_url_small",""))
            sj = esc(it.get("preview_url_small", it.get("preview_url","")))
            al = esc(it.get("h1",""))
            thumb = (f'<picture><source srcset="{sw}" type="image/webp"/>'
                     f'<img src="{sj}" alt="{al}" loading="lazy" width="400" height="400"/></picture>')
        pg  = idx // ITEMS_PER_PAGE
        return (f'<a href="/png-library/{it["category"]}/{it["slug"]}/" '
                f'class="mpng-item-card cat-item" data-pg="{pg}" title="{esc(it.get("h1",""))}">'
                f'<div class="mpng-item-thumb">{thumb}</div>'
                f'<div class="mpng-item-label">{esc(it.get("h1",""))}</div></a>\n')

    cards = "".join(_card(idx, it) for idx, it in enumerate(items))
    tp    = max(1, (total + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE)
    pag   = (f'<nav class="pg-nav">'
             f'<button class="pg-btn pg-btn-prev" id="cP" onclick="catPage(_cp-1)" disabled>&#8592; Previous</button>'
             f'<span class="pg-info" id="cI">Page 1 of {tp}</span>'
             f'<button class="pg-btn pg-btn-next" id="cN" onclick="catPage(_cp+1)">Next &#8594;</button>'
             f'</nav>') if tp > 1 else ""
    title = f"{label} PNG Images Free Download ({total}+) | UltraPNG"
    desc  = f"Download {total}+ free {label} transparent PNG. HD Photoshop Canva."
    return (f'{_head(title, desc, url, img)}\n</head><body>{_header()}\n'
            f'<nav class="breadcrumb"><a href="/">Home</a><span>&rsaquo;</span>'
            f'<a href="/png-library/">PNG Library</a><span>&rsaquo;</span>'
            f'<span>{esc(label)}</span></nav>\n'
            f'<div class="cat-pg-wrap">'
            f'<h1 class="cat-pg-title">{esc(label)} PNG Images <span class="sec-count">({total})</span></h1>'
            f'<div class="png-masonry" id="catMasonry">{cards}</div>{pag}</div>\n'
            f'{_footer()}\n'
            f'<script>var _cp=0,_tp={tp};'
            f'function catPage(p){{if(p<0||p>=_tp)return;_cp=p;'
            f'document.querySelectorAll(".cat-item").forEach(function(c){{'
            f'c.style.display=parseInt(c.dataset.pg||0)===p?"":"none";}});'
            f'var P=document.getElementById("cP"),N=document.getElementById("cN");'
            f'if(P)P.disabled=p===0;if(N)N.disabled=p===_tp-1;'
            f'var I=document.getElementById("cI");if(I)I.textContent="Page "+(p+1)+" of "+_tp;}}'
            f'catPage(0);</script></body></html>')

def build_main_page(all_data):
    url    = f"{SITE_URL}/png-library/"
    by_cat = {}
    for item in all_data:
        by_cat.setdefault(item["category"], []).append(item)
    cats   = sorted(by_cat.keys())
    ti, tc = len(all_data), len(cats)

    def _cat_card(cat):
        items = by_cat[cat]
        name  = esc(items[0].get("subject_name", cat).lower())
        tags  = esc(",".join(set(
            t.strip() for it in items[:3]
            for t in it.get("tags","").lower().split(",") if t.strip()
        ))[:200])
        previews = "".join(
            f'<img src="{esc(it.get("preview_url_small",it.get("preview_url","")))} " '
            f'alt="{esc(items[0].get("subject_name",cat))}" loading="lazy" '
            f'width="200" height="200" onerror="this.remove()"/>'
            for it in items[:4] if it.get("preview_url_small") or it.get("preview_url")
        )
        return (f'<a href="/png-library/{cat}/" class="mpng-cat-card" '
                f'data-cat="{name}" data-search="{tags}">'
                f'<div class="mpng-cat-previews">{previews}</div>'
                f'<div class="mpng-cat-footer">'
                f'<span class="mpng-cat-name">{esc(items[0].get("subject_name",cat))}</span>'
                f'<span class="mpng-cat-cnt">{len(items)} images</span>'
                f'</div></a>\n')

    cat_cards = "".join(_cat_card(cat) for cat in cats)

    recent = sorted(all_data, key=lambda x: x.get("date_added",""), reverse=True)[:ITEMS_PER_PAGE * 5]

    def _rec_card(idx, it):
        thumb = ""
        if it.get("preview_url_small") or it.get("preview_url"):
            src = esc(it.get("preview_url_small", it.get("preview_url","")))
            alt = esc(it.get("h1",""))
            thumb = f'<img src="{src}" alt="{alt}" loading="lazy" width="400" height="400"/>'
        return (f'<a href="/png-library/{it["category"]}/{it["slug"]}/" '
                f'class="mpng-item-card recent-item" data-rpg="{idx // ITEMS_PER_PAGE}">'
                f'<div class="mpng-item-thumb">{thumb}</div>'
                f'<div class="mpng-item-label">{esc(it.get("h1",""))}</div>'
                f'<div class="mpng-item-cat">{esc(it.get("subject_name",it["category"]))}</div></a>\n')

    rec_cards = "".join(_rec_card(idx, it) for idx, it in enumerate(recent))
    trp  = max(1, (len(recent) + ITEMS_PER_PAGE - 1) // ITEMS_PER_PAGE)
    pag  = (f'<nav class="pg-nav">'
            f'<button class="pg-btn pg-btn-prev" id="rP" onclick="recentPage(_rcp-1)" disabled>&#8592; Previous</button>'
            f'<span class="pg-info" id="rI">Page 1 of {trp}</span>'
            f'<button class="pg-btn pg-btn-next" id="rN" onclick="recentPage(_rcp+1)">Next &#8594;</button>'
            f'</nav>') if trp > 1 else ""
    title = f"UltraPNG {ti}+ Free Transparent PNG Images HD"
    desc  = f"Download {ti}+ free HD transparent PNG. {tc} categories. No watermark no signup."
    img   = all_data[0].get("preview_url","") if all_data else ""
    return (f'{_head(title, desc, url, img)}\n</head><body>{_header()}\n'
            f'<div class="mpng-hero">'
            f'<div class="mpng-hero-badge">&#127881; 100% Free &#8212; No Signup &#8212; Updated Daily</div>'
            f'<h1 class="mpng-hero-title">Ultra<span>PNG</span></h1>'
            f'<p class="mpng-hero-sub">Download <strong>{ti}+</strong> free HD transparent PNG. No watermark. No signup.</p>'
            f'<div class="mpng-stats">'
            f'<div class="mpng-stat"><div class="mpng-stat-num">{ti}+</div><div class="mpng-stat-lbl">PNG Images</div></div>'
            f'<div class="mpng-stat"><div class="mpng-stat-num">{tc}</div><div class="mpng-stat-lbl">Categories</div></div>'
            f'<div class="mpng-stat"><div class="mpng-stat-num">100%</div><div class="mpng-stat-lbl">Free</div></div>'
            f'<div class="mpng-stat"><div class="mpng-stat-num">HD</div><div class="mpng-stat-lbl">Quality</div></div>'
            f'</div>'
            f'<div class="mpng-search-wrap">'
            f'<input type="search" id="mpngSearch" '
            f'placeholder="Search PNG... (Fish, Car, Wedding, Flower, Tool...)" '
            f'oninput="filterMpng(this.value)" aria-label="Search PNG images"/>'
            f'</div></div>\n'
            f'<div class="mpng-section"><div class="mpng-section-head">'
            f'<h2 class="mpng-section-title">Browse <span>Categories</span></h2>'
            f'<span>{tc} collections</span></div>'
            f'<div class="mpng-cat-grid" id="mpngCatGrid">{cat_cards}</div></div>\n'
            f'<div class="mpng-section"><div class="mpng-section-head">'
            f'<h2 class="mpng-section-title">Recently <span>Added</span></h2></div>'
            f'<div class="mpng-masonry" id="mpngMasonry">{rec_cards}</div>{pag}</div>\n'
            f'{_footer()}\n'
            f'<script>var _rcp=0,_rtp={trp};'
            f'function filterMpng(q){{q=(q||"").trim().toLowerCase();'
            f'document.querySelectorAll(".mpng-cat-card").forEach(function(c){{'
            f'var n=(c.getAttribute("data-cat")||"").toLowerCase();'
            f'var s=(c.getAttribute("data-search")||"").toLowerCase();'
            f'c.style.display=q.length<2||n.includes(q)||s.includes(q)?"":"none";}});}}'
            f'function recentPage(p){{if(p<0||p>=_rtp)return;_rcp=p;'
            f'document.querySelectorAll(".recent-item").forEach(function(c){{'
            f'c.style.display=parseInt(c.dataset.rpg||0)===p?"":"none";}});'
            f'var P=document.getElementById("rP"),N=document.getElementById("rN");'
            f'if(P)P.disabled=p===0;if(N)N.disabled=p===_rtp-1;'
            f'var I=document.getElementById("rI");if(I)I.textContent="Page "+(p+1)+" of "+_rtp;}}'
            f'recentPage(0);</script></body></html>')

def build_sitemaps(all_data, out_dir):
    today   = datetime.now().strftime("%Y-%m-%d")
    entries = [
        f'<url><loc>{SITE_URL}/png-library/</loc><lastmod>{today}</lastmod>'
        f'<changefreq>daily</changefreq><priority>0.9</priority></url>']
    for cat in sorted(set(i["category"] for i in all_data)):
        entries.append(
            f'<url><loc>{SITE_URL}/png-library/{cat}/</loc><lastmod>{today}</lastmod>'
            f'<changefreq>weekly</changefreq><priority>0.8</priority></url>')
    for item in all_data:
        img_tag = (f'<image:image><image:loc>{esc(item["preview_url"])}</image:loc>'
                   f'<image:title>{esc(item.get("h1",""))}</image:title></image:image>'
                   if item.get("preview_url") else "")
        entries.append(
            f'<url><loc>{SITE_URL}/png-library/{item["category"]}/{item["slug"]}/</loc>'
            f'<lastmod>{item.get("date_added",today)}</lastmod>'
            f'<changefreq>monthly</changefreq><priority>0.7</priority>{img_tag}</url>')
    ns     = ('xmlns="http://www.sitemaps.org/schemas/sitemap/0.9" '
              'xmlns:image="http://www.google.com/schemas/sitemap-image/1.1"')
    chunks = [entries[i:i+SITEMAP_MAX_URL] for i in range(0, len(entries), SITEMAP_MAX_URL)]
    sm_files = []
    for idx, chunk in enumerate(chunks, 1):
        fname = f"sitemap-png-{idx}.xml"
        (out_dir / fname).write_text(
            f'<?xml version="1.0" encoding="UTF-8"?>\n<urlset {ns}>\n' +
            "\n".join(chunk) + '\n</urlset>', "utf-8")
        sm_files.append(fname)
    index = "\n".join(
        f'<sitemap><loc>{SITE_URL}/{f}</loc><lastmod>{today}</lastmod></sitemap>'
        for f in sm_files)
    (out_dir / "sitemap-png.xml").write_text(
        f'<?xml version="1.0" encoding="UTF-8"?>\n'
        f'<sitemapindex xmlns="http://www.sitemaps.org/schemas/sitemap/0.9">\n'
        f'{index}\n</sitemapindex>', "utf-8")
    log(f"  Sitemaps: {len(sm_files)} files ({len(entries)} URLs)")

def build_robots_txt(out_dir):
    (out_dir / "robots.txt").write_text(
        "User-agent: *\nAllow: /\n\n"
        "User-agent: GPTBot\nDisallow: /\n\n"
        "User-agent: ClaudeBot\nDisallow: /\n\n"
        "User-agent: CCBot\nDisallow: /\n\n"
        "User-agent: anthropic-ai\nDisallow: /\n\n"
        f"Sitemap: {SITE_URL}/sitemap-png.xml\n", "utf-8")

def build_llms_txt(out_dir):
    (out_dir / "llms.txt").write_text(
        f"# {SITE_NAME}\n\n> {SITE_URL} — Free Transparent PNG images\n\n"
        "## About\nFree HD transparent PNG images. Updated daily. No signup.\n\n"
        f"## Collections\n{SITE_URL}/png-library/\n", "utf-8")

def save_data_split(all_data, data_dir):
    data_dir.mkdir(parents=True, exist_ok=True)
    by_cat = {}
    for item in all_data:
        by_cat.setdefault(item["category"], []).append(item)
    for cat, items in by_cat.items():
        fname = data_dir / f"{cat}.json"
        tmp   = str(fname) + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(items, f, ensure_ascii=False, indent=2)
        os.replace(tmp, str(fname))
    (data_dir / "_index.json").write_text(
        json.dumps({cat: len(items) for cat, items in by_cat.items()},
                   ensure_ascii=False, indent=2), "utf-8")
    log(f"  Data saved: {len(by_cat)} category JSON files")

def load_all_data(data_dir):
    all_data = []
    if not data_dir.exists():
        return all_data
    for jf in data_dir.glob("*.json"):
        if jf.name.startswith("_") or jf.name == "png-library.json":
            continue
        try:
            entries = json.loads(jf.read_text("utf-8"))
            if isinstance(entries, list):
                all_data.extend(entries)
        except Exception:
            pass
    return all_data

# ══════════════════════════════════════════════════════════════
# PHASE 5 — JSON-Only Git Push to REPO2  (V7.0: NO HTML build)
# ══════════════════════════════════════════════════════════════
def phase5_build_push(new_posts):
    log("=" * 56)
    log("PHASE 5: JSON-Only Push to REPO2/data/ (V7.0 — no HTML build)")
    log("=" * 56)

    if not GITHUB_TOKEN or not GITHUB_REPO2:
        log("  WARNING: GITHUB_REPO2 / GITHUB_TOKEN_REPO2 not set — skipping")
        return
    if not new_posts:
        log("No new posts"); return

    repo_url = f"https://x-access-token:{GITHUB_TOKEN}@github.com/{GITHUB_REPO2}.git"

    # ── Sparse checkout — only data/ folder (fast, no GB of HTML) ──
    if REPO2_DIR.exists() and (REPO2_DIR / ".git").exists():
        log("REPO2 exists — pulling data/ only...")
        try:
            # BUG FIX: set-url FIRST so pull uses the authenticated token URL
            subprocess.run(["git", "remote", "set-url", "origin", repo_url],
                           cwd=str(REPO2_DIR), capture_output=True)
            subprocess.run(["git", "pull", "--rebase", "--autostash"],
                           cwd=str(REPO2_DIR), capture_output=True, check=True)
        except Exception as e:
            log(f"  Pull failed ({e}) — re-cloning sparse...")
            shutil.rmtree(str(REPO2_DIR), ignore_errors=True)
            _sparse_clone(repo_url)
    else:
        log(f"Sparse-cloning REPO2 data/ only: {GITHUB_REPO2}...")
        _sparse_clone(repo_url)

    data_dir = REPO2_DIR / "data"
    data_dir.mkdir(parents=True, exist_ok=True)

    # ── Merge new posts into existing JSON ──
    all_data = load_all_data(data_dir)
    log(f"Existing entries: {len(all_data)}")
    existing_keys = set(f"{d['category']}/{d['slug']}" for d in all_data)
    added = 0
    for post in new_posts:
        key = f"{post['category']}/{post['slug']}"
        if key not in existing_keys:
            all_data.append(post)
            existing_keys.add(key)
            added += 1
    log(f"Merged: +{added} new | Total: {len(all_data)}")

    # ── Save JSON files only ──
    save_data_split(all_data, data_dir)
    log(f"  JSON files saved to data/")

    # ── Git push — only data/ folder changes ──
    today    = datetime.now().strftime("%Y-%m-%d")
    orig_dir = os.getcwd()
    try:
        os.chdir(str(REPO2_DIR))
        subprocess.run(["git", "config", "user.name", "github-actions[bot]"],
                       check=True, capture_output=True)
        subprocess.run(["git", "config", "user.email",
                        "github-actions[bot]@users.noreply.github.com"],
                       check=True, capture_output=True)
        subprocess.run(["git", "add", "data/"], check=True, capture_output=True)
        diff = subprocess.run(["git", "diff", "--staged", "--quiet"], capture_output=True)
        if diff.returncode != 0:
            msg = f"data: +{added} images ({len(all_data)} total) [{today}]"
            subprocess.run(["git", "commit", "-m", msg], check=True, capture_output=True)
            push = subprocess.run(["git", "push"], capture_output=True, text=True)
            if push.returncode == 0:
                log(f"REPO2 JSON pushed! +{added} new | Total: {len(all_data)} images")
            else:
                log(f"Push failed: {push.stderr[:300]}")
        else:
            log("Nothing to commit — data already up to date.")
    except subprocess.CalledProcessError as e:
        log(f"Git error: {e}")
    finally:
        os.chdir(orig_dir)

def _sparse_clone(repo_url):
    """Clone only the data/ folder using git sparse-checkout (fast)."""
    shutil.rmtree(str(REPO2_DIR), ignore_errors=True)
    REPO2_DIR.mkdir(parents=True, exist_ok=True)

    r = subprocess.run(
        ["git", "clone", "--depth", "1", "--filter=blob:none",
         "--sparse", repo_url, str(REPO2_DIR)],
        capture_output=True)

    if r.returncode != 0:
        # Sparse clone not supported — fall back to regular shallow clone
        log("  Sparse clone failed — falling back to regular shallow clone...")
        shutil.rmtree(str(REPO2_DIR), ignore_errors=True)
        subprocess.run(
            ["git", "clone", "--depth", "1", repo_url, str(REPO2_DIR)],
            capture_output=True, check=True)
        return

    # Init cone mode then restrict to data/ only
    subprocess.run(
        ["git", "sparse-checkout", "init", "--cone"],
        cwd=str(REPO2_DIR), capture_output=True)
    subprocess.run(
        ["git", "sparse-checkout", "set", "data"],
        cwd=str(REPO2_DIR), capture_output=True)
    log("  Sparse clone OK — data/ folder only")



# ══════════════════════════════════════════════════════════════
# PHASE 6 — Save Run Logs to REPO1
# ══════════════════════════════════════════════════════════════
def phase6_save_logs(stats: dict):
    log("=" * 56)
    log("PHASE 6: Saving run logs to REPO1/logs/")
    log("=" * 56)

    token = GITHUB_TOKEN_REPO1 or GITHUB_TOKEN
    if not token or not GITHUB_REPO1:
        log("  No REPO1 token/repo — skipping")
        return

    repo_url = f"https://x-access-token:{token}@github.com/{GITHUB_REPO1}.git"
    try:
        if not REPO1_DIR.exists():
            subprocess.run(["git", "clone", "--depth", "1", repo_url, str(REPO1_DIR)],
                           capture_output=True, check=True)
        else:
            # BUG FIX: set-url FIRST so pull uses the authenticated token URL
            subprocess.run(["git", "remote", "set-url", "origin", repo_url],
                           cwd=str(REPO1_DIR), capture_output=True)
            subprocess.run(["git", "pull", "--rebase", "--autostash"],
                           cwd=str(REPO1_DIR), capture_output=True)

        logs_dir = REPO1_DIR / "logs"
        logs_dir.mkdir(parents=True, exist_ok=True)
        now      = datetime.now().strftime("%Y-%m-%d_%H-%M")
        log_file = logs_dir / f"{now}.log"

        summary = (
            f"ULTRAPNG PIPELINE V7.0 — RUN REPORT\n"
            f"{'='*60}\n"
            f"Date       : {datetime.now().strftime('%Y-%m-%d %H:%M IST')}\n"
            f"Batch      : {START_INDEX} -> {END_INDEX}\n"
            f"Generated  : {stats.get('generated',0)}\n"
            f"Deleted    : {stats.get('deleted',0)}\n"
            f"Approved   : {stats.get('approved',0)}\n"
            f"Transparent: {stats.get('transparent',0)}\n"
            f"Uploaded   : {stats.get('uploaded',0)}\n"
            f"Posts      : {stats.get('posts',0)}\n"
            f"Duration   : {stats.get('duration','?')}\n"
            f"Status     : {stats.get('status','unknown')}\n"
            f"{'='*60}\n\nFULL LOG\n{'='*60}\n"
        )
        full_log = summary + "\n".join(_LOG_LINES)
        log_file.write_text(full_log, "utf-8")

        for old in sorted(logs_dir.glob("????-??-??_??-??.log"))[:-30]:
            old.unlink()
        (logs_dir / "latest.log").write_text(full_log, "utf-8")

        readme = logs_dir / "README.md"
        rows   = [l for l in (readme.read_text("utf-8").split("\n")
                               if readme.exists() else []) if l.startswith("| 20")]
        rows.insert(0, (f"| {now.replace('_',' ')} "
                        f"| {stats.get('generated',0)} | {stats.get('deleted',0)} "
                        f"| {stats.get('approved',0)} | {stats.get('uploaded',0)} "
                        f"| {stats.get('posts',0)} | {stats.get('duration','?')} "
                        f"| {stats.get('status','?')} |"))
        readme.write_text(
            "# UltraPNG Pipeline — Run History\n\n"
            "| Date & Time | Generated | Deleted | Approved | Uploaded | Posts | Duration | Status |\n"
            "|-------------|-----------|---------|----------|----------|-------|----------|--------|\n"
            + "\n".join(rows[:50]) + "\n", "utf-8")

        orig_dir = os.getcwd()
        try:
            os.chdir(str(REPO1_DIR))
            subprocess.run(["git", "config", "user.name", "github-actions[bot]"],
                           check=True, capture_output=True)
            subprocess.run(["git", "config", "user.email",
                            "github-actions[bot]@users.noreply.github.com"],
                           check=True, capture_output=True)
            subprocess.run(["git", "add", "logs/"], check=True, capture_output=True)
            diff = subprocess.run(["git", "diff", "--staged", "--quiet"], capture_output=True)
            if diff.returncode != 0:
                msg  = (f"[logs] {now} | gen={stats.get('generated',0)} "
                        f"del={stats.get('deleted',0)} posts={stats.get('posts',0)} "
                        f"| {stats.get('status','?')}")
                subprocess.run(["git", "commit", "-m", msg], check=True, capture_output=True)
                push = subprocess.run(["git", "push"], capture_output=True, text=True)
                if push.returncode == 0:
                    log(f"Logs pushed! View: https://github.com/{GITHUB_REPO1}/tree/main/logs")
                else:
                    log(f"  Log push failed: {push.stderr[:100]}")
        finally:
            os.chdir(orig_dir)
    except Exception as e:
        log(f"  Log save error (non-fatal): {e}")

# ══════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════
def main():
    t0    = time.time()
    stats = {"status": "running", "generated": 0, "deleted": 0,
             "approved": 0, "transparent": 0, "uploaded": 0, "posts": 0, "duration": "?"}

    print("╔══════════════════════════════════════════════════════╗")
    print("║  UltraPNG — HuggingFace Direct Pipeline           ║")
    print("║  FLUX -> RMBG -> Drive(PNG+WEBP) -> ultradata.xlsx ║")
    print("╚══════════════════════════════════════════════════════╝")
    print(f"  Batch : {START_INDEX} -> {END_INDEX} ({END_INDEX-START_INDEX} prompts)")
    print(f"  REPO2 : {GITHUB_REPO2}")
    print(f"  FLUX  : {FLUX_HF_ID}")
    print(f"  RMBG  : {RMBG_HF_ID} (2048x2048 — highest quality)")
    print(f"  Cache : {HF_CACHE}")
    if torch.cuda.is_available():
        p = torch.cuda.get_device_properties(0)
        print(f"  GPU   : {p.name} | VRAM: {p.total_memory/1e9:.0f}GB")
    print()

    try:
        # Manual drop processing (independent of Flux generation)
        try:
            process_manual_drop()
        except Exception as e:
            log(f"Manual drop error (non-fatal): {e}")

        prompts = load_prompts()
        if not prompts:
            raise Exception("No prompts loaded!")

        skip_set = load_skip_set_from_json()
        # Also skip images already in ultradata.xlsx (generated but SEO not yet done)
        skip_set |= load_skip_set_from_ultradata()
        log(f"  Combined skip-set: {len(skip_set)} filenames total\n")

        # ── AUTO-ADVANCE: if batch is fully skipped, scan forward to find real work ──
        BATCH_SIZE   = END_INDEX - START_INDEX
        real_start   = START_INDEX
        real_end     = END_INDEX
        MAX_SCAN     = min(len(prompts), START_INDEX + max(BATCH_SIZE * 50, 5000))

        scan_start = START_INDEX
        while scan_start < MAX_SCAN:
            window = prompts[scan_start:scan_start + BATCH_SIZE]
            fresh  = [p for p in window if p["filename"] not in skip_set]
            if fresh:
                if scan_start != START_INDEX:
                    log(f"  AUTO-ADVANCE: {START_INDEX} → {scan_start} "
                        f"(skipped {scan_start - START_INDEX} already-done items)")
                real_start = scan_start
                real_end   = scan_start + BATCH_SIZE
                break
            log(f"  All skipped in {scan_start}→{scan_start+BATCH_SIZE}, scanning ahead...")
            scan_start += BATCH_SIZE
        else:
            log("  All remaining prompts already done — nothing to generate.")
            return

        batch = prompts[real_start:real_end]
        log(f"Batch: {real_start} → {real_end} ({len(batch)} prompts, "
            f"{len([p for p in batch if p['filename'] not in skip_set])} fresh)\n")

        # ── Update START/END so batch_tracker.txt saves the correct position ──
        import builtins as _bt
        _bt.__dict__["_EFFECTIVE_START"] = real_start
        _bt.__dict__["_EFFECTIVE_END"]   = real_end
        os.environ["START_INDEX"] = str(real_start)
        os.environ["END_INDEX"]   = str(real_end)

        # PHASE 1
        generated = phase1_generate(batch, skip_set)
        stats["generated"] = len(generated)
        if not generated:
            log("No new images generated."); return

        # PHASE 2 (REMOVED FROM WORKFLOW):
        # Filter + SEO must not run in Section 1.
        # We only convert generated items into minimal "posts" needed for RMBG + upload.
        log("=" * 56)
        log("PHASE 2: (SKIPPED) Filter + SEO removed from Section 1 workflow")
        log("=" * 56)
        posts = []
        for gd in generated:
            item = gd["item"]
            src  = Path(gd["path"])
            rel  = src.relative_to(GENERATED_DIR)
            dst  = APPROVED_DIR / rel
            dst.parent.mkdir(parents=True, exist_ok=True)
            if not dst.exists():
                shutil.copy2(str(src), str(dst))
            subject = (item.get("subject_name") or item.get("subject") or
                       item.get("name") or item.get("title") or
                       item.get("subcategory") or item.get("category") or "Untitled")
            subject = str(subject).replace("_", " ").replace("-", " ").title()
            posts.append({
                "category": item.get("category", "general"),
                "subcategory": item.get("subcategory", "general"),
                "subject_name": subject,
                "filename": item.get("filename", src.name),
                "original_prompt": item.get("prompt", ""),
                "approved_path": str(dst),
                # Keep downstream schema compatibility
                "slug": slugify(subject),
                "title": "",
                "h1": "",
                "meta_desc": "",
                "alt_text": "",
                "tags": "",
                "description": "",
                "word_count": 0,
                "ai_generated": False,
                "qwen_confidence": 0,
                "qwen_group": "",
                "png_file_id": "", "jpg_file_id": "", "webp_file_id": "",
                "download_url": "", "preview_url": "", "preview_url_small": "",
                "webp_preview_url": "", "preview_w": 800, "preview_h": 800,
                "date_added": datetime.now().strftime("%Y-%m-%d"),
            })
        stats["approved"] = len(posts)
        stats["deleted"]  = 0
        if not posts:
            log("No items to process after Phase 1."); return

        # PHASE 3
        transparent = phase3_bg_remove(posts)
        stats["transparent"] = len(transparent)
        if not transparent:
            log("BG removal produced no results."); return

        # PHASE 4
        uploaded = phase4_upload(transparent)
        stats["uploaded"] = len(uploaded)
        if not uploaded:
            log("Drive upload failed."); return

        # PHASE 5 (REMOVED FROM WORKFLOW):
        # Section 1 must NOT push SEO/JSON into Repo 2 anymore.
        stats["posts"] = len(uploaded)

        # ultradata.xlsx — MUST succeed. If push fails, job fails → batch_tracker
        # does NOT advance → no Drive duplicates on next run.
        _append_ultradata_and_push(uploaded)

        # Clear checkpoints only AFTER xlsx is safely pushed
        for ck in CHECKPOINT_DIR.glob("*.json"):
            ck.unlink()

        # ── Write effective_end.txt so GitHub Actions saves the correct batch position ──
        eff_end = int(os.environ.get("END_INDEX", str(END_INDEX)))
        eff_end_path = Path("/kaggle/working/effective_end.txt")
        eff_end_path.write_text(str(eff_end))
        log(f"  effective_end.txt written: {eff_end} (batch_tracker will be set to this)")

        hrs = (time.time() - t0) / 3600
        stats["duration"] = f"{hrs:.1f}h"
        stats["status"]   = "SUCCESS"

        print(f"\n╔══════════════════════════════════════════════════════╗")
        print(f"║  DONE in {hrs:.1f}h")
        print(f"║  Gen:{len(generated)} Del:{stats['deleted']} OK:{len(posts)}")
        print(f"║  Trans:{len(transparent)} Up:{len(uploaded)}")
        print(f"╚══════════════════════════════════════════════════════╝")

    except Exception as e:
        hrs = (time.time() - t0) / 3600
        stats["duration"] = f"{hrs:.1f}h"
        stats["status"]   = f"FAILED: {str(e)[:80]}"
        log(f"FATAL: {e}")
        raise
    finally:
        phase6_save_logs(stats)

def _append_ultradata_and_push(items):
    """
    Append-only writer for ultradata.xlsx stored in REPO1 (this GitHub repo).
    This runs inside Kaggle; it clones/pulls REPO1 using GITHUB_TOKEN_REPO1.
    """
    token = GITHUB_TOKEN_REPO1 or ""
    repo  = (GITHUB_REPO1 or "").strip()
    if not token or not repo:
        raise Exception("Missing GITHUB_TOKEN_REPO1 or GITHUB_REPO (repo1)")

    repo_url = f"https://x-access-token:{token}@github.com/{repo}.git"
    xrepo_dir = WORKING_DIR / "repo1_xlsx"

    if xrepo_dir.exists() and (xrepo_dir / ".git").exists():
        # BUG FIX: set-url FIRST so pull uses the authenticated token URL
        subprocess.run(["git", "remote", "set-url", "origin", repo_url],
                       cwd=str(xrepo_dir), capture_output=True)
        subprocess.run(["git", "pull", "--rebase", "--autostash"],
                       cwd=str(xrepo_dir), capture_output=True)
    else:
        shutil.rmtree(str(xrepo_dir), ignore_errors=True)
        subprocess.run(["git", "clone", "--depth", "1", repo_url, str(xrepo_dir)],
                       capture_output=True, check=True)

    try:
        import openpyxl
        from openpyxl import Workbook
    except Exception:
        subprocess.run([sys.executable, "-m", "pip", "install", "-q", "openpyxl>=3.1.2"],
                       capture_output=True, check=True)
        import openpyxl
        from openpyxl import Workbook

    xlsx_path = xrepo_dir / ULTRADATA_XLSX_NAME
    headers = [
        "date_added", "subject_name", "category", "subcategory", "filename",
        "png_file_id", "webp_file_id", "download_url", "preview_url"
    ]

    if xlsx_path.exists():
        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb.active
        existing_header = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        if existing_header != headers:
            # If header differs, keep existing sheet and append missing columns at end.
            # We do NOT overwrite old data.
            for h in headers[len(existing_header):]:
                ws.cell(row=1, column=len(existing_header) + 1, value=h)
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)

    def _row(it):
        return [
            it.get("date_added", ""),
            it.get("subject_name", ""),
            it.get("category", ""),
            it.get("subcategory", ""),
            it.get("filename", ""),
            it.get("png_file_id", ""),
            it.get("webp_file_id", ""),
            it.get("download_url", ""),
            it.get("preview_url", ""),
        ]

    appended = 0
    for it in items:
        if not it.get("subject_name") or not it.get("download_url") or not it.get("preview_url"):
            continue
        ws.append(_row(it))
        appended += 1

    wb.save(str(xlsx_path))

    if appended == 0:
        log("ultradata.xlsx: nothing new to append")
        return

    orig_dir = os.getcwd()
    try:
        os.chdir(str(xrepo_dir))
        subprocess.run(["git", "config", "user.name", "github-actions[bot]"],
                       check=True, capture_output=True)
        subprocess.run(["git", "config", "user.email",
                        "github-actions[bot]@users.noreply.github.com"],
                       check=True, capture_output=True)
        subprocess.run(["git", "add", ULTRADATA_XLSX_NAME], check=True, capture_output=True)
        diff = subprocess.run(["git", "diff", "--staged", "--quiet"], capture_output=True)
        if diff.returncode != 0:
            msg = f"ultradata: append {appended} rows [{datetime.now().strftime('%Y-%m-%d')}]"
            subprocess.run(["git", "commit", "-m", msg], check=True, capture_output=True)

            # ── Push with rebase-retry (handles concurrent push conflicts) ──
            pushed = False
            for push_attempt in range(1, 5):
                result = subprocess.run(["git", "push"], capture_output=True, text=True)
                if result.returncode == 0:
                    pushed = True
                    break
                log(f"  Push attempt {push_attempt}/4 failed — pulling rebase...")
                subprocess.run(
                    ["git", "pull", "--rebase", "--autostash"],
                    capture_output=True, cwd=str(xrepo_dir)
                )
                time.sleep(5 * push_attempt)

            if not pushed:
                raise Exception(
                    f"git push failed after 4 attempts:\n{result.stderr.strip()}"
                )
            log(f"ultradata.xlsx pushed: +{appended} rows")
        else:
            log("ultradata.xlsx: no staged changes")
    finally:
        os.chdir(orig_dir)


if __name__ == "__main__":
    main()
