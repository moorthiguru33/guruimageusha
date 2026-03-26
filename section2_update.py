import os
import re
import io
import json
import math
import time
import shutil
import base64
import urllib.parse
import subprocess
from pathlib import Path
from datetime import datetime

import requests


SITE_URL = "https://www.ultrapng.com"
SITE_NAME = "UltraPNG"
WATERMARK_TEXT = "www.ultrapng.com"


def log(msg: str):
    line = f"[{datetime.now().strftime('%H:%M:%S')}] {msg}"
    print(line, flush=True)


def slugify(s: str, max_len: int = 60) -> str:
    s = re.sub(r"[^a-z0-9]+", "-", str(s or "").lower()).strip("-")
    if len(s) > max_len:
        cut = s[:max_len]
        idx = cut.rfind("-")
        s = cut[:idx] if idx > max_len // 2 else cut
    return s or "untitled"


def preview_url(fid: str, size: int = 800) -> str:
    return f"https://lh3.googleusercontent.com/d/{fid}=s{size}"


def download_url(fid: str) -> str:
    return f"https://drive.usercontent.google.com/download?id={fid}&export=download&authuser=0"


def get_env(name: str, default: str = "") -> str:
    return os.environ.get(name, default)


# ----------------------------
# Google Drive helpers
# ----------------------------
_token_cache = {"value": None, "expires": 0}


def get_drive_token() -> str:
    client_id = get_env("GOOGLE_CLIENT_ID")
    client_secret = get_env("GOOGLE_CLIENT_SECRET")
    refresh_token = get_env("GOOGLE_REFRESH_TOKEN")
    if _token_cache["value"] and time.time() < _token_cache["expires"]:
        return _token_cache["value"]
    r = requests.post(
        "https://oauth2.googleapis.com/token",
        data={
            "client_id": client_id,
            "client_secret": client_secret,
            "refresh_token": refresh_token,
            "grant_type": "refresh_token",
        },
        timeout=30,
    )
    d = r.json()
    if "access_token" not in d:
        raise RuntimeError(f"Token error: {d}")
    _token_cache.update({"value": d["access_token"], "expires": time.time() + 3200})
    return _token_cache["value"]


def drive_list_children(token: str, parent_id: str, mime_type: str | None = None) -> list[dict]:
    h = {"Authorization": f"Bearer {token}"}
    q = f"'{parent_id}' in parents and trashed=false"
    if mime_type:
        q += f" and mimeType='{mime_type}'"
    r = requests.get(
        "https://www.googleapis.com/drive/v3/files",
        headers=h,
        params={"q": q, "fields": "files(id,name,mimeType)", "pageSize": 200},
        timeout=30,
    )
    return r.json().get("files", [])


def drive_folder_id(token: str, name: str, parent_id: str | None = None, create: bool = True) -> str:
    h = {"Authorization": f"Bearer {token}"}
    q = f"name='{name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    if parent_id:
        q += f" and '{parent_id}' in parents"
    r = requests.get(
        "https://www.googleapis.com/drive/v3/files",
        headers=h,
        params={"q": q, "fields": "files(id)", "pageSize": 10},
        timeout=30,
    )
    files = r.json().get("files", [])
    if files:
        return files[0]["id"]
    if not create:
        raise RuntimeError(f"Drive folder not found: {name}")
    meta = {"name": name, "mimeType": "application/vnd.google-apps.folder"}
    if parent_id:
        meta["parents"] = [parent_id]
    r2 = requests.post(
        "https://www.googleapis.com/drive/v3/files",
        headers={**h, "Content-Type": "application/json"},
        json=meta,
        timeout=30,
    )
    return r2.json()["id"]


def drive_download(token: str, file_id: str) -> bytes:
    h = {"Authorization": f"Bearer {token}"}
    r = requests.get(
        f"https://www.googleapis.com/drive/v3/files/{file_id}",
        headers=h,
        params={"alt": "media"},
        timeout=120,
    )
    if not r.ok:
        raise RuntimeError(f"Drive download failed: {r.status_code} {r.text[:200]}")
    return r.content


def drive_move_file(token: str, file_id: str, new_parent_id: str) -> None:
    # Move by updating parents.
    h = {"Authorization": f"Bearer {token}"}
    r = requests.patch(
        f"https://www.googleapis.com/drive/v3/files/{file_id}",
        headers={**h, "Content-Type": "application/json"},
        params={"addParents": new_parent_id, "removeParents": "*"},
        timeout=30,
    )
    if not r.ok:
        # Fallback: attempt without wildcard removal.
        requests.patch(
            f"https://www.googleapis.com/drive/v3/files/{file_id}",
            headers={**h, "Content-Type": "application/json"},
            params={"addParents": new_parent_id},
            timeout=30,
        )


def drive_upload_bytes(token: str, folder_id: str, name: str, data: bytes, mime: str, retries: int = 3) -> dict:
    # Keep the same low-call multipart upload style as Section 1.
    for attempt in range(1, retries + 1):
        try:
            h = {"Authorization": f"Bearer {token}"}
            metadata = json.dumps({"name": name, "parents": [folder_id]})
            boundary = "----UltraPNGPipe"
            body = (
                f"--{boundary}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n{metadata}\r\n"
                f"--{boundary}\r\nContent-Type: {mime}\r\n\r\n"
            ).encode() + data + f"\r\n--{boundary}--".encode()
            r = requests.post(
                "https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart&fields=id,name",
                headers={**h, "Content-Type": f'multipart/related; boundary="{boundary}"'},
                data=body,
                timeout=120,
            )
            if r.ok:
                return r.json()
            raise RuntimeError(f"HTTP {r.status_code}: {r.text[:200]}")
        except Exception as e:
            if attempt < retries:
                log(f"  Upload retry {attempt}/{retries}: {e}")
                time.sleep(5 * attempt)
            else:
                raise


def drive_share_public(token: str, file_id: str) -> None:
    try:
        requests.post(
            f"https://www.googleapis.com/drive/v3/files/{file_id}/permissions",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json={"role": "reader", "type": "anyone"},
            timeout=30,
        )
    except Exception:
        pass


def make_previews_webp(png_bytes: bytes) -> bytes:
    # Uses the same watermark/footer approach as Section 1.
    from PIL import Image, ImageDraw, ImageFont
    import piexif

    with Image.open(io.BytesIO(png_bytes)).convert("RGBA") as img_rgba:
        w, h = img_rgba.size
        if max(w, h) > 800:
            r = 800 / max(w, h)
            img_rgba = img_rgba.resize((int(w * r), int(h * r)), Image.LANCZOS)
        w, h = img_rgba.size
        bg = Image.new("RGB", (w, h), (255, 255, 255))
        drw = ImageDraw.Draw(bg)
        for ry in range(0, h, 20):
            for cx in range(0, w, 20):
                if (ry // 20 + cx // 20) % 2 == 1:
                    drw.rectangle([cx, ry, cx + 20, ry + 20], fill=(232, 232, 232))
        bg.paste(img_rgba.convert("RGB"), mask=img_rgba.split()[3])

        try:
            fnt = ImageFont.truetype("/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", 13)
        except Exception:
            fnt = ImageFont.load_default()

        wm_layer = Image.new("RGBA", (w, h), (0, 0, 0, 0))
        wm_draw = ImageDraw.Draw(wm_layer)
        for ry in range(-h, h + 110, 110):
            for cx in range(-w, w + 110, 110):
                wm_draw.text((cx, ry), WATERMARK_TEXT, fill=(0, 0, 0, 42), font=fnt)
        wm_rot = wm_layer.rotate(-30, expand=False)
        bg_rgba = bg.convert("RGBA")
        bg_rgba.alpha_composite(wm_rot)
        bg = bg_rgba.convert("RGB")

        drw2 = ImageDraw.Draw(bg)
        drw2.rectangle([0, h - 36, w, h], fill=(13, 13, 20))
        try:
            fnt2 = ImageFont.truetype("/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", 15)
        except Exception:
            fnt2 = fnt
        drw2.text((w // 2 - 72, h - 26), WATERMARK_TEXT, fill=(245, 166, 35), font=fnt2)

        webp_buf = io.BytesIO()
        bg.save(webp_buf, "WEBP", quality=82, method=4)
        return webp_buf.getvalue()


# ----------------------------
# Groq + free web context
# ----------------------------
def fetch_wikipedia_context(subject: str) -> str:
    subject_q = urllib.parse.quote(subject)
    url = f"https://en.wikipedia.org/api/rest_v1/page/summary/{subject_q}"
    try:
        r = requests.get(url, timeout=20, headers={"User-Agent": "Mozilla/5.0"})
        if r.ok:
            d = r.json()
            extract = d.get("extract") or ""
            if extract:
                return extract[:1500]
    except Exception:
        pass
    # Fallback search snippets
    url2 = "https://en.wikipedia.org/w/api.php"
    try:
        r = requests.get(
            url2,
            params={
                "action": "query",
                "list": "search",
                "srsearch": subject,
                "format": "json",
                "srlimit": 3,
            },
            timeout=20,
        )
        if r.ok:
            d = r.json()
            hits = d.get("query", {}).get("search", [])
            snippets = []
            for h in hits[:3]:
                snippets.append(str(h.get("snippet", "")))
            return "\n".join(snippets)[:1500]
    except Exception:
        pass
    return ""


def groq_generate_seo(subject_name: str, category: str = "", subcategory: str = "") -> dict:
    groq_key = get_env("GROQ_API_KEY") or get_env("GROQ_KEY")
    if not groq_key:
        raise RuntimeError("Missing GROQ_API_KEY/GROQ_KEY env var")

    model = get_env("GROQ_MODEL", "llama-3.1-70b-versatile")
    context = fetch_wikipedia_context(subject_name)

    prompt = (
        "You are an SEO writer for a free transparent PNG library.\n"
        "Create ONLY a JSON object (no markdown) with keys:\n"
        "title (20+ words), h1, meta_desc (<=155 chars), alt_text, tags (comma-separated), description (300+ words).\n\n"
        f"SUBJECT: {subject_name}\n"
        f"CATEGORY: {category}\n"
        f"SUBCATEGORY: {subcategory}\n\n"
        "FREE WEB CONTEXT (Wikipedia extract/snippets):\n"
        f"{context}\n\n"
        "Constraints:\n"
        "- Title must be 20+ words and must include the subject name.\n"
        "- Description must be 300+ words, unique, and should discuss how this specific PNG can be used (design ideas) and technical quality (transparent background).\n"
        "- No keyword stuffing. Natural language.\n"
    )

    url = "https://api.groq.com/openai/v1/chat/completions"
    r = requests.post(
        url,
        headers={"Authorization": f"Bearer {groq_key}", "Content-Type": "application/json"},
        json={
            "model": model,
            "temperature": 0.4,
            "messages": [{"role": "user", "content": prompt}],
            "response_format": {"type": "json_object"},
        },
        timeout=120,
    )
    if not r.ok:
        raise RuntimeError(f"Groq error: {r.status_code} {r.text[:300]}")
    data = r.json()
    content = data["choices"][0]["message"]["content"]
    try:
        return json.loads(content)
    except Exception:
        # Last-resort parse: find first {...}
        j0 = content.find("{")
        j1 = content.rfind("}") + 1
        if j0 >= 0 and j1 > j0:
            return json.loads(content[j0:j1])
        raise


# ----------------------------
# Repo2 JSON sharding
# ----------------------------
def load_repo2_entries(repo2_data_dir: Path) -> list[dict]:
    all_entries = []
    if not repo2_data_dir.exists():
        return all_entries
    for jf in repo2_data_dir.glob("*.json"):
        if jf.name.startswith("_"):
            continue
        try:
            entries = json.loads(jf.read_text("utf-8"))
            if isinstance(entries, list):
                all_entries.extend(entries)
        except Exception:
            pass
    return all_entries


def write_sharded_entries(
    repo2_data_dir: Path,
    entries: list[dict],
    shard_size: int = 200,
    shard_prefix: str = "json",
) -> None:
    repo2_data_dir.mkdir(parents=True, exist_ok=True)
    # Remove existing json shards (keep _index.json or other underscore files)
    for jf in repo2_data_dir.glob("*.json"):
        if jf.name.startswith("_"):
            continue
        if re.match(rf"^{re.escape(shard_prefix)}(\d+)?\.json$", jf.name):
            jf.unlink(missing_ok=True)

    shards = []
    for i in range(0, len(entries), shard_size):
        shards.append(entries[i : i + shard_size])

    for idx, shard in enumerate(shards, 1):
        name = f"{shard_prefix}{idx}.json"
        (repo2_data_dir / name).write_text(json.dumps(shard, ensure_ascii=False, indent=2), "utf-8")


# ----------------------------
# ultradata.xlsx
# ----------------------------
def read_ultradata_xlsx(xlsx_path: Path) -> tuple[list[dict], list[str]]:
    import openpyxl

    wb = openpyxl.load_workbook(str(xlsx_path))
    ws = wb["ultradata"] if "ultradata" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [str(x or "").strip() for x in rows[0]]
    out = []
    for r in rows[1:]:
        if not any(r):
            continue
        d = {}
        for i, h in enumerate(headers):
            d[h] = r[i] if i < len(r) else ""
        out.append(d)
    return out, headers


def append_ultradata_rows(xlsx_path: Path, rows_to_append: list[dict]) -> int:
    import openpyxl

    if xlsx_path.exists():
        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb["ultradata"] if "ultradata" in wb.sheetnames else wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ultradata"

    headers = [c.value for c in ws[1]] if ws.max_row >= 1 else []
    if not headers or any(h is None for h in headers):
        headers = [
            "subject_name",
            "category",
            "subcategory",
            "filename",
            "slug",
            "download_url",
            "webp_download_url",
            "preview_url",
            "preview_url_small",
            "webp_preview_url",
            "png_file_id",
            "webp_file_id",
            "date_added",
        ]
        ws.delete_rows(1, ws.max_row)
        ws.append(headers)
    else:
        headers = [str(h) for h in headers]

    col = {h: i for i, h in enumerate(headers)}
    added = 0
    for d in rows_to_append:
        ws.append([d.get(h, "") for h in headers])
        added += 1

    tmp_path = str(xlsx_path) + ".tmp"
    wb.save(tmp_path)
    os.replace(tmp_path, str(xlsx_path))
    return added


def drive_file_exists(token: str, file_id: str) -> bool:
    h = {"Authorization": f"Bearer {token}"}
    try:
        r = requests.get(
            f"https://www.googleapis.com/drive/v3/files/{file_id}",
            headers=h,
            params={"fields": "id"},
            timeout=30,
        )
        return bool(r.ok)
    except Exception:
        return False


def parse_subject_from_filename(filename: str) -> str:
    # filename examples: "chicken_biryani_with_white_plate.png" or "img_000123.png"
    stem = Path(filename).stem
    stem = re.sub(r"^img[_-]?\d+$", "", stem, flags=re.I).strip("_- ").strip()
    stem = stem.replace("_", " ").replace("-", " ")
    stem = re.sub(r"\s+", " ", stem).strip()
    if not stem:
        return "Untitled PNG"
    return " ".join(w[:1].upper() + w[1:].lower() for w in stem.split(" "))


def main():
    repo_root = Path(".").resolve()
    ultradata_path = repo_root / "ultradata.xlsx"
    if not ultradata_path.exists():
        log("ultradata.xlsx not found. Nothing to do.")
        return

    google_token = get_drive_token()

    # Drive folder names (can be changed via env if needed)
    manual_drop_name = get_env("DRIVE_MANUAL_DROP_FOLDER", "Manual_Drop")
    png_root_name = get_env("DRIVE_PNG_LIBRARY_IMAGES_FOLDER", "png_library_images")
    preview_root_name = get_env("DRIVE_PNG_LIBRARY_PREVIEWS_FOLDER", "png_library_previews")
    manual_category = get_env("DRIVE_MANUAL_CATEGORY", "manual")
    manual_subcategory = get_env("DRIVE_MANUAL_SUBCATEGORY", "manual")

    manual_drop_id = drive_folder_id(google_token, manual_drop_name, parent_id=None, create=True)

    # Locate (or create) library folders
    png_root_id = drive_folder_id(google_token, png_root_name, parent_id=None, create=True)
    preview_root_id = drive_folder_id(google_token, preview_root_name, parent_id=None, create=True)

    png_lib_cat_id = drive_folder_id(google_token, manual_category, parent_id=png_root_id, create=True)
    png_lib_sub_id = drive_folder_id(google_token, manual_subcategory, parent_id=png_lib_cat_id, create=True)
    prev_lib_cat_id = drive_folder_id(google_token, manual_category, parent_id=preview_root_id, create=True)
    prev_lib_sub_id = drive_folder_id(google_token, manual_subcategory, parent_id=prev_lib_cat_id, create=True)

    ultradata_rows, headers = read_ultradata_xlsx(ultradata_path)
    if not ultradata_rows:
        log("ultradata.xlsx has no rows.")
        return

    # Build ultradata index by png_file_id (NOT filename) so duplicate-named
    # images in manual drop are all processed — only true same-file duplicates skip.
    uda_by_file_id = {}
    uda_by_filename = {}
    for r in ultradata_rows:
        fn = str(r.get("filename", "") or "").strip()
        fid = str(r.get("png_file_id", "") or "").strip()
        if fn:
            uda_by_filename[fn] = r
        if fid:
            uda_by_file_id[fid] = r

    # --- Clone REPO2 data/ only (best-effort) ---
    repo2_slug = get_env("GITHUB_REPO2") or get_env("REPO_2")
    repo2_token = get_env("REPO_2_TOKEN") or get_env("GITHUB_TOKEN_REPO2") or get_env("GITHUB_TOKEN_REPO2_VAL")
    # In GitHub Actions secrets, the workflow typically passes REPO_2 and REPO_2_TOKEN into these.
    repo2_token = repo2_token or get_env("GITHUB_TOKEN_REPO2_VAL")
    repo2_slug = repo2_slug or get_env("GITHUB_REPO2_VAL") or get_env("REPO_2")
    if not repo2_slug or not repo2_token:
        log("REPO2 token/repo not set — cannot update JSON.")
        return

    work_dir = repo_root / ".section2_work"
    shutil.rmtree(str(work_dir), ignore_errors=True)
    work_dir.mkdir(parents=True, exist_ok=True)
    repo2_dir = work_dir / "repo2"

    git_url = f"https://x-access-token:{repo2_token}@github.com/{repo2_slug}.git"
    # Full clone is simplest (still typically small with sparse disabled).
    subprocess.run(["git", "clone", "--depth", "1", git_url, str(repo2_dir)], check=True, capture_output=True)
    repo2_data_dir = repo2_dir / "data"

    existing_entries = load_repo2_entries(repo2_data_dir)
    existing_by_filename = {}
    for e in existing_entries:
        fn = str(e.get("filename", "") or "").strip()
        if fn:
            existing_by_filename[fn] = e

    new_entries = []
    to_process = [r for r in ultradata_rows if str(r.get("filename", "") or "").strip() and str(r.get("filename", "") or "").strip() not in existing_by_filename]

    log(f"Section2: ultradata rows={len(ultradata_rows)} existing repo2={len(existing_entries)} missing={len(to_process)}")

    # --- 1) Process missing ultradata rows into repo2 JSON ---
    for idx, r in enumerate(to_process, 1):
        fn = str(r.get("filename", "") or "").strip()
        if not fn:
            continue

        png_file_id = str(r.get("png_file_id", "") or "").strip()
        if not png_file_id:
            log(f"  Skip (no PNG file id in ultradata): {fn}")
            continue
        if not drive_file_exists(google_token, png_file_id):
            log(f"  Skip (missing PNG in library): {fn}")
            continue

        subject = str(r.get("subject_name", "") or "").strip() or parse_subject_from_filename(fn)
        category = str(r.get("category", "") or "").strip()
        subcategory = str(r.get("subcategory", "") or "").strip()

        log(f"  [{idx}/{len(to_process)}] SEO via Groq: {subject}")
        seo = groq_generate_seo(subject, category, subcategory)

        description = seo.get("description", "") or ""
        word_count = len(description.split()) if description else 0

        entry = {
            "category": category,
            "subcategory": subcategory,
            "subject_name": subject,
            "filename": fn,
            "slug": str(r.get("slug", "") or slugify(subject)),
            "download_url": str(r.get("download_url", "") or ""),
            "preview_url": str(r.get("preview_url", "") or ""),
            "preview_url_small": str(r.get("preview_url_small", "") or ""),
            "webp_preview_url": str(r.get("webp_preview_url", "") or ""),
            "png_file_id": str(r.get("png_file_id", "") or ""),
            "webp_file_id": str(r.get("webp_file_id", "") or ""),
            "webp_download_url": str(r.get("webp_download_url", "") or ""),
            "title": seo.get("title", ""),
            "h1": seo.get("h1", seo.get("title", "")),
            "meta_desc": seo.get("meta_desc", ""),
            "alt_text": seo.get("alt_text", seo.get("title", "")),
            "tags": seo.get("tags", ""),
            "description": description,
            "word_count": word_count,
            "ai_generated": True,
            "date_added": str(r.get("date_added", "") or datetime.now().strftime("%Y-%m-%d")),
        }
        new_entries.append(entry)

        time.sleep(0.2)

    # --- 2) Manual drop processing (move png to library + upload webp + append ultradata + create repo2 entry) ---
    manual_files = drive_list_children(google_token, manual_drop_id, mime_type="image/png")
    if manual_files:
        log(f"Manual drop: found {len(manual_files)} png files")
    rows_append = []
    for f in manual_files:
        name = f.get("name", "")
        file_id = f.get("id", "")
        if not name or not file_id:
            continue
        # Skip only if this exact Drive file_id was already processed (true duplicate).
        # Same filename with different file_id = different image → always process.
        if file_id in uda_by_file_id:
            log(f"  Skip (already processed file_id): {name}")
            continue

        # Move to png library manual folder (so they can be deleted without breaking)
        try:
            drive_move_file(google_token, file_id, png_lib_sub_id)
        except Exception as e:
            log(f"  Move failed for {name}: {e}")

        png_bytes = drive_download(google_token, file_id)
        webp_bytes = make_previews_webp(png_bytes)

        webp_name = Path(name).stem + ".webp"
        webp_upload = drive_upload_bytes(
            google_token, prev_lib_sub_id, webp_name, webp_bytes, "image/webp"
        )
        webp_id = webp_upload["id"]
        drive_share_public(google_token, webp_id)

        subject = parse_subject_from_filename(name)
        category = manual_category
        subcategory = manual_subcategory
        # Unique slug: if same filename already used, append short file_id suffix
        base_slug = slugify(subject)
        slug = base_slug if base_slug not in {r.get("slug","") for r in uda_by_filename.values()} \
               else f"{base_slug}-{file_id[:6]}"

        png_dl = download_url(file_id)
        webp_dl = download_url(webp_id)
        prev_webp = preview_url(webp_id, 800)
        prev_webp_small = preview_url(webp_id, 400)

        row = {
            "subject_name": subject,
            "category": category,
            "subcategory": subcategory,
            "filename": name,
            "slug": slug,
            "download_url": png_dl,
            "webp_download_url": webp_dl,
            "preview_url": prev_webp,
            "preview_url_small": prev_webp_small,
            "webp_preview_url": prev_webp,
            "png_file_id": file_id,
            "webp_file_id": webp_id,
            "date_added": datetime.now().strftime("%Y-%m-%d"),
        }
        rows_append.append(row)
        uda_by_filename[name] = row

    if rows_append:
        added = append_ultradata_rows(ultradata_path, rows_append)
        log(f"  Appended {added} manual rows to ultradata.xlsx")

        # Commit ultradata.xlsx
        try:
            subprocess.run(["git", "add", "ultradata.xlsx"], cwd=str(repo_root), check=True)
            diff = subprocess.run(["git", "diff", "--cached", "--quiet"], cwd=str(repo_root), capture_output=True)
            # diff.returncode==0 => no changes
            if diff.returncode != 0:
                msg = f"Update ultradata.xlsx (+{added}) [{datetime.now().strftime('%Y-%m-%d')}]"
                subprocess.run(["git", "config", "user.name", "github-actions[bot]"], cwd=str(repo_root), check=True)
                subprocess.run(
                    ["git", "config", "user.email", "github-actions[bot]@users.noreply.github.com"],
                    cwd=str(repo_root),
                    check=True,
                )
                subprocess.run(["git", "commit", "-m", msg], cwd=str(repo_root), check=True)
                subprocess.run(["git", "push"], cwd=str(repo_root), check=True)
        except Exception as e:
            log(f"  ultradata.xlsx commit/push failed: {e}")

        # Generate SEO for manual rows and create repo2 entries right away.
        for j, row in enumerate(rows_append, 1):
            subject = row["subject_name"]
            log(f"  Groq SEO for manual [{j}/{len(rows_append)}]: {subject}")
            seo = groq_generate_seo(subject, row.get("category", ""), row.get("subcategory", ""))
            description = seo.get("description", "") or ""
            word_count = len(description.split()) if description else 0
            new_entries.append(
                {
                    "category": row.get("category", ""),
                    "subcategory": row.get("subcategory", ""),
                    "subject_name": subject,
                    "filename": row.get("filename", ""),
                    "slug": row.get("slug", ""),
                    "download_url": row.get("download_url", ""),
                    "preview_url": row.get("preview_url", ""),
                    "preview_url_small": row.get("preview_url_small", ""),
                    "webp_preview_url": row.get("webp_preview_url", ""),
                    "png_file_id": row.get("png_file_id", ""),
                    "webp_file_id": row.get("webp_file_id", ""),
                    "webp_download_url": row.get("webp_download_url", ""),
                    "title": seo.get("title", ""),
                    "h1": seo.get("h1", seo.get("title", "")),
                    "meta_desc": seo.get("meta_desc", ""),
                    "alt_text": seo.get("alt_text", seo.get("title", "")),
                    "tags": seo.get("tags", ""),
                    "description": description,
                    "word_count": word_count,
                    "ai_generated": True,
                    "date_added": row.get("date_added", datetime.now().strftime("%Y-%m-%d")),
                }
            )

            time.sleep(0.2)

    # --- 3) Write merged JSON shards into REPO2 and push ---
    if new_entries:
        # Merge + dedupe by filename
        merged_by_fn = {str(e.get("filename", "") or "").strip(): e for e in existing_entries}
        for e in new_entries:
            fn = str(e.get("filename", "") or "").strip()
            if fn:
                merged_by_fn[fn] = e
        merged_entries = list(merged_by_fn.values())
        log(f"Repo2: writing merged entries total={len(merged_entries)} (+{len(new_entries)} new/updated)")

        # Write shards (200 per file)
        shard_prefix = get_env("REPO2_JSON_SHARD_PREFIX", "json")
        write_sharded_entries(
            repo2_data_dir,
            merged_entries,
            shard_size=200,
            shard_prefix=shard_prefix,
        )

        # Push
        subprocess.run(["git", "add", "data/"], cwd=str(repo2_dir), check=True)
        diff = subprocess.run(["git", "diff", "--cached", "--quiet"], cwd=str(repo2_dir), capture_output=True)
        if diff.returncode != 0:
            msg = f"Update REPO2 data shards [+{len(new_entries)}] [{datetime.now().strftime('%Y-%m-%d')}]"
            subprocess.run(["git", "commit", "-m", msg], cwd=str(repo2_dir), check=True, capture_output=True)
            subprocess.run(["git", "push"], cwd=str(repo2_dir), check=True)
            log("Repo2 JSON pushed.")
        else:
            log("Repo2: nothing to commit.")
    else:
        log("Repo2: no new entries generated.")


if __name__ == "__main__":
    main()
