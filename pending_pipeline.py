"""
╔══════════════════════════════════════════════════════════════════╗
║   Pending Drive Pipeline  V3.4                                   ║
╠══════════════════════════════════════════════════════════════════╣
║  PHASE 1 → Discovery     BFS recursive — ALL nested subfolders   ║
║  PHASE 2 → Download      Batch download JPG/PNG from Drive       ║
║  PHASE 3 → BRIA RMBG-2.0 SOTA bg removal (HF / direct PyTorch) ║
║  PHASE 4 → WebP Preview  <80 KB + checkered BG + watermark      ║
║  PHASE 5 → Drive Upload  Original transparent PNG → Drive        ║
║  PHASE 6 → GitHub Upload WebP preview → guruimageusha/preview_webp║
║  PHASE 7 → Move          Original → pending/finished/{subfolder} ║
║  PHASE 8 → ultradata     Append rows → ultrapng/ultradata.xlsx   ║
╚══════════════════════════════════════════════════════════════════╝

inject_creds_pending.py prepends env vars before this file runs.
Required secrets  : GOOGLE_CLIENT_ID, GOOGLE_CLIENT_SECRET,
                    GOOGLE_REFRESH_TOKEN, GH_TOKEN, GH_OWNER
Tunable variables : RUN_ITEMS_COUNT,
                    WATERMARK_TEXT, PENDING_FOLDER_NAME
"""

import os, sys, json, time, gc, io, subprocess, math, base64
from pathlib import Path
from datetime import datetime

try:
    sys.stdout.reconfigure(line_buffering=True)
    sys.stderr.reconfigure(line_buffering=True)
except AttributeError:
    pass

HF_CACHE = Path("/kaggle/working/hf_cache")
HF_CACHE.mkdir(parents=True, exist_ok=True)
os.environ["HF_HOME"]               = str(HF_CACHE)
os.environ["HUGGINGFACE_HUB_CACHE"] = str(HF_CACHE)
os.environ["TRANSFORMERS_CACHE"]    = str(HF_CACHE)

# ══════════════════════════════════════════════════════════════════
# CONFIG  — all values read from env vars (set via GitHub Actions)
#           defaults are used when running locally
# ══════════════════════════════════════════════════════════════════
def _env_int(key, default):
    try:
        return int(os.environ.get(key, default))
    except ValueError:
        return default

RUN_ITEMS_COUNT     = _env_int("RUN_ITEMS_COUNT", 50)   # selectable per run
WATERMARK_TEXT      = os.environ.get("WATERMARK_TEXT",      "www.ultrapng.com")
PENDING_FOLDER_NAME = os.environ.get("PENDING_FOLDER_NAME", "pending")
REMBG_MODEL         = "briaai/RMBG-2.0"                 # SOTA — fixed, no longer configurable
BRIA_INPUT_SIZE     = 1024                               # model expects 1024×1024
BATCH_SIZE          = min(RUN_ITEMS_COUNT, 50)           # Drive upload sub-batch

WEBP_MAX_SIDE  = 800
WEBP_MAX_BYTES = 80 * 1024          # 80 KB hard limit

PREVIEW_REPO   = os.environ.get("PREVIEW_REPO",   "guruimageusha")
PREVIEW_BRANCH = os.environ.get("PREVIEW_BRANCH", "main")
PREVIEW_FOLDER = os.environ.get("PREVIEW_FOLDER", "preview_webp")

ULTRADATA_REPO   = os.environ.get("ULTRADATA_REPO",   "ultrapng")
ULTRADATA_FILE   = os.environ.get("ULTRADATA_FILE",   "ultradata.xlsx")
ULTRADATA_BRANCH = os.environ.get("ULTRADATA_BRANCH", "main")

# Secrets (injected by inject_creds_pending.py)
GOOGLE_CLIENT_ID     = os.environ.get("GOOGLE_CLIENT_ID",     "")
GOOGLE_CLIENT_SECRET = os.environ.get("GOOGLE_CLIENT_SECRET", "")
GOOGLE_REFRESH_TOKEN = os.environ.get("GOOGLE_REFRESH_TOKEN", "")
GH_TOKEN             = os.environ.get("GH_TOKEN",             "")
GH_OWNER             = os.environ.get("GH_OWNER",             "")
HF_TOKEN             = os.environ.get("HF_TOKEN",             "")   # needed for gated briaai/RMBG-2.0

# Working directories
WORKING_DIR     = Path("/kaggle/working")
DOWNLOAD_DIR    = WORKING_DIR / "downloads"
TRANSPARENT_DIR = WORKING_DIR / "transparent"
WEBP_DIR        = WORKING_DIR / "webp_previews"

for _d in [DOWNLOAD_DIR, TRANSPARENT_DIR, WEBP_DIR]:
    _d.mkdir(parents=True, exist_ok=True)

# ══════════════════════════════════════════════════════════════════
# INSTALL DEPENDENCIES
# NOTE: birefnet-general uses PyTorch+transformers (pre-installed on
#       Kaggle GPU).  onnxruntime is NOT needed and NOT installed.
# ══════════════════════════════════════════════════════════════════
print("=" * 64)
print("Installing dependencies...")
for _pkg in [
    "Pillow>=10.0",
    "numpy",
    "requests",
    "openpyxl",
    "transformers>=4.40.0",   # BRIA RMBG-2.0 — direct HuggingFace load
    "torchvision",             # image transforms for RMBG-2.0 preprocessing
]:
    _r = subprocess.run(
        [sys.executable, "-m", "pip", "install", "-q",
         "--no-warn-script-location", _pkg],
        capture_output=True, text=True,
    )
    print(f"  {'OK  ' if _r.returncode == 0 else 'WARN'} {_pkg}")
print("Done!\n")

import torch
from PIL import Image, ImageDraw, ImageFont
import requests as req

# ══════════════════════════════════════════════════════════════════
# LOGGING
# ══════════════════════════════════════════════════════════════════
_LOG_LINES = []

def log(msg):
    line = f"[{datetime.now().strftime('%H:%M:%S')}] {msg}"
    print(line, flush=True)
    _LOG_LINES.append(line)

def log_section(title):
    log("=" * 64)
    log(title)
    log("=" * 64)

# ══════════════════════════════════════════════════════════════════
# GPU UTILITIES
# ══════════════════════════════════════════════════════════════════
def free_memory():
    gc.collect()
    if torch.cuda.is_available():
        torch.cuda.empty_cache()
        torch.cuda.synchronize()
    used = torch.cuda.memory_allocated(0) / 1e9 if torch.cuda.is_available() else 0
    log(f"  GPU freed -> {used:.2f} GB remaining")

def gpu_info():
    if torch.cuda.is_available():
        name  = torch.cuda.get_device_name(0)
        total = torch.cuda.get_device_properties(0).total_memory / 1e9
        used  = torch.cuda.memory_allocated(0) / 1e9
        log(f"  GPU: {name} | {used:.1f}/{total:.1f} GB")
    else:
        log("  GPU: not available — running on CPU")

# ══════════════════════════════════════════════════════════════════
# GOOGLE DRIVE — token
# ══════════════════════════════════════════════════════════════════
_token_cache = {"value": None, "expires": 0}

def get_drive_token():
    if _token_cache["value"] and time.time() < _token_cache["expires"]:
        return _token_cache["value"]
    r = req.post("https://oauth2.googleapis.com/token", data={
        "client_id":     GOOGLE_CLIENT_ID,
        "client_secret": GOOGLE_CLIENT_SECRET,
        "refresh_token": GOOGLE_REFRESH_TOKEN,
        "grant_type":    "refresh_token",
    }, timeout=30)
    d = r.json()
    if "access_token" not in d:
        raise RuntimeError(f"Drive token error: {d}")
    _token_cache.update({"value": d["access_token"],
                         "expires": time.time() + 3200})
    return _token_cache["value"]

def _dh(token):
    return {"Authorization": f"Bearer {token}"}

# ══════════════════════════════════════════════════════════════════
# GOOGLE DRIVE — folder  (get-or-create with cache + retry)
# ══════════════════════════════════════════════════════════════════
_folder_cache = {}   # cache: "parent_id::name" -> folder_id

def drive_folder_get_or_create(token, name, parent_id=None):
    """
    - Search for folder by name (+ parent).
    - If found  → return its ID immediately  (never create a duplicate).
    - If absent → create it and return the new ID.
    - In-process cache avoids redundant API calls within the same run.
    - Retries up to 3 times on transient network errors.
    """
    cache_key = f"{parent_id or 'root'}::{name}"
    if cache_key in _folder_cache:
        return _folder_cache[cache_key]

    q = (f"name='{name}' and mimeType='application/vnd.google-apps.folder'"
         f" and trashed=false")
    if parent_id:
        q += f" and '{parent_id}' in parents"

    for attempt in range(1, 4):
        try:
            # 1 — search first
            r = req.get(
                "https://www.googleapis.com/drive/v3/files",
                headers=_dh(token),
                params={"q": q, "fields": "files(id,name)", "pageSize": 1},
                timeout=30,
            )
            r.raise_for_status()
            files = r.json().get("files", [])
            if files:
                fid = files[0]["id"]
                _folder_cache[cache_key] = fid
                return fid

            # 2 — not found: create
            meta = {"name": name,
                    "mimeType": "application/vnd.google-apps.folder"}
            if parent_id:
                meta["parents"] = [parent_id]
            r2 = req.post(
                "https://www.googleapis.com/drive/v3/files",
                headers={**_dh(token), "Content-Type": "application/json"},
                json=meta, timeout=30,
            )
            r2.raise_for_status()
            fid = r2.json()["id"]
            _folder_cache[cache_key] = fid
            log(f"  Created folder '{name}' -> {fid}")
            return fid

        except Exception as e:
            if attempt < 3:
                log(f"  folder retry {attempt}/3 [{name}]: {e}")
                time.sleep(4 * attempt)
            else:
                raise RuntimeError(
                    f"drive_folder_get_or_create failed '{name}': {e}"
                ) from e

# ══════════════════════════════════════════════════════════════════
# GOOGLE DRIVE — file helpers
# ══════════════════════════════════════════════════════════════════
def drive_list(token, folder_id, mime_filter=None):
    q = f"'{folder_id}' in parents and trashed=false"
    if mime_filter:
        q += f" and mimeType='{mime_filter}'"
    results, page_token = [], None
    while True:
        params = {"q": q, "pageSize": 1000,
                  "fields": "nextPageToken,files(id,name,mimeType,parents)"}
        if page_token:
            params["pageToken"] = page_token
        r = req.get("https://www.googleapis.com/drive/v3/files",
                    headers=_dh(token), params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
        results.extend(data.get("files", []))
        page_token = data.get("nextPageToken")
        if not page_token:
            break
    return results

def drive_list_images(token, folder_id):
    q = (f"'{folder_id}' in parents and trashed=false and "
         "(mimeType='image/jpeg' or mimeType='image/png' or "
         " name contains '.jpg' or name contains '.jpeg' or name contains '.png')")
    results, page_token = [], None
    while True:
        params = {"q": q, "pageSize": 1000,
                  "fields": "nextPageToken,files(id,name,mimeType,parents)"}
        if page_token:
            params["pageToken"] = page_token
        r = req.get("https://www.googleapis.com/drive/v3/files",
                    headers=_dh(token), params=params, timeout=30)
        r.raise_for_status()
        data = r.json()
        results.extend(data.get("files", []))
        page_token = data.get("nextPageToken")
        if not page_token:
            break
    return results

def drive_upload(token, folder_id, name, data, mime="image/png", retries=3):
    for attempt in range(1, retries + 1):
        try:
            metadata = json.dumps({"name": name, "parents": [folder_id]})
            boundary = "----PendingPipe32"
            body = (
                f"--{boundary}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n"
                f"{metadata}\r\n--{boundary}\r\nContent-Type: {mime}\r\n\r\n"
            ).encode() + data + f"\r\n--{boundary}--".encode()
            r = req.post(
                "https://www.googleapis.com/upload/drive/v3/files"
                "?uploadType=multipart&fields=id,name",
                headers={**_dh(token),
                         "Content-Type": f'multipart/related; boundary="{boundary}"'},
                data=body, timeout=180,
            )
            if r.ok:
                return r.json()
            raise RuntimeError(f"HTTP {r.status_code}: {r.text[:200]}")
        except Exception as e:
            if attempt < retries:
                log(f"    Drive upload retry {attempt}/{retries}: {e}")
                time.sleep(6 * attempt)
            else:
                raise

def drive_share(token, fid):
    try:
        req.post(
            f"https://www.googleapis.com/drive/v3/files/{fid}/permissions",
            headers={**_dh(token), "Content-Type": "application/json"},
            json={"role": "reader", "type": "anyone"}, timeout=30)
    except Exception:
        pass

def drive_move(token, fid, add_parent_id, remove_parent_id):
    r = req.patch(
        f"https://www.googleapis.com/drive/v3/files/{fid}",
        headers=_dh(token),
        params={"addParents": add_parent_id, "removeParents": remove_parent_id,
                "fields": "id,parents"},
        timeout=30)
    r.raise_for_status()
    return r.json()

def drive_download_bytes(token, fid):
    r = req.get(f"https://www.googleapis.com/drive/v3/files/{fid}",
                headers=_dh(token), params={"alt": "media"}, timeout=180)
    r.raise_for_status()
    return r.content

# ══════════════════════════════════════════════════════════════════
# GITHUB API
# ══════════════════════════════════════════════════════════════════
def _gh(token):
    return {"Authorization": f"token {token}",
            "Accept": "application/vnd.github.v3+json"}

def github_get_sha(token, owner, repo, path, branch="main"):
    """Return blob SHA of an existing file, or None if not found."""
    url = f"https://api.github.com/repos/{owner}/{repo}/contents/{path}"
    r   = req.get(url, headers=_gh(token), params={"ref": branch}, timeout=30)
    return r.json().get("sha") if r.status_code == 200 else None

def github_upload_file(token, owner, repo, path, content_bytes,
                       message, branch="main", retries=3):
    """Create or update a file (fetches SHA first so updates never fail)."""
    url  = f"https://api.github.com/repos/{owner}/{repo}/contents/{path}"
    sha  = github_get_sha(token, owner, repo, path, branch)
    body = {"message": message,
            "content": base64.b64encode(content_bytes).decode(),
            "branch":  branch}
    if sha:
        body["sha"] = sha
    for attempt in range(1, retries + 1):
        try:
            r = req.put(url,
                        headers={**_gh(token), "Content-Type": "application/json"},
                        json=body, timeout=90)
            r.raise_for_status()
            return r.json()
        except Exception as e:
            if attempt < retries:
                log(f"    GH retry {attempt}/{retries}: {e}")
                time.sleep(5 * attempt)
            else:
                raise

def jsdelivr_url(owner, repo, branch, path):
    return f"https://cdn.jsdelivr.net/gh/{owner}/{repo}@{branch}/{path}"

# ══════════════════════════════════════════════════════════════════
# WEBP PREVIEW  (checkered BG + diagonal watermark + footer bar)
# ══════════════════════════════════════════════════════════════════
def _get_font(size=13):
    for fpath in [
        "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
        "/usr/share/fonts/truetype/freefont/FreeSansBold.ttf",
    ]:
        try:
            return ImageFont.truetype(fpath, size)
        except Exception:
            continue
    return ImageFont.load_default()

def _render_webp_canvas(img_rgba):
    w, h = img_rgba.size

    # 1 — checkered background
    cs  = 20
    bg  = Image.new("RGB", (w, h), (255, 255, 255))
    drw = ImageDraw.Draw(bg)
    for ry in range(0, h, cs):
        for cx in range(0, w, cs):
            if (ry // cs + cx // cs) % 2 == 1:
                drw.rectangle([cx, ry, cx + cs, ry + cs], fill=(232, 232, 232))
    bg.paste(img_rgba.convert("RGB"), mask=img_rgba.split()[3])

    # 2 — diagonal watermark  (−30°, alpha 42)
    fnt_wm   = _get_font(12)
    wm_layer = Image.new("RGBA", (w, h), (0, 0, 0, 0))
    wm_draw  = ImageDraw.Draw(wm_layer)
    for ry in range(-h, h + 120, 120):
        for cx in range(-w, w + 120, 120):
            wm_draw.text((cx, ry), WATERMARK_TEXT, fill=(0, 0, 0, 42), font=fnt_wm)
    bg_rgba = bg.convert("RGBA")
    bg_rgba.alpha_composite(wm_layer.rotate(-30, expand=False))
    bg = bg_rgba.convert("RGB")

    # 3 — footer bar
    fnt_ft = _get_font(14)
    drw2   = ImageDraw.Draw(bg)
    drw2.rectangle([0, h - 32, w, h], fill=(13, 13, 20))
    try:
        bbox = drw2.textbbox((0, 0), WATERMARK_TEXT, font=fnt_ft)
        tx   = (w - (bbox[2] - bbox[0])) // 2
    except Exception:
        tx = w // 2 - 72
    drw2.text((tx, h - 22), WATERMARK_TEXT, fill=(245, 166, 35), font=fnt_ft)
    return bg

def make_webp_preview(img_rgba):
    """
    Produce WebP bytes under WEBP_MAX_BYTES (80 KB).
      1. Resize longest side to WEBP_MAX_SIDE (800 px)
      2. Binary-search WebP quality 5→92 for best quality under limit
      3. If quality=5 still too big, shrink dimensions by 12% per step
    Returns (webp_bytes, width, height).
    """
    img_rgba = img_rgba.convert("RGBA")
    ow, oh   = img_rgba.size

    scale  = min(1.0, WEBP_MAX_SIDE / max(ow, oh, 1))
    rw     = max(int(ow * scale), 60)
    rh     = max(int(oh * scale), 60)
    canvas = _render_webp_canvas(img_rgba.resize((rw, rh), Image.LANCZOS))

    lo, hi, best = 5, 92, None
    while lo <= hi:
        mid = (lo + hi) // 2
        buf = io.BytesIO()
        canvas.save(buf, "WEBP", quality=mid, method=6)
        if buf.tell() <= WEBP_MAX_BYTES:
            best = buf.getvalue()
            lo   = mid + 1
        else:
            hi   = mid - 1

    if best is not None:
        return best, canvas.width, canvas.height

    # Still too large — shrink dimensions
    while max(rw, rh) > 80:
        scale  *= 0.88
        rw      = max(int(ow * scale), 60)
        rh      = max(int(oh * scale), 60)
        canvas  = _render_webp_canvas(img_rgba.resize((rw, rh), Image.LANCZOS))
        buf     = io.BytesIO()
        canvas.save(buf, "WEBP", quality=5, method=6)
        if buf.tell() <= WEBP_MAX_BYTES:
            return buf.getvalue(), canvas.width, canvas.height

    buf = io.BytesIO()
    canvas.save(buf, "WEBP", quality=5, method=6)
    return buf.getvalue(), canvas.width, canvas.height

# ══════════════════════════════════════════════════════════════════
# PHASE 1: DISCOVERY
# ══════════════════════════════════════════════════════════════════
def phase1_discovery():
    """
    BFS recursive discovery — finds ALL images in every nested subfolder
    under the pending folder (excluding 'finished' at any depth).

    Each item gets:
      fid              – Drive file ID of the image
      name             – original filename
      stem             – filename without extension
      subfolder_id     – immediate parent folder ID  (used for move in phase7)
      subfolder_name   – immediate parent folder name (used as subcategory)
      top_category     – first-level subfolder name  (used as category in UltraData)
      folder_path      – full relative path e.g. "fruits/tropical"  (logging only)
    """
    log_section("PHASE 1: Discovery  [recursive — all nested subfolders]")
    token      = get_drive_token()
    pending_id = drive_folder_get_or_create(token, PENDING_FOLDER_NAME)
    log(f"  '{PENDING_FOLDER_NAME}' folder ID: {pending_id}")

    # ------------------------------------------------------------------
    # BFS queue entries: (folder_id, folder_name, top_category, path_str)
    # ------------------------------------------------------------------
    queue = []

    # Seed with first-level subfolders (skip 'finished')
    top_subs = drive_list(token, pending_id,
                          mime_filter="application/vnd.google-apps.folder")
    top_subs = [sf for sf in top_subs if sf["name"].lower() != "finished"]
    log(f"  Top-level subfolders: {len(top_subs)}")

    for sf in top_subs:
        queue.append((sf["id"], sf["name"], sf["name"], sf["name"]))

    # Also check for images sitting directly in the pending root
    root_images = drive_list_images(token, pending_id)
    if root_images:
        log(f"  Images in pending root (no subfolder): {len(root_images)}")
    _root_items = []
    for img in root_images:
        _root_items.append({
            "fid":            img["id"],
            "name":           img["name"],
            "stem":           Path(img["name"]).stem,
            "subfolder_id":   pending_id,
            "subfolder_name": "uncategorised",
            "top_category":   "uncategorised",
            "folder_path":    "",
        })

    # BFS traversal
    visited    = set()
    all_items  = list(_root_items)   # root-level images first

    while queue:
        folder_id, folder_name, top_cat, path_str = queue.pop(0)

        if folder_id in visited:
            continue
        visited.add(folder_id)

        # Images directly in this folder
        token  = get_drive_token()   # refresh before each API burst
        images = drive_list_images(token, folder_id)
        log(f"  [{path_str}]: {len(images)} image(s)")

        for img in images:
            all_items.append({
                "fid":            img["id"],
                "name":           img["name"],
                "stem":           Path(img["name"]).stem,
                "subfolder_id":   folder_id,
                "subfolder_name": folder_name,
                "top_category":   top_cat,
                "folder_path":    path_str,
            })

        # Enqueue nested subfolders (skip 'finished' at every level)
        nested = drive_list(token, folder_id,
                            mime_filter="application/vnd.google-apps.folder")
        for sf in nested:
            if sf["name"].lower() == "finished":
                continue
            if sf["id"] in visited:
                continue
            child_path = f"{path_str}/{sf['name']}"
            queue.append((sf["id"], sf["name"], top_cat, child_path))

    log(f"  Total images found   : {len(all_items)}")

    # Slice AFTER full discovery — no folders are skipped
    selected = all_items[:RUN_ITEMS_COUNT]
    log(f"  Items selected       : {len(selected)} / {RUN_ITEMS_COUNT}")
    return selected, pending_id

# ══════════════════════════════════════════════════════════════════
# PHASE 2: DOWNLOAD
# ══════════════════════════════════════════════════════════════════
def phase2_download(items):
    log_section(f"PHASE 2: Download  ({len(items)} images from Drive)")
    downloaded = []
    for i, item in enumerate(items):
        try:
            token = get_drive_token()
            log(f"  [{i+1:>3}/{len(items)}] {item['name']}")
            data  = drive_download_bytes(token, item["fid"])
            ext   = Path(item["name"]).suffix.lower()
            if ext not in (".jpg", ".jpeg", ".png"):
                ext = ".jpg"
            local = DOWNLOAD_DIR / f"{item['stem']}{ext}"
            local.write_bytes(data)
            item["local_path"] = str(local)
            downloaded.append(item)
        except Exception as e:
            log(f"  FAIL [{item['name']}]: {e}")
    log(f"  Downloaded: {len(downloaded)}/{len(items)}")
    return downloaded

# ══════════════════════════════════════════════════════════════════
# PHASE 3: BRIA RMBG-2.0 SOTA  (direct HuggingFace / pure PyTorch)
#
#  Why direct instead of rembg wrapper?
#   • Full 1024×1024 inference — rembg defaults to 512px
#   • No intermediate library overhead / version drift
#   • Official briaai/RMBG-2.0 pipeline with trust_remote_code
#   • Cleaner alpha mask — better hair, fur, fine-detail edges
# ══════════════════════════════════════════════════════════════════
def _bria_select_device():
    """
    Pick inference device.
    PyTorch 2.3+ dropped compiled CUDA kernels for sm_60 (Pascal / P100).
    Anything below compute capability 7.0 (Volta) must fall back to CPU.
    """
    if not torch.cuda.is_available():
        log("  Device: CPU (no CUDA)")
        return "cpu"
    cap = torch.cuda.get_device_capability(0)
    name = torch.cuda.get_device_name(0)
    if cap[0] < 7:
        log(f"  GPU {name} is sm_{cap[0]}{cap[1]} (Pascal/older) — "
            f"PyTorch 2.3+ has no CUDA kernels for sm_<70. Using CPU.")
        return "cpu"
    log(f"  Device: CUDA ({name}, sm_{cap[0]}{cap[1]})")
    return "cuda"


def _bria_load_model():
    """Load briaai/RMBG-2.0 from HuggingFace once and return (model, device)."""
    from transformers import AutoModelForImageSegmentation

    # briaai/RMBG-2.0 is a gated repo — authenticate before downloading
    if HF_TOKEN:
        try:
            from huggingface_hub import login as hf_login
            hf_login(token=HF_TOKEN, add_to_git_credential=False)
            log("  HuggingFace login OK ✓")
        except Exception as hf_e:
            log(f"  HuggingFace login warning: {hf_e}")
    else:
        log("  WARNING: HF_TOKEN not set — gated repo access may fail")

    device = _bria_select_device()
    log(f"  Loading BRIA RMBG-2.0 on {device.upper()} ...")
    model = AutoModelForImageSegmentation.from_pretrained(
        "briaai/RMBG-2.0",
        trust_remote_code=True,
        token=HF_TOKEN or None,
    )
    model.eval()
    model.to(device)
    log("  BRIA RMBG-2.0 ready ✓")
    return model, device


def _bria_remove_bg(model, device, img_path: str) -> bytes:
    """
    Run BRIA RMBG-2.0 on a single image file.
    Returns transparent PNG bytes (RGBA).
    """
    import torchvision.transforms as T

    transform = T.Compose([
        T.Resize((BRIA_INPUT_SIZE, BRIA_INPUT_SIZE)),
        T.ToTensor(),
        T.Normalize(mean=[0.485, 0.456, 0.406],
                    std=[0.229, 0.224, 0.225]),
    ])

    # ── Load & convert ────────────────────────────────────────────
    orig = Image.open(img_path).convert("RGB")
    orig_w, orig_h = orig.size

    inp = transform(orig).unsqueeze(0).to(device)

    # ── Inference ─────────────────────────────────────────────────
    with torch.no_grad():
        preds = model(inp)
        # model returns list; last element is the finest-grained mask
        pred_tensor = preds[-1].sigmoid().squeeze().cpu()

    # ── Build alpha mask at original resolution ───────────────────
    mask_pil = T.ToPILImage()(pred_tensor).resize(
        (orig_w, orig_h), Image.LANCZOS
    )

    # ── Composite: paste subject onto transparent canvas ──────────
    orig_rgba = orig.convert("RGBA")
    orig_rgba.putalpha(mask_pil)

    buf = io.BytesIO()
    orig_rgba.save(buf, "PNG")
    return buf.getvalue()


def phase3_remove_bg(items):
    log_section(f"PHASE 3: BRIA RMBG-2.0 SOTA  [{REMBG_MODEL}]  (direct HuggingFace)")
    gpu_info()

    try:
        model, device = _bria_load_model()
    except Exception as e:
        log(f"  BRIA load error: {e}  — falling back to original image (no bg removal)")
        for item in items:
            item["transparent_path"] = item["local_path"]
        return items

    result = []
    for i, item in enumerate(items):
        try:
            log(f"  [{i+1:>3}/{len(items)}] {item['name']}")
            out_bytes = _bria_remove_bg(model, device, item["local_path"])
            dst = TRANSPARENT_DIR / f"{item['stem']}.png"
            dst.write_bytes(out_bytes)
            item["transparent_path"] = str(dst)    # -> phase4
            chk  = Image.open(io.BytesIO(out_bytes))
            w, h = chk.size
            chk.close()
            log(f"    -> {w}x{h}  {len(out_bytes)//1024} KB  (transparent PNG)")
            result.append(item)
        except Exception as e:
            log(f"  FAIL BRIA [{item['name']}]: {e}")
            item["transparent_path"] = item["local_path"]  # fallback: original
            result.append(item)

    del model
    free_memory()
    log(f"  Done: {len(result)}/{len(items)}")
    return result

# ══════════════════════════════════════════════════════════════════
# PHASE 4: WEBP PREVIEWS  (<80 KB, checkered BG, watermark)
# ══════════════════════════════════════════════════════════════════
def phase4_make_webp_previews(items):
    log_section(f"PHASE 4: WebP previews  (<{WEBP_MAX_BYTES//1024} KB + watermark)")
    result = []
    for i, item in enumerate(items):
        try:
            img_rgba           = Image.open(item["transparent_path"]).convert("RGBA")
            webp_bytes, pw, ph = make_webp_preview(img_rgba)
            img_rgba.close()

            out_path = WEBP_DIR / f"{item['stem']}.webp"
            out_path.write_bytes(webp_bytes)

            item["webp_bytes"] = webp_bytes  # -> phase6 GitHub upload
            item["webp_path"]  = str(out_path)
            item["preview_w"]  = pw          # -> phase8 ultradata
            item["preview_h"]  = ph

            log(f"  [{i+1:>3}/{len(items)}] {item['stem']}.webp  "
                f"{len(webp_bytes)//1024} KB  ({pw}x{ph})")
            result.append(item)
        except Exception as e:
            log(f"  FAIL webp [{item['name']}]: {e}")
    return result

# ══════════════════════════════════════════════════════════════════
# PHASE 5: DRIVE UPLOAD  (transparent PNG -> png_library_images/)
# ══════════════════════════════════════════════════════════════════
def phase5_drive_upload(items):
    log_section(f"PHASE 5: Drive upload — transparent PNG  "
                f"({len(items)} items, batch={BATCH_SIZE})")
    token    = get_drive_token()
    png_root = drive_folder_get_or_create(token, "png_library_images")
    log(f"  png_library_images -> {png_root}")

    _sub_cache = {}
    def get_sub(sf):
        if sf not in _sub_cache:
            tok = get_drive_token()
            _sub_cache[sf] = drive_folder_get_or_create(tok, sf, png_root)
            log(f"  Sub-folder '{sf}' -> {_sub_cache[sf]}")
        return _sub_cache[sf]

    uploaded  = []
    n_batches = math.ceil(len(items) / BATCH_SIZE)
    for b in range(n_batches):
        batch = items[b * BATCH_SIZE : (b + 1) * BATCH_SIZE]
        log(f"\n  -- Batch {b+1}/{n_batches}  ({len(batch)} items) --")
        for j, item in enumerate(batch):
            try:
                token     = get_drive_token()
                stem      = item["stem"]
                sf        = item["subfolder_name"]
                png_bytes = Path(item["transparent_path"]).read_bytes()
                res       = drive_upload(token, get_sub(sf),
                                         f"{stem}.png", png_bytes, "image/png")
                drive_share(token, res["id"])
                item["png_drive_id"]     = res["id"]
                # Permanent direct-download URL (no login needed)
                item["png_download_url"] = (
                    f"https://drive.google.com/uc?export=download&id={res['id']}"
                )
                uploaded.append(item)
                log(f"    [{j+1:>2}/{len(batch)}] OK  {stem}.png -> Drive  id={res['id']}")
            except Exception as e:
                log(f"    FAIL Drive [{item['name']}]: {e}")

    log(f"\n  Drive uploaded: {len(uploaded)}/{len(items)}")
    return uploaded

# ══════════════════════════════════════════════════════════════════
# PHASE 6: GITHUB UPLOAD  (WebP -> guruimageusha/preview_webp/)
# ══════════════════════════════════════════════════════════════════
def phase6_github_upload(items):
    log_section(
        f"PHASE 6: GitHub upload — WebP  "
        f"-> {GH_OWNER}/{PREVIEW_REPO}/{PREVIEW_FOLDER}/"
    )
    if not GH_TOKEN or not GH_OWNER:
        log("  SKIP — GH_TOKEN or GH_OWNER not set")
        for item in items:
            item["preview_cdn_url"] = ""
        return items

    result = []
    for i, item in enumerate(items):
        try:
            stem    = item["stem"]
            gh_path = f"{PREVIEW_FOLDER}/{stem}.webp"
            log(f"  [{i+1:>3}/{len(items)}] {stem}.webp  "
                f"{len(item['webp_bytes'])//1024} KB  -> {PREVIEW_REPO}/{gh_path}")

            github_upload_file(
                token         = GH_TOKEN,
                owner         = GH_OWNER,
                repo          = PREVIEW_REPO,
                path          = gh_path,
                content_bytes = item["webp_bytes"],
                message       = f"preview: add {stem}.webp",
                branch        = PREVIEW_BRANCH,
            )
            cdn = jsdelivr_url(GH_OWNER, PREVIEW_REPO, PREVIEW_BRANCH, gh_path)
            item["preview_cdn_url"] = cdn       # -> phase8 ultradata (preview_url)
            item["webp_file_id"]    = gh_path   # -> phase8 ultradata (webp_file_id)
            log(f"    CDN: {cdn}")
            result.append(item)
            time.sleep(0.4)

        except Exception as e:
            log(f"  FAIL GitHub [{item['name']}]: {e}")
            item["preview_cdn_url"] = ""
            result.append(item)

    ok = len([x for x in result if x.get("preview_cdn_url")])
    log(f"  GitHub uploaded: {ok}/{len(items)}")
    return result

# ══════════════════════════════════════════════════════════════════
# PHASE 7: MOVE ORIGINALS  (pending/* -> pending/finished/*)
# ══════════════════════════════════════════════════════════════════
def phase7_move_originals(items, pending_id):
    log_section("PHASE 7: Move originals -> pending/finished/{subfolder}/")
    token       = get_drive_token()
    finished_id = drive_folder_get_or_create(token, "finished", pending_id)
    log(f"  finished/ -> {finished_id}")

    _fin_cache = {}
    def get_fin_sub(sf):
        if sf not in _fin_cache:
            tok = get_drive_token()
            _fin_cache[sf] = drive_folder_get_or_create(tok, sf, finished_id)
        return _fin_cache[sf]

    moved = 0
    for item in items:
        try:
            token = get_drive_token()
            sf    = item["subfolder_name"]
            drive_move(token, item["fid"], get_fin_sub(sf), item["subfolder_id"])
            log(f"  OK  {item['name']}  -> finished/{sf}/")
            moved += 1
        except Exception as e:
            log(f"  FAIL move [{item['name']}]: {e}")
    log(f"  Moved: {moved}/{len(items)}")

# ══════════════════════════════════════════════════════════════════
# PHASE 8: UPDATE ultradata.xlsx  (ultrapng repo)
# ══════════════════════════════════════════════════════════════════
def phase8_update_ultradata(items):
    log_section(
        f"PHASE 8: Update {ULTRADATA_FILE}  -> {GH_OWNER}/{ULTRADATA_REPO}"
    )
    if not GH_TOKEN or not GH_OWNER:
        log("  SKIP — GH_TOKEN or GH_OWNER not set")
        return

    import openpyxl

    url = (f"https://api.github.com/repos/{GH_OWNER}/{ULTRADATA_REPO}"
           f"/contents/{ULTRADATA_FILE}")
    r   = req.get(url, headers=_gh(GH_TOKEN),
                  params={"ref": ULTRADATA_BRANCH}, timeout=30)

    if r.status_code == 200:
        data     = r.json()
        file_sha = data.get("sha")
        wb       = openpyxl.load_workbook(
            io.BytesIO(base64.b64decode(data["content"].replace("\n", ""))))
        log(f"  Loaded existing {ULTRADATA_FILE}  (SHA: {file_sha[:8]}...)")
    else:
        log(f"  {ULTRADATA_FILE} not found — creating new workbook")
        wb       = openpyxl.Workbook()
        file_sha = None

    ws = wb.active

    # ── Column schema — matches ultradata format exactly ──────────
    HEADERS = [
        "date_added",    # today YYYY-MM-DD
        "subject_name",  # filename stem  (e.g. "red_apple_01")
        "category",      # Drive subfolder  (e.g. "fruits")
        "subcategory",   # second-level category — derived or blank
        "filename",      # stem + ".png"
        "png_file_id",   # Google Drive file ID of the original PNG
        "webp_file_id",  # GitHub path  e.g. preview_webp/red_apple_01.webp
        "download_url",  # permanent Drive download link (no login needed)
        "preview_url",   # jsDelivr CDN URL for the WebP preview
        "seo_status",    # default "pending" — update manually when done
    ]

    # Ensure header row exists; add any missing columns (non-destructive)
    if ws.max_row == 0 or ws.cell(row=1, column=1).value is None:
        ws.append(HEADERS)
        log("  Headers written (new sheet)")
    else:
        existing = [ws.cell(row=1, column=c).value
                    for c in range(1, ws.max_column + 1)]
        for col_name in HEADERS:
            if col_name not in existing:
                ws.cell(row=1, column=ws.max_column + 1, value=col_name)
                existing.append(col_name)
                log(f"  Added missing column: '{col_name}'")
        # Re-read to pick up any newly added columns
        HEADERS = [ws.cell(row=1, column=c).value
                   for c in range(1, ws.max_column + 1)]

    today = datetime.now().strftime("%Y-%m-%d")
    added = 0
    for item in items:
        stem = item.get("stem", "")

        # top_category  = first-level subfolder (e.g. "fruits")
        # subfolder_name = immediate parent     (e.g. "tropical" for nested)
        # Falls back gracefully if keys missing (older items without BFS metadata)
        top_cat = item.get("top_category", item.get("subfolder_name", ""))
        sub_cat = item.get("subfolder_name", "")

        # If top_cat == sub_cat the item is in a top-level folder (no nesting)
        if top_cat == sub_cat:
            sub_cat = ""

        row = {
            "date_added":   today,
            "subject_name": stem,                               # e.g. red_apple_01
            "category":     top_cat,                           # top-level Drive folder
            "subcategory":  sub_cat,                           # nested folder name
            "filename":     f"{stem}.png",                     # original PNG filename
            "png_file_id":  item.get("png_drive_id",     ""),  # Drive file ID  (phase5)
            "webp_file_id": item.get("webp_file_id",     ""),  # GH path        (phase6)
            "download_url": item.get("png_download_url", ""),  # Drive dl URL   (phase5)
            "preview_url":  item.get("preview_cdn_url",  ""),  # jsDelivr URL   (phase6)
            "seo_status":   "pending",                         # section2_seo fills this
        }
        ws.append([row.get(h, "") for h in HEADERS])
        added += 1

    buf = io.BytesIO()
    wb.save(buf)

    body = {
        "message": f"ultradata: +{added} images [{today}]",
        "content": base64.b64encode(buf.getvalue()).decode(),
        "branch":  ULTRADATA_BRANCH,
    }
    if file_sha:
        body["sha"] = file_sha

    r2 = req.put(url,
                 headers={**_gh(GH_TOKEN), "Content-Type": "application/json"},
                 json=body, timeout=90)
    r2.raise_for_status()
    log(f"  OK  {ULTRADATA_FILE} updated  (+{added} rows)")

# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════
def main():
    log_section("Pending Drive Pipeline V3.4 — START")
    log(f"  RUN_ITEMS_COUNT  : {RUN_ITEMS_COUNT}  (env: RUN_ITEMS_COUNT)")
    log(f"  BG MODEL         : {REMBG_MODEL}  (SOTA — direct HuggingFace / 1024px)")
    log(f"  WEBP PREVIEW     : <{WEBP_MAX_BYTES//1024} KB  max_side={WEBP_MAX_SIDE}px")
    log(f"  WATERMARK        : {WATERMARK_TEXT}")
    log(f"  PENDING FOLDER   : {PENDING_FOLDER_NAME}")
    log(f"  PREVIEW -> GH    : {GH_OWNER}/{PREVIEW_REPO}/{PREVIEW_FOLDER}/")
    log(f"  ULTRADATA -> GH  : {GH_OWNER}/{ULTRADATA_REPO}/{ULTRADATA_FILE}")
    log(f"  BATCH_SIZE       : {BATCH_SIZE}")
    gpu_info()

    t0 = time.time()

    items, pending_id = phase1_discovery()
    if not items:
        log("No items found in pending folder. Done.")
        return

    items = phase2_download(items)
    if not items:
        log("Download failed for all items. Done.")
        return

    items = phase3_remove_bg(items)

    items = phase4_make_webp_previews(items)
    if not items:
        log("No items survived WebP preview generation. Done.")
        return

    uploaded = phase5_drive_upload(items)
    if not uploaded:
        log("No items uploaded to Drive. Done.")
        return

    uploaded = phase6_github_upload(uploaded)
    phase7_move_originals(uploaded, pending_id)
    phase8_update_ultradata(uploaded)

    elapsed = time.time() - t0
    log_section("DONE")
    log(f"  Processed  : {len(uploaded)}/{len(items)} images")
    log(f"  Time       : {elapsed/60:.1f} min  "
        f"({elapsed/max(len(uploaded),1):.0f} sec/image)")


if __name__ == "__main__":
    main()
